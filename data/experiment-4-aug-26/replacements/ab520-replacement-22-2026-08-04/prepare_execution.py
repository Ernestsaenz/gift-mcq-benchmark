"""Create frozen, harness-ready execution inputs after the QA gate passes."""

from __future__ import annotations

import csv
import hashlib
import json
import sys
from collections import Counter
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HERE = Path(__file__).resolve().parent
REPO = HERE.parents[3]
sys.path.insert(0, str(REPO / "code"))

from medrag_eval.excel_io import import_questions_from_workbook  # noqa: E402
from medrag_eval.prompting import render_benchmark_prompt  # noqa: E402


QA_SUMMARY = HERE / "protocol-qa/final-review-summary.json"
QA_COVERAGE = HERE / "protocol-qa/qa-coverage.csv"
RUN_MATRIX = HERE / "run-matrix-264.csv"
INPUTS = HERE / "inputs"
RUNS = HERE / "runs"
LOGS = HERE / "logs"
MANIFESTS = HERE / "manifests"
EXPORTS = HERE / "exports"
PRESENTATION = HERE / "presentation"
DB_PATH = RUNS / "ab520-replacement22-2026-08-05.sqlite"
LEDGER_PATH = MANIFESTS / "frozen-replacement-cell-ledger.csv"
NONE_TEXT = "Ninguna de las respuestas anteriores es correcta."
RAW_FIELDS = (
    "question_text",
    "option_a",
    "option_b",
    "option_c",
    "option_d",
    "correct_letter",
    "correct_option_text",
)
NUMERIC_COLUMNS = {"year", "question_number", "page_in_exam_pdf"}
BOOLEAN_COLUMNS = {"negated_stem"}
MODEL_ORDER = {
    "google/gemini-3.6-flash": 0,
    "google/gemma-4-26b-a4b-it": 1,
    "qwen/qwen3.6-35b-a3b": 2,
    "z-ai/glm-5.2": 3,
}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def sha256_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        raise AssertionError(f"Refusing to write empty CSV: {path}")
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


def lp_hash(row: dict[str, Any]) -> str:
    payload = bytearray()
    for field in RAW_FIELDS:
        raw = str(row[field]).encode("utf-8")
        payload.extend(len(raw).to_bytes(8, "big"))
        payload.extend(raw)
    return hashlib.sha256(payload).hexdigest()


def typed_value(column: str, value: str) -> Any:
    if value == "":
        return None
    if column in NUMERIC_COLUMNS:
        try:
            return int(value)
        except ValueError:
            return value
    if column in BOOLEAN_COLUMNS:
        return value.casefold() in {"true", "1", "yes"}
    return value


def write_workbook(csv_path: Path, xlsx_path: Path) -> dict[str, Any]:
    with csv_path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        columns = list(reader.fieldnames or [])
        rows = list(reader)
    assert len(rows) == 22
    assert len({row["question_id"] for row in rows}) == 22

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "questions"
    workbook.properties.title = "August replacement cohort execution input"
    workbook.properties.subject = "22 protocol-approved Spanish digestive-medicine MCQs"
    workbook.properties.creator = "ab520 replacement QA workflow"
    sheet.append(columns)
    for row in rows:
        sheet.append([typed_value(column, row[column]) for column in columns])

    header_fill = PatternFill("solid", fgColor="17365D")
    for cell in sheet[1]:
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 30
    for index, column in enumerate(columns, start=1):
        width = 14
        if column == "question_text":
            width = 60
        elif column.startswith("option_") or column == "correct_option_text":
            width = 48
        elif column in {"source_exam_pdf", "source_answer_key_pdf", "source_key"}:
            width = 42
        elif column in {"candidate_id", "question_id", "correct_letter"}:
            width = 14
        sheet.column_dimensions[get_column_letter(index)].width = width
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(xlsx_path)

    reopened = load_workbook(xlsx_path, read_only=False, data_only=False)
    assert reopened.sheetnames == ["questions"]
    check = reopened["questions"]
    assert check.max_row == 23 and check.max_column == len(columns)
    assert [cell.value for cell in check[1]] == columns
    assert all(
        cell.font.name == "Arial"
        for row in check.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert not any(
        cell.data_type == "f" or cell.data_type == "e"
        for row in check.iter_rows()
        for cell in row
    )
    imported = import_questions_from_workbook(xlsx_path)
    assert len(imported.questions) == 22 and not imported.warnings
    imported_by_id = {question.question_id: question for question in imported.questions}
    for row in rows:
        question = imported_by_id[row["question_id"]]
        assert question.question_text == row["question_text"]
        assert question.correct_letter == row["correct_letter"]
        assert question.correct_option_text == row["correct_option_text"]
        assert all(question.options[letter] == row[f"option_{letter}"] for letter in "abcd")
    return {
        "path": str(xlsx_path.relative_to(HERE)),
        "sha256": sha256_file(xlsx_path),
        "sheet": "questions",
        "rows": 22,
        "columns": len(columns),
        "font": "Arial",
        "formulas": 0,
        "excel_error_cells": 0,
        "harness_imported_rows": 22,
    }


def main() -> None:
    qa = json.loads(QA_SUMMARY.read_text(encoding="utf-8"))
    assert qa["status"] == "PASS_QA_COMPLETE_READY_FOR_EXECUTION"
    assert qa["candidate_count"] == 22
    assert qa["formal_sourcing_passes"] == 22
    assert qa["blinded_qa1_passes"] == qa["blinded_qa2_passes"] == 22
    coverage = read_csv(QA_COVERAGE)
    assert len(coverage) == 22 and {row["protocol_gate"] for row in coverage} == {"PASS"}

    for directory in (INPUTS, RUNS, LOGS, MANIFESTS, EXPORTS, PRESENTATION):
        directory.mkdir(parents=True, exist_ok=True)

    csv_a = HERE / "replacement-22-A.csv"
    csv_b = HERE / "replacement-22-B.csv"
    xlsx_a = INPUTS / "replacement-22-A.xlsx"
    xlsx_b = INPUTS / "replacement-22-B.xlsx"
    workbooks = {
        "A": write_workbook(csv_a, xlsx_a),
        "B": write_workbook(csv_b, xlsx_b),
    }

    rows_by_condition = {
        "A": {row["question_id"]: row for row in read_csv(csv_a)},
        "B": {row["question_id"]: row for row in read_csv(csv_b)},
    }
    matrix = read_csv(RUN_MATRIX)
    assert len(matrix) == 264 and {row["status"] for row in matrix} == {"READY_NOT_RUN"}
    coverage_by_replacement = {row["replacement_id"]: row for row in coverage}
    ledger: list[dict[str, Any]] = []
    for row in sorted(
        matrix,
        key=lambda item: (
            ("openrouter_A", "openrouter_B", "tailscale_A").index(item["arm"]),
            int(item["replacement_id"][1:]),
            MODEL_ORDER[item["model"]],
        ),
    ):
        condition = row["condition"]
        source = rows_by_condition[condition][row["replacement_id"]]
        provider = row["provider"]
        prompt = render_benchmark_prompt(
            source,
            provider=provider,
            prompt_version=row["prompt_version"],
        )
        coverage_row = coverage_by_replacement[row["replacement_id"]]
        assert coverage_row["candidate_id"] == row["candidate_id"] == source["candidate_id"]
        assert coverage_row["source_key"] == source["source_key"]
        ledger.append(
            {
                "cell_key": (
                    f"{provider}|{condition}|{row['replacement_id']}|{source['source_key']}|"
                    f"{row['model']}|1"
                ),
                "arm": row["arm"],
                "provider": provider,
                "condition": condition,
                "replacement_id": row["replacement_id"],
                "replaces_question_id": row["replaces_question_id"],
                "candidate_id": row["candidate_id"],
                "source_key": source["source_key"],
                "model": row["model"],
                "run_index": 1,
                "dataset": row["dataset_name"],
                "experiment": row["planned_experiment_id"],
                "prompt_version": row["prompt_version"],
                "temperature": row["temperature"],
                "tailscale_prompt_id": row["tailscale_prompt_id"],
                "input_fields_sha256": lp_hash(source),
                "system_prompt_sha256": prompt.system_sha256,
                "user_prompt_sha256": prompt.user_sha256,
                "user_prompt_characters": len(prompt.user_prompt),
                "qa_coverage_sha256": sha256_file(QA_COVERAGE),
                "status": "READY_NOT_RUN",
            }
        )
    assert len(ledger) == len({row["cell_key"] for row in ledger}) == 264
    assert Counter(row["arm"] for row in ledger) == Counter(
        {"openrouter_A": 88, "openrouter_B": 88, "tailscale_A": 88}
    )
    write_csv(LEDGER_PATH, ledger)

    generated_at = datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    manifest = {
        "artifact_version": "ab520-replacement22-execution-input-v1",
        "generated_at_utc": generated_at,
        "status": "READY_NOT_RUN",
        "qa_summary": {
            "path": str(QA_SUMMARY.relative_to(HERE)),
            "sha256": sha256_file(QA_SUMMARY),
        },
        "qa_coverage": {
            "path": str(QA_COVERAGE.relative_to(HERE)),
            "sha256": sha256_file(QA_COVERAGE),
        },
        "source_csvs": {
            "A": {"path": csv_a.name, "sha256": sha256_file(csv_a)},
            "B": {"path": csv_b.name, "sha256": sha256_file(csv_b)},
        },
        "workbooks": workbooks,
        "database": str(DB_PATH.relative_to(HERE)),
        "cell_ledger": {
            "path": str(LEDGER_PATH.relative_to(HERE)),
            "sha256": sha256_file(LEDGER_PATH),
            "cells": 264,
            "cells_per_arm": 88,
        },
        "datasets": {
            "aug26_replacement22_A": workbooks["A"]["path"],
            "aug26_replacement22_B": workbooks["B"]["path"],
        },
        "experiments": sorted({row["experiment"] for row in ledger}),
        "models": sorted(MODEL_ORDER, key=MODEL_ORDER.get),
        "prompt_version": "mcq_es_v4",
        "temperature": 0,
        "runs": 1,
        "tailscale_prompt_id": 13,
        "excluded_arm": "tailscale_B",
    }
    (MANIFESTS / "execution-input-manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    ledger_header = (
        "invocation_id,arm,condition,model,dataset,experiment_id,target_count,concurrency,"
        "start_time_utc,end_time_utc,status,scored_count,failure_count,redacted_command,"
        "log_path,notes\n"
    )
    run_ledger = HERE / "RUN_LEDGER.csv"
    if not run_ledger.exists():
        run_ledger.write_text(ledger_header, encoding="utf-8")
    status_path = HERE / "STATUS.md"
    if not status_path.exists():
        status_path.write_text(
            "# Replacement execution status\n\n"
            f"Updated: {generated_at}\n\n"
            "QA gate: **PASS — 22/22 sourcing, 22/22 QA1, 22/22 QA2**.\n\n"
            "| Arm | Retained scored | Replacement scored | Queued | Adjusted required |\n"
            "|---|---:|---:|---:|---:|\n"
            "| OpenRouter A | 1,912 | 0 | 88 | 2,000 |\n"
            "| OpenRouter B | 1,912 | 0 | 88 | 2,000 |\n"
            "| GIFT/TailScale A | 1,912 | 0 | 88 | 2,000 |\n\n"
            "Active concurrency: 0. Provider calls issued for this cohort: 0.\n",
            encoding="utf-8",
        )
    print(json.dumps(manifest, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
