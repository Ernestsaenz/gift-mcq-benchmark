"""Fail-closed finalizer for the 22-question replacement QA gate.

The script performs only local validation.  It does not contact either model
provider.  It writes final QA summaries only after every protocol invariant has
passed, and then promotes the already-frozen execution matrix from QA-blocked to
ready-not-run.
"""

from __future__ import annotations

import csv
import hashlib
import json
import re
import unicodedata
from collections import Counter
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


HERE = Path(__file__).resolve().parent
REPO = HERE.parents[3]
QA_ROOT = HERE / "protocol-qa"
SOURCE_WORKBOOK = Path(
    "/Users/ernestsaenz/Programming/gift-project-compile/second-project/"
    "workbook-repairs-2026-07-30/outputs/all-regions-aparato-digestivo.corrected.xlsx"
)
CANONICAL_BENCHMARK = REPO / "data/experiment-4-aug-26/benchmark-500.json"
CANONICAL_RESULTS = (
    REPO
    / "data/experiment-4-aug-26/results/ab520-gapfill-2026-08-04/"
    "exports/benchmark-6000-cell-results.csv"
)
PROTOCOL_VERSION = "ab182-readonly-v1"
AS_OF_DATE = "2026-08-02"
CORPUS_SHA256 = "18f6becd4e51f1b9ef6a5a8ab68421e905cfe2584ec32a0e303b76f3cacf1e46"
CANONICAL_BENCHMARK_SHA256 = "057f8b805d928e90079b8ba80b326581e6e14f65fb9a1644a0b4d53cbc294abc"
NONE_TEXT = "Ninguna de las respuestas anteriores es correcta."
RAW_FIELDS = (
    "question_text",
    "option_a",
    "option_b",
    "option_c",
    "option_d",
    "correct_letter",
    "correct_option_text",
)
MODELS = (
    "google/gemini-3.6-flash",
    "google/gemma-4-26b-a4b-it",
    "qwen/qwen3.6-35b-a3b",
    "z-ai/glm-5.2",
)
ARMS = ("openrouter_A", "openrouter_B", "tailscale_A")
CALIBRATION_EXPECTED = {
    "b2": "PASS",
    "b193": "PASS",
    "b211": "PASS",
    "b480": "PASS",
    "b10": "REJECT",
    "b152": "REJECT",
    "b178": "REJECT",
    "b205": "REJECT",
}
REQUIRED_EVIDENCE = {"official_exam", "definitive_key", "medical_guidance"}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def text(value: Any) -> str:
    return "" if value is None else str(value)


def normalized(value: Any) -> str:
    value = unicodedata.normalize("NFKC", text(value))
    return re.sub(r"\s+", " ", value).strip().casefold()


def lp_serialize(values: Iterable[Any]) -> bytes:
    payload = bytearray()
    for value in values:
        raw = text(value).encode("utf-8")
        payload.extend(len(raw).to_bytes(8, "big"))
        payload.extend(raw)
    return bytes(payload)


def field_hash(value: Any) -> str:
    return sha256_bytes(lp_serialize([value]))


def fields_hash(row: dict[str, Any], *, normalize: bool = False) -> str:
    values = [row.get(field) for field in RAW_FIELDS]
    if normalize:
        values = [normalized(value) for value in values]
    return sha256_bytes(lp_serialize(values))


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    return [
        json.loads(line)
        for line in path.read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def write_json(path: Path, value: Any) -> None:
    path.write_text(
        json.dumps(value, ensure_ascii=False, indent=2, sort_keys=False) + "\n",
        encoding="utf-8",
    )


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        raise AssertionError(f"Refusing to write empty CSV: {path}")
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


def calibration_score(
    path: Path, reviewer_id: str, *, expected_blinded: bool = True
) -> dict[str, Any]:
    rows = read_jsonl(path)
    assert len(rows) == 8
    assert len({row["candidate_id"] for row in rows}) == 8
    assert {row["candidate_id"] for row in rows} == set(CALIBRATION_EXPECTED)
    assert {row.get("reviewer_id") for row in rows} == {reviewer_id}
    assert {row.get("blinded") for row in rows} == {expected_blinded}
    actual = {row["candidate_id"]: row["verdict"] for row in rows}
    score = sum(actual[item] == expected for item, expected in CALIBRATION_EXPECTED.items())
    return {
        "reviewer_id": reviewer_id,
        "path": str(path.relative_to(HERE)),
        "sha256": sha256_file(path),
        "score": score,
        "required_score": 8,
        "eligible": score == 8,
        "verdict_counts": dict(sorted(Counter(actual.values()).items())),
    }


def validate_review(
    review: dict[str, Any],
    packet: dict[str, Any],
    *,
    expected_reviewer: str | None,
    expected_role: str | None,
    expected_blinded: bool | None,
) -> None:
    candidate_id = packet["candidate_id"]
    assert "proposed_fix" not in review, candidate_id
    assert review.get("protocol_version") == PROTOCOL_VERSION, candidate_id
    assert review.get("candidate_id") == candidate_id, candidate_id
    assert review.get("source_key") == packet["source_key"], candidate_id
    assert review.get("corpus_sha256") == packet["corpus_sha256"] == CORPUS_SHA256
    assert review.get("raw_fields_hash") == packet["raw_fields_hash"], candidate_id
    if expected_reviewer is not None:
        assert review.get("reviewer_id") == expected_reviewer, candidate_id
    else:
        assert isinstance(review.get("reviewer_id"), str) and review["reviewer_id"]
    if expected_blinded is None:
        assert isinstance(review.get("blinded"), bool), candidate_id
    else:
        assert review.get("blinded") is expected_blinded, candidate_id
    if expected_role is not None:
        assert review.get("qa_role") == expected_role, candidate_id
    assert review.get("verdict") == "PASS", candidate_id
    assert review.get("a_validity") == "PASS", candidate_id
    assert review.get("b_validity") == "PASS", candidate_id
    assert isinstance(review.get("rationale"), str) and review["rationale"].strip()
    # Completed upstream records are immutable.  The original coordinator
    # accepted an empty reason-code list, so preserve that valid historical
    # representation while requiring the field and string values when present.
    assert isinstance(review.get("reason_codes"), list)
    assert all(isinstance(value, str) and value for value in review["reason_codes"])
    assert isinstance(review.get("model"), str) and review["model"].strip()

    evidence = review.get("evidence")
    assert isinstance(evidence, list) and evidence, candidate_id
    assert REQUIRED_EVIDENCE <= {item.get("kind") for item in evidence}, candidate_id
    provenance = packet["provenance"]
    exam_matches = []
    key_matches = []
    for item in evidence:
        assert isinstance(item, dict), candidate_id
        assert isinstance(item.get("kind"), str) and item["kind"]
        locator = item.get("locator")
        assert isinstance(locator, str) and locator.strip()
        if item["kind"] == "official_exam":
            exam_matches.append(
                item.get("sha256") == provenance["exam_pdf_sha256"]
                and (
                    provenance["exam_pdf_path"] in locator
                    or item.get("path") == provenance["exam_pdf_path"]
                )
            )
        elif item["kind"] == "definitive_key":
            key_matches.append(
                item.get("sha256") == provenance["key_pdf_sha256"]
                and (
                    provenance["key_pdf_path"] in locator
                    or item.get("path") == provenance["key_pdf_path"]
                )
            )
        elif item["kind"] == "medical_guidance":
            guidance_locator = item.get("url") or locator
            assert isinstance(guidance_locator, str) and guidance_locator.startswith(
                ("https://", "http://", "/")
            ), (candidate_id, guidance_locator)
            if guidance_locator.startswith(("https://", "http://")):
                assert (item.get("accessed_date") or item.get("accessed")) == AS_OF_DATE
    assert any(exam_matches), candidate_id
    assert any(key_matches), candidate_id


def validate_batch(
    *,
    role: str,
    batch: int,
    reviewer_id: str,
    qa_role: str,
    packets: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    assignment = read_jsonl(QA_ROOT / "inputs" / f"{role}-batch-{batch:02d}.jsonl")
    reviews = read_jsonl(QA_ROOT / "reviews" / role / f"batch-{batch:02d}.jsonl")
    assigned = {row["candidate_id"]: row for row in assignment}
    assert len(assigned) == len(assignment) <= 10
    assert len(reviews) == len(assignment)
    assert len({row["candidate_id"] for row in reviews}) == len(reviews)
    assert {row["candidate_id"] for row in reviews} == set(assigned)
    for review in reviews:
        packet = packets[review["candidate_id"]]
        assert packet["raw_fields"] == assigned[review["candidate_id"]]["raw_fields"]
        validate_review(
            review,
            packet,
            expected_reviewer=reviewer_id,
            expected_role=qa_role,
            expected_blinded=True,
        )
    return reviews


def validate_workbook_and_packets(
    packets: dict[str, dict[str, Any]],
) -> tuple[dict[str, str], int]:
    assert sha256_file(SOURCE_WORKBOOK) == CORPUS_SHA256
    workbook = load_workbook(SOURCE_WORKBOOK, read_only=True, data_only=True)
    sheet = workbook["questions"]
    rows = sheet.iter_rows(values_only=True)
    headers = [text(value) for value in next(rows)]
    required_rows = {int(packet["source_row"]): packet for packet in packets.values()}
    workbook_rows: dict[int, dict[str, Any]] = {}
    for row_number, values in enumerate(rows, start=2):
        if row_number in required_rows:
            workbook_rows[row_number] = dict(zip(headers, values))
    assert set(workbook_rows) == set(required_rows)

    pdf_hashes: dict[str, str] = {}
    for source_row, packet in required_rows.items():
        row = workbook_rows[source_row]
        raw = {field: text(row.get(field)) for field in RAW_FIELDS}
        assert raw == packet["raw_fields"], packet["candidate_id"]
        assert fields_hash(row) == packet["raw_fields_hash"], packet["candidate_id"]
        assert {
            field: field_hash(raw[field]) for field in RAW_FIELDS
        } == packet["raw_field_hashes"], packet["candidate_id"]
        assert text(row.get("region")) == packet["region"]
        assert int(row["year"]) == int(packet["year"])
        assert text(row.get("exam_part")) == packet["exam_part"]
        assert int(row["question_number"]) == int(packet["question_number"])
        assert not packet.get("context_chunks"), packet["candidate_id"]
        assert field_hash(raw["question_text"]) == packet["assembled_context_and_stem_sha256"]

        letter = raw["correct_letter"].casefold()
        assert letter in "bcd"
        assert raw["correct_option_text"] == raw[f"option_{letter}"]
        b_fields = dict(raw)
        b_fields[f"option_{letter}"] = NONE_TEXT
        b_fields["correct_option_text"] = NONE_TEXT
        changed = [field for field in RAW_FIELDS if raw[field] != b_fields[field]]
        assert changed == [f"option_{letter}", "correct_option_text"]
        expected_b = {
            "replacement_text": NONE_TEXT,
            "changed_fields": changed,
            "only_authorized_two_field_diff": True,
            "b_fields_hash": sha256_bytes(lp_serialize(b_fields[field] for field in RAW_FIELDS)),
        }
        assert packet["b_simulation"] == expected_b, packet["candidate_id"]

        provenance = packet["provenance"]
        for path_key, hash_key in (
            ("exam_pdf_path", "exam_pdf_sha256"),
            ("key_pdf_path", "key_pdf_sha256"),
        ):
            path = Path(provenance[path_key])
            assert path.is_file(), path
            digest = pdf_hashes.setdefault(str(path), sha256_file(path))
            assert digest == provenance[hash_key], path
    return pdf_hashes, len(workbook_rows)


def validate_forms_and_matrix(
    packets: dict[str, dict[str, Any]], mappings: list[dict[str, Any]]
) -> dict[str, Any]:
    rows_a = read_csv(HERE / "replacement-22-A.csv")
    rows_b = read_csv(HERE / "replacement-22-B.csv")
    by_a = {row["question_id"]: row for row in rows_a}
    by_b = {row["question_id"]: row for row in rows_b}
    expected_ids = {mapping["replacement_id"] for mapping in mappings}
    assert len(rows_a) == len(by_a) == len(rows_b) == len(by_b) == 22
    assert set(by_a) == set(by_b) == expected_ids
    mapping_by_replacement = {row["replacement_id"]: row for row in mappings}
    max_gift = 0
    for replacement_id in sorted(expected_ids):
        mapping = mapping_by_replacement[replacement_id]
        packet = packets[mapping["candidate_id"]]
        a = by_a[replacement_id]
        b = by_b[replacement_id]
        letter = packet["raw_fields"]["correct_letter"]
        for field in RAW_FIELDS:
            assert a[field] == packet["raw_fields"][field], (replacement_id, field)
        changed = [field for field in RAW_FIELDS if a[field] != b[field]]
        assert changed == [f"option_{letter}", "correct_option_text"]
        assert b[f"option_{letter}"] == b["correct_option_text"] == NONE_TEXT
        assert a["source_key"] == b["source_key"] == packet["source_key"]
        assert a["candidate_id"] == b["candidate_id"] == packet["candidate_id"]
        assert a["replaces_question_id"] == b["replaces_question_id"] == mapping["replaces_question_id"]
        for row in (a, b):
            user_prompt = (
                f"question_id: {replacement_id}\n\n{row['question_text']}\n\n"
                f"a) {row['option_a']}\n"
                f"b) {row['option_b']}\n"
                f"c) {row['option_c']}\n"
                f"d) {row['option_d']}"
            ).strip()
            max_gift = max(max_gift, len(user_prompt))
            assert len(user_prompt) <= 4500

    matrix = read_csv(HERE / "run-matrix-264.csv")
    assert len(matrix) == 22 * 4 * 3 == 264
    identities = {
        (row["arm"], row["replacement_id"], row["model"]) for row in matrix
    }
    assert len(identities) == len(matrix)
    assert {row["arm"] for row in matrix} == set(ARMS)
    assert {row["model"] for row in matrix} == set(MODELS)
    for arm in ARMS:
        arm_rows = [row for row in matrix if row["arm"] == arm]
        assert len(arm_rows) == 88
        assert {row["replacement_id"] for row in arm_rows} == expected_ids
        assert Counter(row["model"] for row in arm_rows) == Counter({model: 22 for model in MODELS})
    return {"rows_A": 22, "rows_B": 22, "matrix_cells": 264, "max_gift_characters": max_gift}


def validate_cohort_and_adjusted_baseline(
    mappings: list[dict[str, Any]], packets: dict[str, dict[str, Any]]
) -> dict[str, Any]:
    assert sha256_file(CANONICAL_BENCHMARK) == CANONICAL_BENCHMARK_SHA256
    canonical = read_json(CANONICAL_BENCHMARK)
    proposed = read_json(HERE / "benchmark-500-with-provisional-replacements.json")
    assert len(canonical) == len(proposed) == 500
    assert len({row["question_id"] for row in proposed}) == 500
    old_ids = {row["replaces_question_id"] for row in mappings}
    new_ids = {row["replacement_id"] for row in mappings}
    assert not old_ids & {row["question_id"] for row in proposed}
    assert new_ids <= {row["question_id"] for row in proposed}
    mapping_by_old = {row["replaces_question_id"]: row for row in mappings}
    retained_count = 0
    for before, after in zip(canonical, proposed):
        mapping = mapping_by_old.get(before["question_id"])
        if mapping is None:
            assert before == after
            retained_count += 1
        else:
            assert after["question_id"] == mapping["replacement_id"]
            assert after["candidate_id"] == mapping["candidate_id"]
    assert retained_count == 478

    retained = [row for row in canonical if row["question_id"] not in old_ids]
    retained_source_keys = {row["source_key"] for row in retained}
    retained_stems = {normalized(row["stem"]) for row in retained}
    selected_source_keys = {packet["source_key"] for packet in packets.values()}
    selected_stems = {normalized(packet["raw_fields"]["question_text"]) for packet in packets.values()}
    assert len(selected_source_keys) == len(selected_stems) == 22
    assert not selected_source_keys & retained_source_keys
    assert not selected_stems & retained_stems
    assert len({packet["raw_fields_hash"] for packet in packets.values()}) == 22

    old_letters = Counter(
        row["correct_letter"] for row in canonical if row["question_id"] in old_ids
    )
    new_letters = Counter(packet["raw_fields"]["correct_letter"] for packet in packets.values())
    assert old_letters == new_letters == Counter({"b": 7, "c": 7, "d": 8})

    cell_rows = read_csv(CANONICAL_RESULTS)
    assert len(cell_rows) == 6000
    retained_cells = [row for row in cell_rows if row["question_id"] not in old_ids]
    removed_cells = [row for row in cell_rows if row["question_id"] in old_ids]
    assert len(retained_cells) == 478 * 4 * 3 == 5736
    assert len(removed_cells) == 22 * 4 * 3 == 264
    assert all(row["final_execution_status"] == "scored" for row in retained_cells)
    assert Counter(row["arm"] for row in retained_cells) == Counter({arm: 1912 for arm in ARMS})
    assert sum(row["final_execution_status"] != "scored" for row in removed_cells) == 70
    return {
        "canonical_questions": 500,
        "retained_questions": 478,
        "replacement_questions": 22,
        "retained_scored_cells": 5736,
        "retained_scored_cells_per_arm": 1912,
        "removed_original_cells": 264,
        "removed_original_unresolved_cells": 70,
        "target_replacement_cells_per_arm": 88,
        "target_adjusted_cells_per_arm": 2000,
    }


def validate_c0369_duplicate_adjudication(
    packets: dict[str, dict[str, Any]], canonical: list[dict[str, Any]]
) -> dict[str, Any]:
    packet = packets["c0369"]
    workbook = load_workbook(SOURCE_WORKBOOK, read_only=True, data_only=True)
    sheet = workbook["questions"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    historical_values = next(
        sheet.iter_rows(min_row=1370, max_row=1370, values_only=True)
    )
    historical = dict(zip(headers, historical_values))
    assert historical["region"] == "castilla-y-leon"
    assert int(historical["year"]) == 2022
    assert historical["exam_part"] == "reserva"
    assert int(historical["question_number"]) == 95
    assert text(historical["question_text"]) == packet["raw_fields"]["question_text"]
    assert text(historical["correct_letter"]) == packet["raw_fields"]["correct_letter"] == "b"
    assert text(historical["correct_option_text"]) == packet["raw_fields"]["correct_option_text"]
    assert historical["option_d"] is None
    assert text(historical["option_a"]) != packet["raw_fields"]["option_a"]
    assert "castilla-y-leon|2022|reserva|95" not in {
        row["source_key"] for row in canonical
    }
    return {
        "candidate_id": "c0369",
        "historical_source_key": "castilla-y-leon|2022|reserva|95",
        "same_stem_and_keyed_answer": True,
        "historical_option_count": 3,
        "selected_option_count": 4,
        "different_distractor_set": True,
        "historical_item_in_active_benchmark": False,
        "full_content_duplicate": False,
        "adjudication": "PASS_NOT_AN_ACTIVE_OR_FULL_CONTENT_DUPLICATE",
    }


def promote_package(summary: dict[str, Any], coverage: list[dict[str, Any]]) -> None:
    manifest_path = HERE / "replacement-manifest.json"
    manifest = read_json(manifest_path)
    assert {row["candidate_id"] for row in manifest["replacements"]} == {
        row["candidate_id"] for row in coverage
    }
    manifest["status"] = "QA_COMPLETE_READY_FOR_EXECUTION"
    manifest["execution_status"] = "READY_NOT_RUN"
    manifest["protocol_qa"] = {
        "summary_path": "protocol-qa/final-review-summary.json",
        "coverage_path": "protocol-qa/qa-coverage.csv",
        "source_passes": 22,
        "qa1_passes": 22,
        "qa2_passes": 22,
        "all_candidates_ready": True,
        "validated_at_utc": summary["validated_at_utc"],
    }
    coverage_by_id = {row["candidate_id"]: row for row in coverage}
    for replacement in manifest["replacements"]:
        row = coverage_by_id[replacement["candidate_id"]]
        replacement["protocol_gate"] = {
            "status": "PASS",
            "sourcing_reviewer": row["sourcing_reviewer"],
            "qa1_reviewer": row["qa1_reviewer"],
            "qa2_reviewer": row["qa2_reviewer"],
        }
    write_json(manifest_path, manifest)

    spec_path = HERE / "selection-spec.json"
    spec = read_json(spec_path)
    spec["status"] = "QA_COMPLETE_READY_FOR_EXECUTION"
    write_json(spec_path, spec)

    benchmark_path = HERE / "benchmark-500-with-provisional-replacements.json"
    benchmark = read_json(benchmark_path)
    for row in benchmark:
        if row.get("origin") == "replacement22_2026-08-04":
            row["replacement_status"] = "QA_COMPLETE_READY_FOR_EXECUTION"
    write_json(benchmark_path, benchmark)

    matrix_path = HERE / "run-matrix-264.csv"
    matrix = read_csv(matrix_path)
    for row in matrix:
        row["status"] = "READY_NOT_RUN"
    write_csv(matrix_path, matrix)


def main() -> None:
    packets_list = read_jsonl(HERE / "selected-source-packets.jsonl")
    packets = {row["candidate_id"]: row for row in packets_list}
    spec = read_json(HERE / "selection-spec.json")
    mappings = spec["mappings"]
    assert len(packets) == len(mappings) == 22
    assert len({row["candidate_id"] for row in mappings}) == 22
    assert set(packets) == {row["candidate_id"] for row in mappings}

    calibrations = [
        calibration_score(
            QA_ROOT / "reviews/calibration/RSRC22.jsonl",
            "RSRC22",
            expected_blinded=False,
        ),
        calibration_score(QA_ROOT / "reviews/calibration/RQA22A.jsonl", "RQA22A"),
        calibration_score(QA_ROOT / "reviews/calibration/RQA22B.jsonl", "RQA22B"),
        calibration_score(
            QA_ROOT / "reviews/calibration/RQA22C-attempt2.jsonl", "RQA22C"
        ),
    ]
    assert all(row["eligible"] for row in calibrations)
    failed_calibration = calibration_score(
        QA_ROOT / "reviews/calibration/RQA22C.jsonl", "RQA22C"
    )
    assert failed_calibration["score"] == 7 and not failed_calibration["eligible"]

    prior_sourcing = read_jsonl(HERE / "selected-sourcing-reviews.jsonl")
    new_sourcing = read_jsonl(QA_ROOT / "reviews/sourcing/batch-01.jsonl")
    assert len(prior_sourcing) == 12 and len(new_sourcing) == 10
    assert not {row["candidate_id"] for row in prior_sourcing} & {
        row["candidate_id"] for row in new_sourcing
    }
    sourcing = prior_sourcing + new_sourcing
    assert len({row["candidate_id"] for row in sourcing}) == 22
    assert {row["candidate_id"] for row in sourcing} == set(packets)
    new_sourcing_ids = {row["candidate_id"] for row in new_sourcing}
    for review in sourcing:
        is_new_sourcing = review["candidate_id"] in new_sourcing_ids
        validate_review(
            review,
            packets[review["candidate_id"]],
            expected_reviewer=(
                "RSRC22" if is_new_sourcing else None
            ),
            expected_role=None,
            expected_blinded=False if is_new_sourcing else None,
        )

    qa1 = []
    for batch in (1, 2, 3):
        qa1.extend(
            validate_batch(
                role="qa1",
                batch=batch,
                reviewer_id="RQA22A",
                qa_role="QA1",
                packets=packets,
            )
        )
    qa2 = validate_batch(
        role="qa2",
        batch=1,
        reviewer_id="RQA22B",
        qa_role="QA2",
        packets=packets,
    )
    for batch in (2, 3):
        qa2.extend(
            validate_batch(
                role="qa2",
                batch=batch,
                reviewer_id="RQA22C",
                qa_role="QA2",
                packets=packets,
            )
        )
    assert len(qa1) == len(qa2) == 22
    assert {row["candidate_id"] for row in qa1} == set(packets)
    assert {row["candidate_id"] for row in qa2} == set(packets)
    qa1_by_id = {row["candidate_id"]: row for row in qa1}
    qa2_by_id = {row["candidate_id"]: row for row in qa2}
    source_by_id = {row["candidate_id"]: row for row in sourcing}

    pdf_hashes, workbook_rows = validate_workbook_and_packets(packets)
    forms = validate_forms_and_matrix(packets, mappings)
    cohort = validate_cohort_and_adjusted_baseline(mappings, packets)
    canonical = read_json(CANONICAL_BENCHMARK)
    duplicate = validate_c0369_duplicate_adjudication(packets, canonical)

    coverage = []
    mapping_by_candidate = {row["candidate_id"]: row for row in mappings}
    for candidate_id in sorted(packets):
        source = source_by_id[candidate_id]
        first = qa1_by_id[candidate_id]
        second = qa2_by_id[candidate_id]
        assert first["reviewer_id"] != second["reviewer_id"]
        mapping = mapping_by_candidate[candidate_id]
        coverage.append(
            {
                "replacement_id": mapping["replacement_id"],
                "replaces_question_id": mapping["replaces_question_id"],
                "candidate_id": candidate_id,
                "source_key": packets[candidate_id]["source_key"],
                "raw_fields_hash": packets[candidate_id]["raw_fields_hash"],
                "sourcing_reviewer": source["reviewer_id"],
                "sourcing_verdict": source["verdict"],
                "qa1_reviewer": first["reviewer_id"],
                "qa1_verdict": first["verdict"],
                "qa2_reviewer": second["reviewer_id"],
                "qa2_verdict": second["verdict"],
                "distinct_qa_reviewers": "TRUE",
                "protocol_gate": "PASS",
            }
        )

    validated_at = datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    summary = {
        "protocol_version": PROTOCOL_VERSION,
        "validated_at_utc": validated_at,
        "status": "PASS_QA_COMPLETE_READY_FOR_EXECUTION",
        "candidate_count": 22,
        "formal_sourcing_passes": 22,
        "blinded_qa1_passes": 22,
        "blinded_qa2_passes": 22,
        "distinct_qa_reviewers_per_candidate": True,
        "unresolved_or_adverse_findings": 0,
        "eligible_calibrations": calibrations,
        "quarantined_calibration_attempts": [failed_calibration],
        "contamination_incidents": [
            {
                "reviewer_id": "RQA22B",
                "affected_assignment": "qa2-batch-02",
                "detected_before_output": True,
                "contaminated_candidate_records_counted": 0,
                "action": "ABORTED_AND_REASSIGNED_TO_FRESH_RQA22C",
                "unaffected_frozen_output_retained": "qa2-batch-01",
            }
        ],
        "mechanical_validation": {
            "workbook_rows_exact": workbook_rows,
            "official_pdf_bindings_exact": 44,
            "unique_official_pdfs": len(pdf_hashes),
            "raw_field_hashes_recomputed": 22,
            "b_simulations_recomputed": 22,
            **forms,
            **cohort,
        },
        "duplicate_adjudication": duplicate,
        "execution_authorization": {
            "authorized_arms": list(ARMS),
            "excluded_arm": "tailscale_B",
            "models": list(MODELS),
            "replacement_cells_per_arm": 88,
            "total_replacement_cells": 264,
            "provider_calls_issued_by_this_script": 0,
        },
    }

    write_csv(QA_ROOT / "qa-coverage.csv", coverage)
    write_json(QA_ROOT / "final-review-summary.json", summary)
    write_json(
        QA_ROOT / "review-manifest-final.json",
        {
            "protocol_version": PROTOCOL_VERSION,
            "eligible_reviewers": {
                "sourcing_new": "RSRC22",
                "qa1_batches_01_03": "RQA22A",
                "qa2_batch_01": "RQA22B",
                "qa2_batches_02_03": "RQA22C",
            },
            "coverage_path": "protocol-qa/qa-coverage.csv",
            "summary_path": "protocol-qa/final-review-summary.json",
            "all_22_gates_pass": True,
        },
    )
    (QA_ROOT / "DUPLICATE_ADJUDICATION.md").write_text(
        "# Duplicate adjudication\n\n"
        "`c0369` shares its stem and keyed answer with historical corrected-workbook "
        "row 1370 (`castilla-y-leon|2022|reserva|95`). That historical item has only "
        "three options, uses a different distractor set, is ineligible for this four-option "
        "benchmark, and is absent from the active 500. The selected Andalucía item is not "
        "an exact full-content duplicate and does not collide with any retained question.\n\n"
        "Verdict: **PASS_NOT_AN_ACTIVE_OR_FULL_CONTENT_DUPLICATE**.\n",
        encoding="utf-8",
    )
    promote_package(summary, coverage)

    checksum_files = sorted(
        path
        for path in QA_ROOT.rglob("*")
        if path.is_file()
        and path.name != "checksums.sha256"
        and "__pycache__" not in path.parts
    )
    (QA_ROOT / "checksums.sha256").write_text(
        "".join(
            f"{sha256_file(path)}  {path.relative_to(QA_ROOT)}\n"
            for path in checksum_files
        ),
        encoding="utf-8",
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
