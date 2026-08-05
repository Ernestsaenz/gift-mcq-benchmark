"""Build the experiment-4-aug-26 benchmark dataset = 318 retained (already run in
exp-31-07-26, A & B) + 182 new non-negated replacements.

Outputs (in this folder):
  flat-A.xlsx / flat-B.xlsx   drop-in for medrag_eval (21-col contract + origin/negated cols)
  flat-A.csv  / flat-B.csv    same, readable
  benchmark-500.json          combined, one object per question with A and B forms
  new-182-flat-A.xlsx / new-182-flat-B.xlsx   ONLY the 182 to run

Source of truth:
  318 retained rows  -> data/experiment-31-07-26/balanced-flat-A.xlsx / -B.xlsx (by question_id)
  182 new questions  -> /private/tmp/ab182-q5i3oBTb/selected-packets.jsonl (A) + 2-field B swap
Repo corpus is read-only; nothing outside this folder is written.
"""
from __future__ import annotations
import json, csv
from pathlib import Path
from openpyxl import Workbook, load_workbook

HERE = Path(__file__).resolve().parent
EXP = HERE.parent / "experiment-31-07-26"
AB182 = Path("/private/tmp/ab182-q5i3oBTb")
SWAP = "Ninguna de las respuestas anteriores es correcta."
SPECIALTY = "aparato-digestivo"

COLS = ["question_id","region","year","specialty","exam_part","question_number",
    "question_text","option_a","option_b","option_c","option_d",
    "correct_letter","correct_option_text","flags","page_in_exam_pdf",
    "source_exam_pdf","source_answer_key_pdf","content_sha256","source_key",
    "selection_score","context_ids","origin","negated_stem"]

REGION_DISPLAY = {"andalucia":"Andalucía","aragon":"Aragón","castilla-la-mancha":"Castilla-La Mancha",
    "castilla-y-leon":"Castilla y León","comunidad-de-madrid":"Comunidad de Madrid",
    "comunitat-valenciana":"Comunitat Valenciana","galicia":"Galicia","illes-balears":"Illes Balears",
    "la-rioja":"La Rioja","navarra":"Navarra","region-de-murcia":"Región de Murcia"}

def read_flat(fn):
    wb = load_workbook(EXP/fn, read_only=True); ws = wb["questions"]
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    out = {r[hdr.index("question_id")]: dict(zip(hdr, r)) for r in ws.iter_rows(min_row=2, values_only=True)}
    wb.close(); return out

flatA = read_flat("balanced-flat-A.xlsx")
flatB = read_flat("balanced-flat-B.xlsx")
retained = json.loads((HERE/"inputs/retained_318_ids.json").read_text())
neg_by_id = {r["question_id"]: r["negated_stem"] for r in retained}
ids318 = [r["question_id"] for r in retained]
assert all(q in flatA and q in flatB for q in ids318), "some 318 id missing from flat files"

def row318(q, src):
    r = dict(src[q]); r["origin"] = "retained318"; r["negated_stem"] = neg_by_id[q]
    return {c: r.get(c, "") for c in COLS}

# ---- 182 new (packets) ----
# packet index spans selection + reserves + pool (dedup fix pulls 3 replacements from reserves/pool)
_idx = {}
for fn in ("selected-packets.jsonl", "reserve-packets.jsonl", "fully-passing-pool.jsonl"):
    for l in (AB182/fn).read_text().splitlines():
        if l.strip():
            p = json.loads(l); _idx.setdefault(p["candidate_id"], p)
final_ids = json.loads((HERE/"inputs/final_182_candidate_ids.json").read_text())
packets = [_idx[c] for c in final_ids]
assert len(packets) == 182, len(packets)
# order by frozen_rank to match proposed-selection.csv ranking
packets.sort(key=lambda p: (p["frozen_rank"], p["candidate_id"]))

def packet_rows(p, i):
    rf = p["raw_fields"]; L = rf["correct_letter"].lower(); prov = p["provenance"]
    qid = f"n{i:03d}"
    ctx = "|".join(c.get("context_id","") for c in p.get("context_chunks", [])) or ""
    base = {"question_id": qid, "region": REGION_DISPLAY.get(p["region"], p["region"]),
        "year": p["year"], "specialty": SPECIALTY, "exam_part": p["exam_part"],
        "question_number": p["question_number"],
        "question_text": rf["question_text"], "option_a": rf["option_a"], "option_b": rf["option_b"],
        "option_c": rf["option_c"], "option_d": rf["option_d"], "correct_letter": L,
        "flags": "", "page_in_exam_pdf": prov.get("exam_page",""),
        "source_exam_pdf": prov.get("workbook_exam_name",""), "source_answer_key_pdf": prov.get("workbook_key_name",""),
        "content_sha256": p["raw_fields_hash"], "source_key": p["source_key"],
        "selection_score": "", "context_ids": ctx, "origin": "new182", "negated_stem": False}
    a = dict(base); a["correct_option_text"] = rf["correct_option_text"]
    b = dict(base); b[f"option_{L}"] = SWAP; b["correct_option_text"] = SWAP
    return {c: a.get(c,"") for c in COLS}, {c: b.get(c,"") for c in COLS}

newA, newB, combined = [], [], []
for i, p in enumerate(packets, 1):
    a, b = packet_rows(p, i); newA.append(a); newB.append(b)

rowsA = [row318(q, flatA) for q in ids318] + newA
rowsB = [row318(q, flatB) for q in ids318] + newB
assert len(rowsA) == 500 and len(rowsB) == 500, (len(rowsA), len(rowsB))

# stem-level dedup guard across the full 500 (catches the same-stem-different-options case)
import re as _re, unicodedata as _ud
def _nst(s): return _re.sub(r"[\s\.:;,]+$", "", _re.sub(r"\s+", " ", _ud.normalize("NFKC", str(s or "")).strip().casefold()))
_stems = [_nst(r["question_text"]) for r in rowsA]
_dups = sorted({s for s in _stems if _stems.count(s) > 1})
assert not _dups, f"duplicate stems remain ({len(_dups)}): {_dups[:5]}"

def write_xlsx(path, rows):
    wb = Workbook(); ws = wb.active; ws.title = "questions"; ws.append(COLS)
    for r in rows: ws.append([r[c] for c in COLS])
    wb.save(path)
def write_csv(path, rows):
    with open(path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=COLS); w.writeheader(); w.writerows(rows)

write_xlsx(HERE/"flat-A.xlsx", rowsA); write_csv(HERE/"flat-A.csv", rowsA)
write_xlsx(HERE/"flat-B.xlsx", rowsB); write_csv(HERE/"flat-B.csv", rowsB)
write_xlsx(HERE/"new-182-flat-A.xlsx", newA); write_xlsx(HERE/"new-182-flat-B.xlsx", newB)

# combined JSON (A + B per question)
def as_q(a, b):
    return {"question_id": a["question_id"], "origin": a["origin"], "negated_stem": a["negated_stem"],
        "region": a["region"], "year": a["year"], "exam_part": a["exam_part"], "source_key": a["source_key"],
        "correct_letter": a["correct_letter"], "stem": a["question_text"],
        "A": {"options": {l: a[f"option_{l}"] for l in "abcd"}, "correct_option_text": a["correct_option_text"]},
        "B": {"options": {l: b[f"option_{l}"] for l in "abcd"}, "correct_option_text": b["correct_option_text"]}}
combined = [as_q(a, b) for a, b in zip(rowsA, rowsB)]
(HERE/"benchmark-500.json").write_text(json.dumps(combined, ensure_ascii=False, indent=1))

from collections import Counter
print("WROTE flat-A/B (xlsx+csv), new-182 flat A/B, benchmark-500.json")
print("flat-A rows:", len(rowsA), "| flat-B rows:", len(rowsB))
print("origin split A:", dict(Counter(r["origin"] for r in rowsA)))
print("negated in combined 500:", sum(1 for r in rowsA if r["negated_stem"]),
      "(all from retained318; new182 =", sum(1 for r in newA if r["negated_stem"]), ")")
print("new-182 correct_letter dist:", dict(Counter(r["correct_letter"] for r in newA)))
# sanity: every new182 B-form has the swap in exactly the keyed slot
bad = [r["question_id"] for a,r in zip(newA,newB) if r[f"option_{a['correct_letter']}"]!=SWAP or r["correct_option_text"]!=SWAP]
print("new182 B-swap errors:", bad)
