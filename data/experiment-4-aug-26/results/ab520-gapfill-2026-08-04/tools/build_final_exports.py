#!/usr/bin/env python3
"""Build read-only benchmark exports from the frozen July and August databases."""

from __future__ import annotations

import csv
import hashlib
import json
import os
import re
import sqlite3
import subprocess
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SCRIPT_PATH = Path(__file__).resolve()
OUT_ROOT = SCRIPT_PATH.parents[1]
ROOT = next(parent for parent in SCRIPT_PATH.parents if (parent / ".git").exists())
EXPORT_DIR = OUT_ROOT / "exports"
MANIFEST_DIR = OUT_ROOT / "manifests"
LEDGER_PATH = MANIFEST_DIR / "benchmark-cell-ledger.csv"
RUN_MANIFEST_PATH = MANIFEST_DIR / "run-manifest.json"
JULY_DB = ROOT / "data/experiment-31-07-26/experiment.sqlite"
AUGUST_DB = OUT_ROOT / "runs/ab520-gapfill-2026-08-04.sqlite"
FLAT_A = ROOT / "data/experiment-4-aug-26/flat-A.csv"
FLAT_B = ROOT / "data/experiment-4-aug-26/flat-B.csv"
MASTER_520 = Path("/private/tmp/ab182-q5i3oBTb/ab520-master.csv")
MASTER_520_MANIFEST = Path("/private/tmp/ab182-q5i3oBTb/ab520-master-manifest.json")
AB182 = Path("/private/tmp/ab182-q5i3oBTb")
FINAL_182_IDS = ROOT / "data/experiment-4-aug-26/inputs/final_182_candidate_ids.json"

CELL_RESULTS = EXPORT_DIR / "benchmark-6000-cell-results.csv"
QUESTION_CATALOG = EXPORT_DIR / "benchmark-520-question-catalog.csv"
UNRESOLVED = EXPORT_DIR / "unresolved-cells.csv"
PROVIDER_SUMMARY = EXPORT_DIR / "provider-condition-summary.csv"
MODEL_SUMMARY = EXPORT_DIR / "model-condition-summary.csv"
PAIRED_RESULTS = EXPORT_DIR / "openrouter-paired-ab-results.csv"
WORKBOOK = EXPORT_DIR / "benchmark-results-and-traceability.xlsx"
FINAL_MANIFEST = MANIFEST_DIR / "execution-manifest-final.json"
FINAL_REPORT = OUT_ROOT / "FINAL_EXECUTION_REPORT.md"
REPO_CHECK = OUT_ROOT / "repo-readonly-check.json"
OUTPUT_CHECKSUMS = MANIFEST_DIR / "output-checksums.sha256"

RAW_FIELDS = (
    "question_text",
    "option_a",
    "option_b",
    "option_c",
    "option_d",
    "correct_letter",
    "correct_option_text",
)

MODELS = (
    "google/gemini-3.6-flash",
    "google/gemma-4-26b-a4b-it",
    "qwen/qwen3.6-35b-a3b",
    "z-ai/glm-5.2",
)

DB_PROVIDER = {
    "openrouter": "openrouter",
    "tailscale": "tailscale_medical_rag",
    "tailscale_medical_rag": "tailscale_medical_rag",
}


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, rows: list[dict[str, Any]], fields: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if fields is None:
        fields = list(rows[0]) if rows else []
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def sha256_text(value: str | None) -> str:
    if value is None:
        return ""
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def load_jsonl(path: Path) -> list[dict[str, Any]]:
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]


def lp_hash(values: list[Any]) -> str:
    payload = bytearray()
    for value in values:
        raw = ("" if value is None else str(value)).encode("utf-8")
        payload.extend(len(raw).to_bytes(8, "big"))
        payload.extend(raw)
    return hashlib.sha256(payload).hexdigest()


def normalized_stem(value: str) -> str:
    normalized = unicodedata.normalize("NFKC", str(value or "")).strip().casefold()
    normalized = re.sub(r"\s+", " ", normalized)
    return re.sub(r"[\s\.:;,]+$", "", normalized)


def ro_connect(path: Path) -> sqlite3.Connection:
    connection = sqlite3.connect(f"file:{path}?mode=ro", uri=True)
    connection.row_factory = sqlite3.Row
    return connection


def one(rows: list[sqlite3.Row], label: str) -> sqlite3.Row:
    if len(rows) != 1:
        raise AssertionError(f"Expected exactly one {label}; found {len(rows)}")
    return rows[0]


def result_experiment(cell: dict[str, str]) -> tuple[str, Path]:
    if cell["score_status"] == "july_reused":
        return cell["prior_experiment"], JULY_DB

    arm = cell["arm"]
    question_id = cell["question_id"]
    origin = cell["origin"]
    model = cell["model"]
    if arm == "openrouter_A":
        if origin == "new182":
            return "ab520_gapfill_or_A_new182", AUGUST_DB
        if question_id == "b320" and model == "z-ai/glm-5.2":
            return "ab520_gapfill_or_A_b320_glm", AUGUST_DB
    elif arm == "openrouter_B":
        return "ab520_gapfill_or_B_new182", AUGUST_DB
    elif arm == "tailscale_A":
        if origin == "new182":
            numeric = int(question_id[1:])
            if numeric == 1:
                return "ab520_gapfill_ts_A_pilot_n001", AUGUST_DB
            if 2 <= numeric <= 6:
                return "ab520_gapfill_ts_A_ramp_n002_n006", AUGUST_DB
            return "ab520_gapfill_ts_A_rest_n007_n182", AUGUST_DB
        by_model = {
            "google/gemini-3.6-flash": "ab520_gapfill_ts_A_gap_gemini",
            "google/gemma-4-26b-a4b-it": "ab520_gapfill_ts_A_gap_gemma",
            "qwen/qwen3.6-35b-a3b": "ab520_gapfill_ts_A_gap_qwen",
            "z-ai/glm-5.2": "ab520_gapfill_ts_A_gap_glm",
        }
        return by_model[model], AUGUST_DB
    raise AssertionError(f"No result experiment mapping for {cell['cell_key']}")


def extract_effective_model(response_json: str | None, response_body: str | None) -> str:
    for candidate in (response_json, response_body):
        if not candidate:
            continue
        try:
            payload = json.loads(candidate)
        except (TypeError, json.JSONDecodeError):
            continue
        if isinstance(payload, dict):
            model = payload.get("model")
            if isinstance(model, str):
                return model
            nested = payload.get("data")
            if isinstance(nested, dict) and isinstance(nested.get("model"), str):
                return nested["model"]
    return ""


def user_content_length(request_json: str | None) -> int | str:
    if not request_json:
        return ""
    try:
        payload = json.loads(request_json)
    except json.JSONDecodeError:
        return ""
    messages = payload.get("messages") if isinstance(payload, dict) else None
    if not isinstance(messages, list):
        return ""
    total = 0
    found = False
    for message in messages:
        if not isinstance(message, dict) or message.get("role") != "user":
            continue
        content = message.get("content")
        if isinstance(content, str):
            total += len(content)
            found = True
    return total if found else ""


def fetch_result(
    connection: sqlite3.Connection,
    *,
    experiment: str,
    question_id: str,
    provider: str,
    model: str,
    run_index: int,
    expected_raw: dict[str, str],
) -> dict[str, Any]:
    logical_rows = connection.execute(
        """
        SELECT lc.*, q.question_id AS qid, q.question_text, q.option_a, q.option_b,
               q.option_c, q.option_d, q.correct_letter, q.correct_option_text
        FROM logical_calls lc
        JOIN experiments e ON e.id = lc.experiment_id
        JOIN questions q ON q.id = lc.question_id
        WHERE e.name = ? AND q.question_id = ? AND lc.provider = ?
          AND lc.model = ? AND lc.run_index = ?
        """,
        (experiment, question_id, DB_PROVIDER[provider], model, run_index),
    ).fetchall()
    logical = one(logical_rows, f"logical call for {experiment}/{question_id}/{model}")

    mismatches = [field for field in RAW_FIELDS if str(logical[field]) != str(expected_raw[field])]
    if mismatches:
        raise AssertionError(
            f"Exact input mismatch for {experiment}/{question_id}/{model}: {mismatches}"
        )

    attempts = connection.execute(
        """
        SELECT * FROM provider_attempts
        WHERE logical_call_id = ? ORDER BY attempt_index, id
        """,
        (logical["id"],),
    ).fetchall()
    parsed_rows = connection.execute(
        """
        SELECT * FROM parsed_answers
        WHERE logical_call_id = ? ORDER BY id
        """,
        (logical["id"],),
    ).fetchall()
    parsed_by_attempt: dict[int, sqlite3.Row] = {}
    for parsed in parsed_rows:
        if parsed["provider_attempt_id"] is not None:
            parsed_by_attempt[int(parsed["provider_attempt_id"])] = parsed

    score_rows = connection.execute(
        """
        SELECT s.*, p.parse_status, p.parse_method, p.selected_letter,
               p.selected_option_text, p.provider_attempt_id
        FROM scores s
        JOIN parsed_answers p ON p.id = s.parsed_answer_id
        WHERE s.logical_call_id = ? ORDER BY s.id
        """,
        (logical["id"],),
    ).fetchall()
    if len(score_rows) > 1:
        raise AssertionError(f"Multiple scores for logical call {logical['id']}")
    score = score_rows[0] if score_rows else None
    latest_attempt = attempts[-1] if attempts else None
    score_attempt = None
    if score is not None and score["provider_attempt_id"] is not None:
        score_attempt = next(
            (attempt for attempt in attempts if attempt["id"] == score["provider_attempt_id"]),
            None,
        )
    chosen_attempt = score_attempt or latest_attempt
    latest_parsed = parsed_by_attempt.get(int(latest_attempt["id"])) if latest_attempt else None
    chosen_parsed = (
        parsed_by_attempt.get(int(score["provider_attempt_id"]))
        if score is not None and score["provider_attempt_id"] is not None
        else latest_parsed
    )

    history = []
    for attempt in attempts:
        parsed = parsed_by_attempt.get(int(attempt["id"]))
        history.append(
            {
                "attempt_index": attempt["attempt_index"],
                "status_code": attempt["status_code"],
                "latency_ms": attempt["latency_ms"],
                "finish_reason": attempt["finish_reason"],
                "error_type": attempt["error_type"],
                "parse_status": parsed["parse_status"] if parsed else None,
                "selected_letter": parsed["selected_letter"] if parsed else None,
            }
        )

    result: dict[str, Any] = {
        "logical_call_id": logical["id"],
        "prompt_version": logical["prompt_version"],
        "attempt_count": len(attempts),
        "attempt_history_json": json.dumps(history, ensure_ascii=False, separators=(",", ":")),
        "exact_input_match_db": "TRUE",
        "final_execution_status": "scored" if score is not None else "unresolved_after_retries",
        "selected_letter": score["selected_letter"] if score else "",
        "selected_option_text": score["selected_option_text"] if score else "",
        "parse_status": score["parse_status"] if score else (chosen_parsed["parse_status"] if chosen_parsed else ""),
        "parse_method": score["parse_method"] if score else (chosen_parsed["parse_method"] if chosen_parsed else ""),
        "strict_correct": score["strict_correct"] if score else "",
        "lenient_correct": score["lenient_correct"] if score else "",
        "letter_correct": score["letter_correct"] if score else "",
        "text_correct": score["text_correct"] if score else "",
        "answer_text_matches_provided": score["answer_text_matches_provided"] if score else "",
        "latest_attempt_index": chosen_attempt["attempt_index"] if chosen_attempt else "",
        "latest_status_code": chosen_attempt["status_code"] if chosen_attempt else "",
        "latest_latency_ms": chosen_attempt["latency_ms"] if chosen_attempt else "",
        "latest_finish_reason": chosen_attempt["finish_reason"] if chosen_attempt else "",
        "latest_error_type": chosen_attempt["error_type"] if chosen_attempt else "",
        "request_sha256": chosen_attempt["request_sha256"] if chosen_attempt else "",
        "response_sha256": sha256_text(chosen_attempt["response_body"]) if chosen_attempt else "",
        "effective_model": extract_effective_model(
            chosen_attempt["response_json"] if chosen_attempt else None,
            chosen_attempt["response_body"] if chosen_attempt else None,
        ),
        "prompt_tokens": chosen_attempt["prompt_tokens"] if chosen_attempt else "",
        "completion_tokens": chosen_attempt["completion_tokens"] if chosen_attempt else "",
        "total_tokens": chosen_attempt["total_tokens"] if chosen_attempt else "",
        "request_user_content_char_count": user_content_length(chosen_attempt["request_json"] if chosen_attempt else None),
    }
    return result


def classify_failure(row: dict[str, Any]) -> str:
    if row["final_execution_status"] == "scored":
        return ""
    history = json.loads(row["attempt_history_json"])
    if row["provider"] == "openrouter":
        if any(item.get("finish_reason") == "length" for item in history):
            return "openrouter_glm_length_no_parse_after_retries"
        return "openrouter_unresolved_after_retries"
    if int(row["source_form_input_char_count"]) > 5000:
        immediate_500 = [
            item for item in history
            if item.get("status_code") == 500 and (item.get("latency_ms") or 0) < 2000
        ]
        if immediate_500:
            return "tailscale_http500_correlated_overlength_exact_input"
    if any(
        item.get("error_type") == "server_error" and (item.get("latency_ms") or 0) >= 145000
        for item in history
    ):
        return "tailscale_glm_server_error_150s_after_retries"
    if any(str(item.get("parse_status") or "").startswith("failed") for item in history):
        return "tailscale_unparsed_response_after_retries"
    return "tailscale_unresolved_after_retries"


def arm_summary(cell_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    output = []
    for arm in ("openrouter_A", "openrouter_B", "tailscale_A"):
        rows = [row for row in cell_rows if row["arm"] == arm]
        scored = [row for row in rows if row["final_execution_status"] == "scored"]
        output.append(
            {
                "arm": arm,
                "required_cells": len(rows),
                "scored_cells": len(scored),
                "unresolved_cells": len(rows) - len(scored),
                "july_reused_scored": sum(row["score_origin"] == "2026-07-31_reused" for row in scored),
                "august_gapfill_scored": sum(row["score_origin"] == "2026-08-04_gapfill" for row in scored),
                "strict_correct": sum(int(row["strict_correct"]) for row in scored),
                "strict_accuracy_scored": f"{sum(int(row['strict_correct']) for row in scored) / len(scored):.8f}" if scored else "",
                "coverage_fraction": f"{len(scored) / len(rows):.8f}" if rows else "",
            }
        )
    return output


def model_summary(cell_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    output = []
    for arm in ("openrouter_A", "openrouter_B", "tailscale_A"):
        for model in MODELS:
            rows = [row for row in cell_rows if row["arm"] == arm and row["model"] == model]
            scored = [row for row in rows if row["final_execution_status"] == "scored"]
            output.append(
                {
                    "arm": arm,
                    "model": model,
                    "required_cells": len(rows),
                    "scored_cells": len(scored),
                    "unresolved_cells": len(rows) - len(scored),
                    "strict_correct": sum(int(row["strict_correct"]) for row in scored),
                    "strict_accuracy_scored": f"{sum(int(row['strict_correct']) for row in scored) / len(scored):.8f}" if scored else "",
                    "coverage_fraction": f"{len(scored) / len(rows):.8f}" if rows else "",
                }
            )
    return output


def paired_openrouter(cell_rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    indexed = {
        (row["condition"], row["question_id"], row["source_key"], row["model"], int(row["run_index"])): row
        for row in cell_rows if row["provider"] == "openrouter"
    }
    paired_rows: list[dict[str, Any]] = []
    for question_id in [row["question_id"] for row in read_csv(FLAT_A)]:
        source_key = next(row["source_key"] for row in read_csv(FLAT_A) if row["question_id"] == question_id)
        for model in MODELS:
            a = indexed[("A", question_id, source_key, model, 1)]
            b = indexed[("B", question_id, source_key, model, 1)]
            paired = a["final_execution_status"] == b["final_execution_status"] == "scored"
            paired_rows.append(
                {
                    "question_id": question_id,
                    "source_key": source_key,
                    "origin": a["origin"],
                    "model": model,
                    "run_index": 1,
                    "paired_score_available": "TRUE" if paired else "FALSE",
                    "A_status": a["final_execution_status"],
                    "B_status": b["final_execution_status"],
                    "A_selected_letter": a["selected_letter"],
                    "B_selected_letter": b["selected_letter"],
                    "correct_letter": a["correct_letter"],
                    "A_strict_correct": a["strict_correct"],
                    "B_strict_correct": b["strict_correct"],
                    "B_minus_A_strict": int(b["strict_correct"]) - int(a["strict_correct"]) if paired else "",
                    "A_score_origin": a["score_origin"],
                    "B_score_origin": b["score_origin"],
                    "A_failure_class": a["failure_class"],
                    "B_failure_class": b["failure_class"],
                }
            )
    scored = [row for row in paired_rows if row["paired_score_available"] == "TRUE"]
    both_correct = sum(int(row["A_strict_correct"]) == 1 and int(row["B_strict_correct"]) == 1 for row in scored)
    a_only = sum(int(row["A_strict_correct"]) == 1 and int(row["B_strict_correct"]) == 0 for row in scored)
    b_only = sum(int(row["A_strict_correct"]) == 0 and int(row["B_strict_correct"]) == 1 for row in scored)
    both_wrong = len(scored) - both_correct - a_only - b_only
    summary = {
        "required_pairs": len(paired_rows),
        "paired_scored": len(scored),
        "unpaired": len(paired_rows) - len(scored),
        "A_strict_correct": sum(int(row["A_strict_correct"]) for row in scored),
        "B_strict_correct": sum(int(row["B_strict_correct"]) for row in scored),
        "A_accuracy_paired": sum(int(row["A_strict_correct"]) for row in scored) / len(scored),
        "B_accuracy_paired": sum(int(row["B_strict_correct"]) for row in scored) / len(scored),
        "B_minus_A_accuracy": (
            sum(int(row["B_strict_correct"]) for row in scored)
            - sum(int(row["A_strict_correct"]) for row in scored)
        ) / len(scored),
        "both_correct": both_correct,
        "A_only_correct": a_only,
        "B_only_correct": b_only,
        "both_incorrect": both_wrong,
    }
    return paired_rows, summary


def updated_catalog(cell_rows: list[dict[str, Any]], generated_at: str) -> list[dict[str, Any]]:
    """Build the current 500 + 20 reserves; the old master is retained-only lineage."""
    archived_master = read_csv(MASTER_520)
    base_fields = list(archived_master[0])
    retained_master = {
        row["question_id"]: row for row in archived_master if row["cohort"] == "previously_run_ab"
    }
    if len(retained_master) != 318:
        raise AssertionError(f"Archived retained lineage has {len(retained_master)} rows")

    selected_packets = load_jsonl(AB182 / "selected-packets.jsonl")
    reserve_packets_original = load_jsonl(AB182 / "reserve-packets.jsonl")
    pool_packets = load_jsonl(AB182 / "fully-passing-pool.jsonl")
    packet_index: dict[str, dict[str, Any]] = {}
    for packet in selected_packets + reserve_packets_original + pool_packets:
        packet_index.setdefault(packet["candidate_id"], packet)
    final_candidate_ids = json.loads(FINAL_182_IDS.read_text(encoding="utf-8"))
    if len(final_candidate_ids) != 182 or len(set(final_candidate_ids)) != 182:
        raise AssertionError("Final candidate ID list is not 182 unique IDs")
    final_packets = [packet_index[candidate_id] for candidate_id in final_candidate_ids]
    final_packet_by_source = {packet["source_key"]: packet for packet in final_packets}
    if len(final_packet_by_source) != 182:
        raise AssertionError("Final packets do not have unique source keys")

    # Three original reserves were promoted to replace duplicate stems. Keep the
    # surviving 17 and deterministically backfill from the remaining fully
    # passing pool, excluding the three known duplicate-stem removals.
    original_reserve_ids = {packet["candidate_id"] for packet in reserve_packets_original}
    final_ids = set(final_candidate_ids)
    surviving_reserves = [
        packet for packet in reserve_packets_original if packet["candidate_id"] not in final_ids
    ]
    removed_duplicate_ids = {"c2772", "c0225", "c1152"}
    negated_labels = json.loads((AB182 / "negstem-labels.json").read_text(encoding="utf-8"))

    def is_nonnegated(candidate_id: str) -> bool:
        label = negated_labels.get(candidate_id)
        if isinstance(label, dict):
            return label.get("negated") is False
        return label is False

    eligible_backfills = sorted(
        (
            packet for packet in pool_packets
            if packet["candidate_id"] not in final_ids
            and packet["candidate_id"] not in original_reserve_ids
            and packet["candidate_id"] not in removed_duplicate_ids
            and is_nonnegated(packet["candidate_id"])
        ),
        key=lambda packet: (packet["frozen_rank"], packet["candidate_id"]),
    )
    backfill_reserves = eligible_backfills[: 20 - len(surviving_reserves)]
    reserve_packets = surviving_reserves + backfill_reserves
    if len(surviving_reserves) != 17 or len(backfill_reserves) != 3 or len(reserve_packets) != 20:
        raise AssertionError(
            f"Unexpected reserve reconstruction: {len(surviving_reserves)} + {len(backfill_reserves)}"
        )

    primary_a = read_csv(FLAT_A)
    primary_b = {row["question_id"]: row for row in read_csv(FLAT_B)}
    if len(primary_a) != 500 or len(primary_b) != 500:
        raise AssertionError("Final flat forms are not 500 rows")

    # Validate uniqueness at the same bare-stem and source-key level used by the
    # final benchmark guard, now including the reconstructed reserve set.
    seen_stems = {normalized_stem(row["question_text"]): row["question_id"] for row in primary_a}
    seen_source_keys = {row["source_key"] for row in primary_a}
    for packet in reserve_packets:
        stem = normalized_stem(packet["raw_fields"]["question_text"])
        if stem in seen_stems:
            raise AssertionError(f"Reserve stem duplicate: {packet['candidate_id']} / {seen_stems[stem]}")
        if packet["source_key"] in seen_source_keys:
            raise AssertionError(f"Reserve source-key duplicate: {packet['source_key']}")
        seen_stems[stem] = packet["candidate_id"]
        seen_source_keys.add(packet["source_key"])

    sourcing_reviews = load_jsonl(AB182 / "sourcing-reviews.jsonl")
    qa_reviews = load_jsonl(AB182 / "qa-reviews-initial.jsonl") + load_jsonl(AB182 / "qa-reviews-expansion.jsonl")
    sourcing_passes: dict[str, list[dict[str, Any]]] = defaultdict(list)
    qa_passes: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for review in sourcing_reviews:
        if review.get("verdict") == "PASS":
            sourcing_passes[review["candidate_id"]].append(review)
    for review in qa_reviews:
        if review.get("verdict") == "PASS":
            qa_passes[review["candidate_id"]].append(review)

    def review_fields(candidate_id: str) -> dict[str, Any]:
        sourcing = sourcing_passes.get(candidate_id, [])
        qa = qa_passes.get(candidate_id, [])
        distinct_qa: list[dict[str, Any]] = []
        seen_reviewers: set[str] = set()
        for review in qa:
            reviewer = str(review.get("reviewer_id") or "")
            if reviewer and reviewer not in seen_reviewers:
                distinct_qa.append(review)
                seen_reviewers.add(reviewer)
        return {
            "sourcing_reviewer": sourcing[0].get("reviewer_id", "") if sourcing else "",
            "sourcing_verdict": "PASS" if sourcing else "",
            "qa1_reviewer": distinct_qa[0].get("reviewer_id", "") if distinct_qa else "",
            "qa1_verdict": "PASS" if distinct_qa else "",
            "qa2_reviewer": distinct_qa[1].get("reviewer_id", "") if len(distinct_qa) > 1 else "",
            "qa2_verdict": "PASS" if len(distinct_qa) > 1 else "",
            "review_record_count": len(sourcing) + len(distinct_qa),
        }

    def packet_base(
        packet: dict[str, Any],
        *,
        a_row: dict[str, str] | None,
        b_row: dict[str, str] | None,
        master_position: int,
        primary_position: int | str,
        cohort_rank: int,
        is_reserve: bool,
        reserve_note: str,
    ) -> dict[str, Any]:
        raw = packet["raw_fields"]
        correct_letter = str(raw["correct_letter"]).lower()
        if a_row is None:
            a_row = {
                "question_text": raw["question_text"],
                "option_a": raw["option_a"],
                "option_b": raw["option_b"],
                "option_c": raw["option_c"],
                "option_d": raw["option_d"],
                "correct_letter": correct_letter,
                "correct_option_text": raw["correct_option_text"],
                "flags": "",
                "source_exam_pdf": packet["provenance"].get("workbook_exam_name", ""),
                "source_answer_key_pdf": packet["provenance"].get("workbook_key_name", ""),
                "page_in_exam_pdf": str(packet["provenance"].get("exam_page", "")),
                "content_sha256": packet["raw_fields_hash"],
                "context_ids": "|".join(chunk.get("context_id", "") for chunk in packet.get("context_chunks", [])),
            }
        if b_row is None:
            b_row = dict(a_row)
            b_row[f"option_{correct_letter}"] = "Ninguna de las respuestas anteriores es correcta."
            b_row["correct_option_text"] = "Ninguna de las respuestas anteriores es correcta."
        changed = [
            field for field in ("question_text", "option_a", "option_b", "option_c", "option_d", "correct_letter", "correct_option_text")
            if str(a_row.get(field, "")) != str(b_row.get(field, ""))
        ]
        if changed != [f"option_{correct_letter}", "correct_option_text"]:
            raise AssertionError(f"Invalid B diff for {packet['candidate_id']}: {changed}")
        provenance = packet["provenance"]
        context_chunks = packet.get("context_chunks", [])
        contexts = "|".join(chunk.get("context_id", "") for chunk in context_chunks)
        context_hashes = "|".join(
            f"{chunk.get('context_id', '')}:{chunk.get('raw_sha256') or chunk.get('sha256') or ''}"
            for chunk in context_chunks
        )
        review = review_fields(packet["candidate_id"])
        question_id = packet["candidate_id"] if is_reserve else str(a_row.get("question_id", ""))
        pair_values = [
            a_row.get("question_text", ""), a_row.get("option_a", ""), a_row.get("option_b", ""),
            a_row.get("option_c", ""), a_row.get("option_d", ""), correct_letter,
            a_row.get("correct_option_text", ""), b_row.get("option_a", ""), b_row.get("option_b", ""),
            b_row.get("option_c", ""), b_row.get("option_d", ""), b_row.get("correct_option_text", ""),
        ]
        base = {
            "master_520_position": master_position,
            "primary_500_position": primary_position,
            "question_id": question_id,
            "cohort": "reserve_20" if is_reserve else "new_selected_182",
            "target_test_role": "reviewed_reserve" if is_reserve else "replacement_primary",
            "in_primary_500": "FALSE" if is_reserve else "TRUE",
            "is_reserve": "TRUE" if is_reserve else "FALSE",
            "cohort_rank": cohort_rank,
            "legacy_item_number": "",
            "selection_rank": "" if is_reserve else cohort_rank,
            "reserve_rank": cohort_rank if is_reserve else "",
            "already_run_in_ab_comparison": "FALSE",
            "ab_evaluation_status": "NOT_RUN_AS_OF_2026-07-31",
            "ab_evaluation_missing": "TRUE",
            "ab_evaluation_missing_reason": "New replacement or reserve at the historical snapshot",
            "recommended_run_scope": "PRIMARY_GAPFILL" if not is_reserve else "RESERVE_NOT_IN_PRIMARY_SCOPE",
            "prior_run_snapshot_date": "2026-07-31",
            "prior_ab_provider": "",
            "prior_ab_experiment_a": "",
            "prior_ab_experiment_b": "",
            "prior_ab_paired_model_count": 0,
            "prior_ab_expected_model_count": 4,
            "prior_ab_paired_models": "",
            "prior_a_correct_count": "",
            "prior_b_correct_count": "",
            "prior_a_selected_letters_by_model": "",
            "prior_b_selected_letters_by_model": "",
            "prior_a_backends": "",
            "prior_b_backends": "",
            "region": packet["region"],
            "region_source_label": a_row.get("region", packet["region"]),
            "year": packet["year"],
            "specialty": a_row.get("specialty", "aparato-digestivo"),
            "source_pair": packet["pair"],
            "exam_part": packet["exam_part"],
            "formal_type": packet["formal_type"],
            "formal_type_source_label": packet["formal_type"],
            "source_question_number": packet["question_number"],
            "question_text": a_row["question_text"],
            "question_stem_raw": raw["question_text"],
            "option_a_A": a_row["option_a"],
            "option_b_A": a_row["option_b"],
            "option_c_A": a_row["option_c"],
            "option_d_A": a_row["option_d"],
            "correct_letter": correct_letter,
            "correct_option_text_A": a_row["correct_option_text"],
            "option_a_B": b_row["option_a"],
            "option_b_B": b_row["option_b"],
            "option_c_B": b_row["option_c"],
            "option_d_B": b_row["option_d"],
            "correct_option_text_B": b_row["correct_option_text"],
            "b_replacement_text": "Ninguna de las respuestas anteriores es correcta.",
            "b_changed_fields": f"option_{correct_letter},correct_option_text",
            "b_diff_verified": "TRUE",
            "flags": a_row.get("flags", ""),
            "needs_attention": "",
            "case_id": "",
            "visual_id": "",
            "source_key": packet["source_key"],
            "source_sheet": packet["sheet"],
            "source_row": packet["source_row"],
            "source_corpus_path": "/Users/ernestsaenz/Programming/gift-project-compile/second-project/workbook-repairs-2026-07-30/outputs/all-regions-aparato-digestivo.corrected.xlsx",
            "source_corpus_sha256": packet["corpus_sha256"],
            "a_row_source_path": str(FLAT_A) if not is_reserve else str(AB182 / "fully-passing-pool.jsonl"),
            "b_row_source_path": str(FLAT_B) if not is_reserve else "mechanical_two_field_simulation_not_executed",
            "source_exam_pdf_reference": provenance.get("workbook_exam_name", ""),
            "source_exam_pdf_path": provenance.get("exam_pdf_path", ""),
            "source_exam_pdf_sha256": provenance.get("exam_pdf_sha256", ""),
            "page_in_exam_pdf": provenance.get("exam_page", ""),
            "source_answer_key_pdf_reference": provenance.get("workbook_key_name", ""),
            "source_answer_key_pdf_path": provenance.get("key_pdf_path", ""),
            "source_answer_key_pdf_sha256": provenance.get("key_pdf_sha256", ""),
            "legacy_content_sha256": a_row.get("content_sha256", packet["raw_fields_hash"]),
            "raw_fields_hash": packet["raw_fields_hash"],
            "assembled_question_text_sha256": packet["assembled_context_and_stem_sha256"],
            "master_question_pair_sha256": lp_hash(pair_values),
            "normalized_content_signature": packet["normalized_content_signature"],
            "semantic_signature": packet["semantic_signature"],
            "context_ids": a_row.get("context_ids", contexts),
            "context_hashes": context_hashes,
            "context_chunk_count": len(context_chunks),
            "review_protocol": packet["protocol_version"],
            "review_as_of_date": packet["as_of_date"],
            **review,
            "medical_guidance_evidence": "Recorded in sourcing/QA review JSONL",
            "selection_frozen_rank": packet["frozen_rank"],
            "balance_stratum": f"{packet['region']}|{packet['formal_type']}|{packet['year']}",
            "traceability_note": reserve_note if is_reserve else "Final new primary row after non-negated rebuild and stem-dedup QA.",
            "candidate_id": packet["candidate_id"],
            "reserve_selection_note": reserve_note,
        }
        return base

    catalog_base: list[dict[str, Any]] = []
    for position, a_row in enumerate(primary_a, start=1):
        question_id = a_row["question_id"]
        if a_row["origin"] == "retained318":
            master = dict(retained_master[question_id])
            master.update(
                {
                    "master_520_position": position,
                    "primary_500_position": position,
                    "cohort_rank": position,
                    "candidate_id": "",
                    "reserve_selection_note": "",
                }
            )
            catalog_base.append(master)
        else:
            packet = final_packet_by_source[a_row["source_key"]]
            catalog_base.append(
                packet_base(
                    packet,
                    a_row=a_row,
                    b_row=primary_b[question_id],
                    master_position=position,
                    primary_position=position,
                    cohort_rank=position - 318,
                    is_reserve=False,
                    reserve_note="",
                )
            )
    for reserve_rank, packet in enumerate(reserve_packets, start=1):
        note = (
            "Surviving reviewed reserve after three reserve promotions into the final 182."
            if packet["candidate_id"] in {item["candidate_id"] for item in surviving_reserves}
            else "Deterministic frozen-rank backfill from the fully passing pool after three reserve promotions; single expansion QA record."
        )
        catalog_base.append(
            packet_base(
                packet,
                a_row=None,
                b_row=None,
                master_position=500 + reserve_rank,
                primary_position="",
                cohort_rank=reserve_rank,
                is_reserve=True,
                reserve_note=note,
            )
        )

    by_question: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in cell_rows:
        by_question[row["question_id"]].append(row)

    current_fields = [
        "current_execution_snapshot_utc", "current_execution_scope",
        "current_in_scope_required_cells", "current_in_scope_scored_cells",
        "current_in_scope_unresolved_cells", "current_openrouter_A_scored_models",
        "current_openrouter_B_scored_models", "current_tailscale_A_scored_models",
        "current_openrouter_A_missing_models", "current_openrouter_B_missing_models",
        "current_tailscale_A_missing_models", "current_openrouter_ab_paired_models",
        "current_openrouter_ab_status", "current_all_requested_arms_status",
        "current_august_gapfill_scored_cells", "current_strict_correct_openrouter_A",
        "current_strict_correct_openrouter_B", "current_strict_correct_tailscale_A",
        "current_selected_letters_openrouter_A_by_model",
        "current_selected_letters_openrouter_B_by_model",
        "current_selected_letters_tailscale_A_by_model",
        "current_strict_correct_openrouter_A_by_model",
        "current_strict_correct_openrouter_B_by_model",
        "current_strict_correct_tailscale_A_by_model",
        "current_score_origin_by_arm_model", "current_execution_status_by_arm_model",
        "current_unresolved_details",
    ]
    output: list[dict[str, Any]] = []
    for base in catalog_base:
        question_id = str(base["question_id"])
        rows = by_question.get(question_id, [])
        current: dict[str, Any] = {
            "current_execution_snapshot_utc": generated_at,
            "current_execution_scope": "openrouter_A|openrouter_B|tailscale_A",
        }
        if base["is_reserve"] == "TRUE":
            current.update(
                {
                    "current_in_scope_required_cells": 0,
                    "current_in_scope_scored_cells": 0,
                    "current_in_scope_unresolved_cells": 0,
                    "current_openrouter_A_scored_models": 0,
                    "current_openrouter_B_scored_models": 0,
                    "current_tailscale_A_scored_models": 0,
                    "current_openrouter_A_missing_models": "",
                    "current_openrouter_B_missing_models": "",
                    "current_tailscale_A_missing_models": "",
                    "current_openrouter_ab_paired_models": 0,
                    "current_openrouter_ab_status": "RESERVE_NOT_IN_PRIMARY_500_SCOPE",
                    "current_all_requested_arms_status": "RESERVE_NOT_IN_PRIMARY_500_SCOPE",
                    "current_august_gapfill_scored_cells": 0,
                    "current_strict_correct_openrouter_A": "",
                    "current_strict_correct_openrouter_B": "",
                    "current_strict_correct_tailscale_A": "",
                    "current_selected_letters_openrouter_A_by_model": "",
                    "current_selected_letters_openrouter_B_by_model": "",
                    "current_selected_letters_tailscale_A_by_model": "",
                    "current_strict_correct_openrouter_A_by_model": "",
                    "current_strict_correct_openrouter_B_by_model": "",
                    "current_strict_correct_tailscale_A_by_model": "",
                    "current_score_origin_by_arm_model": "",
                    "current_execution_status_by_arm_model": "",
                    "current_unresolved_details": "",
                }
            )
        else:
            if len(rows) != 12:
                raise AssertionError(f"Primary question {question_id} has {len(rows)} cells, expected 12")
            arm_rows = {
                arm: [row for row in rows if row["arm"] == arm]
                for arm in ("openrouter_A", "openrouter_B", "tailscale_A")
            }
            for arm, values in arm_rows.items():
                if len(values) != 4:
                    raise AssertionError(f"{question_id}/{arm} has {len(values)} cells")
            scored_by_arm = {
                arm: [row for row in values if row["final_execution_status"] == "scored"]
                for arm, values in arm_rows.items()
            }
            missing_by_arm = {
                arm: [row["model"] for row in values if row["final_execution_status"] != "scored"]
                for arm, values in arm_rows.items()
            }
            paired_models = sum(
                any(row["model"] == model and row["final_execution_status"] == "scored" for row in arm_rows["openrouter_A"])
                and any(row["model"] == model and row["final_execution_status"] == "scored" for row in arm_rows["openrouter_B"])
                for model in MODELS
            )
            unresolved = [
                f"{row['arm']}:{row['model']}:{row['failure_class']}"
                for row in rows if row["final_execution_status"] != "scored"
            ]
            scored_total = sum(len(values) for values in scored_by_arm.values())

            def model_values(arm: str, field: str, scored_only: bool = False) -> str:
                values: dict[str, Any] = {}
                for model in MODELS:
                    row = next(item for item in arm_rows[arm] if item["model"] == model)
                    if scored_only and row["final_execution_status"] != "scored":
                        values[model] = None
                    else:
                        value = row[field]
                        values[model] = int(value) if field == "strict_correct" and value != "" else value
                return json.dumps(values, ensure_ascii=False, separators=(",", ":"))

            score_origins = {
                arm: {model: next(item for item in arm_rows[arm] if item["model"] == model)["score_origin"] for model in MODELS}
                for arm in arm_rows
            }
            execution_statuses = {
                arm: {model: next(item for item in arm_rows[arm] if item["model"] == model)["final_execution_status"] for model in MODELS}
                for arm in arm_rows
            }
            current.update(
                {
                    "current_in_scope_required_cells": 12,
                    "current_in_scope_scored_cells": scored_total,
                    "current_in_scope_unresolved_cells": 12 - scored_total,
                    "current_openrouter_A_scored_models": len(scored_by_arm["openrouter_A"]),
                    "current_openrouter_B_scored_models": len(scored_by_arm["openrouter_B"]),
                    "current_tailscale_A_scored_models": len(scored_by_arm["tailscale_A"]),
                    "current_openrouter_A_missing_models": " | ".join(missing_by_arm["openrouter_A"]),
                    "current_openrouter_B_missing_models": " | ".join(missing_by_arm["openrouter_B"]),
                    "current_tailscale_A_missing_models": " | ".join(missing_by_arm["tailscale_A"]),
                    "current_openrouter_ab_paired_models": paired_models,
                    "current_openrouter_ab_status": "COMPLETE_4_OF_4_MODELS" if paired_models == 4 else f"PARTIAL_{paired_models}_OF_4_MODELS",
                    "current_all_requested_arms_status": "COMPLETE_12_OF_12_CELLS" if scored_total == 12 else f"PARTIAL_{scored_total}_OF_12_CELLS",
                    "current_august_gapfill_scored_cells": sum(
                        row["score_origin"] == "2026-08-04_gapfill" and row["final_execution_status"] == "scored"
                        for row in rows
                    ),
                    "current_strict_correct_openrouter_A": sum(int(row["strict_correct"]) for row in scored_by_arm["openrouter_A"]),
                    "current_strict_correct_openrouter_B": sum(int(row["strict_correct"]) for row in scored_by_arm["openrouter_B"]),
                    "current_strict_correct_tailscale_A": sum(int(row["strict_correct"]) for row in scored_by_arm["tailscale_A"]),
                    "current_selected_letters_openrouter_A_by_model": model_values("openrouter_A", "selected_letter", True),
                    "current_selected_letters_openrouter_B_by_model": model_values("openrouter_B", "selected_letter", True),
                    "current_selected_letters_tailscale_A_by_model": model_values("tailscale_A", "selected_letter", True),
                    "current_strict_correct_openrouter_A_by_model": model_values("openrouter_A", "strict_correct", True),
                    "current_strict_correct_openrouter_B_by_model": model_values("openrouter_B", "strict_correct", True),
                    "current_strict_correct_tailscale_A_by_model": model_values("tailscale_A", "strict_correct", True),
                    "current_score_origin_by_arm_model": json.dumps(score_origins, ensure_ascii=False, separators=(",", ":")),
                    "current_execution_status_by_arm_model": json.dumps(execution_statuses, ensure_ascii=False, separators=(",", ":")),
                    "current_unresolved_details": " | ".join(unresolved),
                }
            )
        ordered_fields = base_fields + ["candidate_id", "reserve_selection_note"] + current_fields
        output.append({field: ({**base, **current}).get(field, "") for field in ordered_fields})
    if len(output) != 520:
        raise AssertionError(f"Updated catalog has {len(output)} rows")
    if len({row["source_key"] for row in output}) != 520:
        raise AssertionError("520 catalog source keys are not unique")
    return output


def add_sheet(workbook: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet(title=title)
    if not rows:
        sheet.append(["No rows"])
        return
    fields = list(rows[0])
    sheet.append(fields)
    for row in rows:
        sheet.append([row.get(field, "") for field in fields])
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    body_font = Font(name="Arial", size=9)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    for index, field in enumerate(fields, start=1):
        sample = [str(field)] + [str(row.get(field, "") or "") for row in rows[:200]]
        width = min(max(max(len(value.split("\n")[0]) for value in sample) + 2, 10), 55)
        if field in {"question_text", "question_stem_raw", "selected_option_text", "attempt_history_json", "current_unresolved_details", "traceability_note"}:
            width = 55
        sheet.column_dimensions[get_column_letter(index)].width = width


def write_workbook(
    catalog: list[dict[str, Any]],
    cells: list[dict[str, Any]],
    providers: list[dict[str, Any]],
    models: list[dict[str, Any]],
    unresolved: list[dict[str, Any]],
    paired: list[dict[str, Any]],
    generated_at: str,
) -> dict[str, Any]:
    workbook = Workbook()
    workbook.remove(workbook.active)
    add_sheet(workbook, "QuestionCatalog520", catalog)
    add_sheet(workbook, "CellResults6000", cells)
    add_sheet(workbook, "ProviderSummary", providers)
    add_sheet(workbook, "ModelSummary", models)
    add_sheet(workbook, "Unresolved", unresolved)
    add_sheet(workbook, "OpenRouterABPairs", paired)
    documentation = [
        {"field": "generated_at_utc", "value": generated_at},
        {"field": "scope", "value": "OpenRouter A, OpenRouter B, TailScale A; TailScale B excluded by request"},
        {"field": "primary_questions", "value": 500},
        {"field": "reserve_questions", "value": 20},
        {"field": "required_logical_cells", "value": 6000},
        {"field": "score_reuse", "value": "Hash-pinned July scores are labeled separately from August gapfill scores"},
        {"field": "missing_policy", "value": "Operational failures remain missing; no answer or score was inferred"},
        {"field": "condition_B", "value": "Exact verified two-field transformation; source strings were not edited during execution"},
    ]
    add_sheet(workbook, "Documentation", documentation)
    workbook.save(WORKBOOK)

    loaded = load_workbook(WORKBOOK, read_only=True, data_only=False)
    expected = {
        "QuestionCatalog520": (len(catalog) + 1, len(catalog[0])),
        "CellResults6000": (len(cells) + 1, len(cells[0])),
        "ProviderSummary": (len(providers) + 1, len(providers[0])),
        "ModelSummary": (len(models) + 1, len(models[0])),
        "Unresolved": (len(unresolved) + 1, len(unresolved[0])),
        "OpenRouterABPairs": (len(paired) + 1, len(paired[0])),
        "Documentation": (len(documentation) + 1, len(documentation[0])),
    }
    actual = {sheet.title: (sheet.max_row, sheet.max_column) for sheet in loaded.worksheets}
    if actual != expected:
        raise AssertionError(f"Workbook dimensions mismatch: {actual} != {expected}")
    formula_count = 0
    for sheet in loaded.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1
    loaded.close()
    if formula_count:
        raise AssertionError(f"Workbook unexpectedly contains {formula_count} formulas")
    return {"sheets": actual, "formula_count": formula_count}


def main() -> None:
    generated_at = datetime.now(timezone.utc).isoformat()
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    ledger = read_csv(LEDGER_PATH)
    if len(ledger) != 6000:
        raise AssertionError(f"Ledger has {len(ledger)} rows, expected 6000")
    if len({row["cell_key"] for row in ledger}) != 6000:
        raise AssertionError("Cell keys are not unique")

    flat_a_rows = read_csv(FLAT_A)
    flat_b_rows = read_csv(FLAT_B)
    if len(flat_a_rows) != 500 or len(flat_b_rows) != 500:
        raise AssertionError("Unexpected flat row count")
    flat = {
        "A": {row["question_id"]: row for row in flat_a_rows},
        "B": {row["question_id"]: row for row in flat_b_rows},
    }
    if len(flat["A"]) != 500 or len(flat["B"]) != 500:
        raise AssertionError("Question IDs are not unique in a flat form")

    july = ro_connect(JULY_DB)
    august = ro_connect(AUGUST_DB)
    cell_rows: list[dict[str, Any]] = []
    for cell in ledger:
        source = flat[cell["condition"]][cell["question_id"]]
        if source["source_key"] != cell["source_key"]:
            raise AssertionError(f"Source-key mismatch for {cell['cell_key']}")
        if source["origin"] != cell["origin"]:
            raise AssertionError(f"Origin mismatch for {cell['cell_key']}")
        if source["content_sha256"] != cell["content_sha256"]:
            raise AssertionError(f"Content hash mismatch for {cell['cell_key']}")
        experiment, database = result_experiment(cell)
        connection = july if database == JULY_DB else august
        result = fetch_result(
            connection,
            experiment=experiment,
            question_id=cell["question_id"],
            provider=cell["provider"],
            model=cell["model"],
            run_index=int(cell["run_index"]),
            expected_raw=source,
        )
        score_origin = "2026-07-31_reused" if database == JULY_DB else "2026-08-04_gapfill"
        source_form_input_char_count = sum(len(source[field]) for field in RAW_FIELDS[:5])
        raw_additions = {
            "source_csv": str(FLAT_A.relative_to(ROOT) if cell["condition"] == "A" else FLAT_B.relative_to(ROOT)),
            "region": source["region"],
            "year": source["year"],
            "specialty": source["specialty"],
            "exam_part": source["exam_part"],
            "question_number": source["question_number"],
            "question_text": source["question_text"],
            "option_a": source["option_a"],
            "option_b": source["option_b"],
            "option_c": source["option_c"],
            "option_d": source["option_d"],
            "correct_letter": source["correct_letter"],
            "correct_option_text": source["correct_option_text"],
            "flags": source["flags"],
            "page_in_exam_pdf": source["page_in_exam_pdf"],
            "source_exam_pdf": source["source_exam_pdf"],
            "source_answer_key_pdf": source["source_answer_key_pdf"],
            "selection_score": source["selection_score"],
            "context_ids": source["context_ids"],
            "negated_stem": source["negated_stem"],
            "source_form_input_char_count": source_form_input_char_count,
            "score_origin": score_origin,
            "result_database": str(database.relative_to(ROOT)),
            "result_experiment": experiment,
        }
        combined = {**cell, **raw_additions, **result}
        combined["failure_class"] = classify_failure(combined)
        cell_rows.append(combined)
    july.close()
    august.close()

    scored_count = sum(row["final_execution_status"] == "scored" for row in cell_rows)
    unresolved_rows = [row for row in cell_rows if row["final_execution_status"] != "scored"]
    if scored_count + len(unresolved_rows) != len(ledger):
        raise AssertionError(
            "Coverage invariant failed: "
            f"scored={scored_count} + unresolved={len(unresolved_rows)} "
            f"!= required={len(ledger)}"
        )

    provider_rows = arm_summary(cell_rows)
    model_rows = model_summary(cell_rows)
    paired_rows, paired_summary = paired_openrouter(cell_rows)
    catalog_rows = updated_catalog(cell_rows, generated_at)

    write_csv(CELL_RESULTS, cell_rows)
    unresolved_fields = [
        "cell_key", "arm", "provider", "condition", "question_id", "source_key", "origin",
        "model", "run_index", "failure_class", "attempt_count", "attempt_history_json",
        "latest_status_code", "latest_latency_ms", "latest_finish_reason", "latest_error_type",
        "parse_status", "source_form_input_char_count", "request_user_content_char_count",
        "result_experiment", "result_database", "source_workbook", "source_workbook_sha256",
        "source_excel_row", "content_sha256", "raw_form_sha256", "request_sha256", "response_sha256",
    ]
    write_csv(UNRESOLVED, unresolved_rows, unresolved_fields)
    write_csv(PROVIDER_SUMMARY, provider_rows)
    write_csv(MODEL_SUMMARY, model_rows)
    write_csv(PAIRED_RESULTS, paired_rows)
    write_csv(QUESTION_CATALOG, catalog_rows)

    workbook_validation = write_workbook(
        catalog_rows,
        cell_rows,
        provider_rows,
        model_rows,
        unresolved_rows,
        paired_rows,
        generated_at,
    )

    failure_counts = Counter(row["failure_class"] for row in unresolved_rows)
    tracked_unstaged = subprocess.check_output(
        ["git", "diff", "--name-only"], cwd=ROOT, text=True
    ).splitlines()
    tracked_staged = subprocess.check_output(
        ["git", "diff", "--cached", "--name-only"], cwd=ROOT, text=True
    ).splitlines()
    if tracked_unstaged or tracked_staged:
        raise AssertionError(
            f"Tracked repository state changed: unstaged={tracked_unstaged}, staged={tracked_staged}"
        )
    repository_status = subprocess.check_output(
        ["git", "status", "--short"], cwd=ROOT, text=True
    ).splitlines()
    repository_check = {
        "checked_at_utc": generated_at,
        "head": subprocess.check_output(["git", "rev-parse", "HEAD"], cwd=ROOT, text=True).strip(),
        "tracked_unstaged_changes": tracked_unstaged,
        "tracked_staged_changes": tracked_staged,
        "tracked_repository_unchanged": True,
        "status_short": repository_status,
        "note": "The canonical August result tree is under data/experiment-4-aug-26/results; unrelated untracked entries pre-existed this execution.",
    }
    REPO_CHECK.write_text(
        json.dumps(repository_check, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )

    packet_inputs = (
        ROOT / "data/experiment-4-aug-26/flat-A.xlsx",
        ROOT / "data/experiment-4-aug-26/flat-B.xlsx",
        ROOT / "data/experiment-4-aug-26/new-182-flat-A.xlsx",
        ROOT / "data/experiment-4-aug-26/new-182-flat-B.xlsx",
        FINAL_182_IDS,
        AB182 / "selected-packets.jsonl",
        AB182 / "reserve-packets.jsonl",
        AB182 / "fully-passing-pool.jsonl",
        AB182 / "negstem-labels.json",
        AB182 / "sourcing-reviews.jsonl",
        AB182 / "qa-reviews-initial.jsonl",
        AB182 / "qa-reviews-expansion.jsonl",
    )
    input_hashes = {
        str(path.relative_to(ROOT) if path.is_relative_to(ROOT) else path): sha256_file(path)
        for path in (
            LEDGER_PATH, RUN_MANIFEST_PATH, JULY_DB, AUGUST_DB, FLAT_A, FLAT_B,
            MASTER_520, MASTER_520_MANIFEST, *packet_inputs,
        )
    }
    outputs = [
        CELL_RESULTS, QUESTION_CATALOG, UNRESOLVED, PROVIDER_SUMMARY,
        MODEL_SUMMARY, PAIRED_RESULTS, WORKBOOK, REPO_CHECK,
    ]
    output_metadata = {
        str(path.relative_to(ROOT)): {"sha256": sha256_file(path), "size_bytes": path.stat().st_size}
        for path in outputs
    }
    catalog_status = Counter(row["current_all_requested_arms_status"] for row in catalog_rows)
    reserve_rows = [row for row in catalog_rows if row["is_reserve"] == "TRUE"]
    reserve_backfills = [
        row["candidate_id"] for row in reserve_rows
        if str(row["reserve_selection_note"]).startswith("Deterministic frozen-rank backfill")
    ]
    reserve_survivors = [row["candidate_id"] for row in reserve_rows if row["candidate_id"] not in reserve_backfills]
    final_manifest = {
        "artifact_version": "ab520-execution-final-v1",
        "generated_at_utc": generated_at,
        "scope": {
            "primary_questions": 500,
            "reserve_questions_documented_not_run": 20,
            "arms": ["openrouter_A", "openrouter_B", "tailscale_A"],
            "excluded_arm": "tailscale_B",
            "required_cells": 6000,
            "scored_cells": scored_count,
            "unresolved_cells": len(unresolved_rows),
        },
        "provider_condition_summary": provider_rows,
        "model_condition_summary": model_rows,
        "failure_classes": dict(sorted(failure_counts.items())),
        "openrouter_paired_ab": paired_summary,
        "question_catalog_status": dict(sorted(catalog_status.items())),
        "reserve_reconstruction": {
            "surviving_reserves_after_three_promotions": reserve_survivors,
            "deterministic_frozen_rank_backfills": reserve_backfills,
            "backfill_review_caveat": "Each of the three backfills has a sourcing PASS and one expansion-QA PASS; they are outside the executed primary 500.",
        },
        "input_hashes": input_hashes,
        "outputs": output_metadata,
        "validation": {
            "ledger_rows": len(ledger),
            "unique_cell_keys": len({row["cell_key"] for row in ledger}),
            "exact_database_input_matches": sum(row["exact_input_match_db"] == "TRUE" for row in cell_rows),
            "workbook": workbook_validation,
            "secrets_written": False,
        },
        "methodology_notes": [
            "July scores are reused only for hash-pinned retained cells and are labeled by origin.",
            "August executions use the existing mcq_es_v4 protocol with temperature 0 and GIFT prompt ID 13.",
            "Unparseable or provider-error responses remain missing; no answer or score was inferred.",
            f"The {failure_counts.get('tailscale_http500_correlated_overlength_exact_input', 0)} TailScale overlength cells are exact linked-context failures retained without input alteration.",
            "TailScale B was not part of the authorized execution scope.",
            "The archived 520 master predates the final non-negated rebuild and stem-dedup swaps; it was reused only for the unchanged retained-318 traceability, while current new/reserve membership was rebuilt from the final packet IDs.",
        ],
    }
    FINAL_MANIFEST.write_text(json.dumps(final_manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    provider_table = "\n".join(
        f"| {row['arm']} | {row['required_cells']} | {row['scored_cells']} | {row['unresolved_cells']} | {row['strict_correct']} | {float(row['strict_accuracy_scored']):.2%} |"
        for row in provider_rows
    )
    failure_table = "\n".join(f"| {name} | {count} |" for name, count in sorted(failure_counts.items()))
    overlength_count = failure_counts.get(
        "tailscale_http500_correlated_overlength_exact_input", 0
    )
    overlength_questions = len(
        {
            row["source_key"]
            for row in unresolved_rows
            if row["failure_class"]
            == "tailscale_http500_correlated_overlength_exact_input"
        }
    )
    gift_glm_count = failure_counts.get(
        "tailscale_glm_server_error_150s_after_retries", 0
    )
    openrouter_glm_count = failure_counts.get(
        "openrouter_glm_length_no_parse_after_retries", 0
    )
    report = f"""# Final execution report

Generated: {generated_at}

The authorized OpenRouter A, OpenRouter B, and TailScale A gapfill run is complete. The frozen 6,000-cell ledger contains **{scored_count:,} scored cells** and **{len(unresolved_rows)} fail-closed unresolved cells**. TailScale B was excluded from this run.

## Coverage and strict accuracy

| Arm | Required | Scored | Unresolved | Strict correct | Accuracy among scored |
|---|---:|---:|---:|---:|---:|
{provider_table}

OpenRouter has {paired_summary['paired_scored']:,}/{paired_summary['required_pairs']:,} scored A/B pairs. On those same paired cells, A strict accuracy is {paired_summary['A_accuracy_paired']:.2%}, B strict accuracy is {paired_summary['B_accuracy_paired']:.2%}, and B − A is {paired_summary['B_minus_A_accuracy']:.2%}.

## Unresolved cells

| Failure class | Cells |
|---|---:|
{failure_table}

The {overlength_count} correlated TailScale overlength failures cover {overlength_questions} linked-context questions. Their exact A text was preserved; no truncation, chunk removal, or protocol change was used. The {gift_glm_count} remaining TailScale failures are GLM server errors at the 150-second backend boundary. The {openrouter_glm_count} OpenRouter failures are GLM responses ending by length without a parseable answer after bounded retries.

## Deliverables

- `exports/benchmark-520-question-catalog.csv`: the single-table 520-question catalog (500 primary + 20 documented reserves), with historical and current coverage markers and full provenance.
- `exports/benchmark-6000-cell-results.csv`: every authorized question/model/arm cell, including exact input fields, score origin, selected answer, correctness, attempts, hashes, and failure class.
- `exports/openrouter-paired-ab-results.csv`: 2,000 OpenRouter A/B pair slots with paired-score availability and per-cell A/B outcomes.
- `exports/unresolved-cells.csv`: the complete fail-closed exception ledger.
- `exports/benchmark-results-and-traceability.xlsx`: the same material split into documented worksheets.
- `manifests/execution-manifest-final.json`: hashes, counts, validation, and methodology notes.

## Interpretation constraints

- The results combine a hash-pinned July reusable cohort with the August gapfill cohort; use `score_origin` when auditing or analyzing time-dependent performance.
- Accuracy denominators in the summary are scored cells, not the nominal 2,000 when unresolved cells remain.
- TailScale B was not run, so this execution does not supply a TailScale A/B comparison.
- Three original reserves were promoted during the final stem-dedup repair. The catalog restores 20 unique reserves with three deterministic frozen-rank backfills from the fully passing pool; each backfill has a sourcing PASS and one expansion-QA PASS and is explicitly labeled.
- The archived 520 master predates the final non-negated rebuild and duplicate-stem swaps. Only its unchanged retained-318 lineage was reused; all current new and reserve rows were rebuilt from the final candidate IDs and packet records.
- Model-only execution and QA are not a substitute for human specialist certification.

## Security note

A diagnostic command earlier in the session exposed credential values in the tool transcript. No secret is written to these artifacts, but the OpenRouter key and GIFT credential should be rotated after the run.
"""
    FINAL_REPORT.write_text(report, encoding="utf-8")
    checksum_targets = outputs + [FINAL_MANIFEST, FINAL_REPORT]
    OUTPUT_CHECKSUMS.write_text(
        "".join(f"{sha256_file(path)}  {path.relative_to(ROOT)}\n" for path in checksum_targets),
        encoding="utf-8",
    )
    print(json.dumps(final_manifest["scope"], indent=2))
    print(json.dumps(final_manifest["failure_classes"], indent=2))
    print(json.dumps(paired_summary, indent=2))


if __name__ == "__main__":
    main()
