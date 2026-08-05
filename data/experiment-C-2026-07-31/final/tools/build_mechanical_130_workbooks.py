from __future__ import annotations

import argparse
import hashlib
import json
import re
import unicodedata
from itertools import zip_longest
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.properties import PageSetupProperties

from build_mechanical_pair_workbooks import (
    AMBER,
    BLUE,
    GREEN,
    NAVY,
    WHITE,
    append_rows,
    atomic_save,
    chunks,
    require,
    set_font,
    sha256_text,
    style_table,
)
from mechanical_130_pipeline import ARMS, PROTOCOL, insert_only, read_source


def read_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def file_sha(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def normalized_question(value: str) -> str:
    value = unicodedata.normalize("NFC", str(value or ""))
    return re.sub(r"\s+", " ", value).strip().casefold()


def selected_records(source: Path, canonical_path: Path, qa_dir: Path) -> tuple[list[str], dict[str, list[dict]], dict]:
    source_header, source_rows = read_source(source)
    source_by_id = {row["question_id"]: row for row in source_rows}
    canonical = read_json(canonical_path)
    selection = read_json(qa_dir / "selection_manifest.json")
    require(selection["protocol_version"] == PROTOCOL, "selection protocol mismatch")
    legacy_qa_flags = {
        (item["question_id"], item["arm"]): item for item in selection.get("existing_qa_flags", [])
    }
    existing = read_json(qa_dir / "existing_locked.json")
    review_ledger = read_json(qa_dir / "review_ledger.json")["records"]
    reviewed = {(row["question_id"], row["arm"]): row for row in review_ledger}
    records_by_arm = {}
    for arm in ARMS:
        records = []
        for rank, old in enumerate(existing[arm], 1):
            qid = old["base_question_id"]
            source_row = source_by_id[qid]
            legacy_flag = legacy_qa_flags.get((qid, arm))
            records.append({
                **source_row,
                **old,
                "pool_role": "PRIMARY",
                "pool_rank": rank,
                "candidate_origin": "EXISTING_LOCKED",
                "reviewer_id": "existing_pool",
                "semantic_risk": "LEGACY_EXPLORATORY",
                "review_rationale": (
                    "Existing mechanical-pool text preserved exactly by user instruction. "
                    "One of three QA reviewers flagged a potential clinical contradiction; "
                    "the row remains a documented locked-baseline flag."
                    if legacy_flag
                    else "Existing mechanical-pool text preserved exactly by user instruction."
                ),
                "qa_status": "LEGACY_QA_FLAG_2_OF_3" if legacy_flag else "PASS_3_OF_3",
            })
        for addition in selection["arms"][arm]["additions"]:
            qid = addition["question_id"]
            source_row = source_by_id[qid]
            review = reviewed[(qid, arm)]
            control = str(source_row["question_text"])
            altered, separator = insert_only(
                control,
                addition["inserted_sentence"],
                addition["boundary_start"],
                addition["boundary_end"],
            )
            require(separator == addition["separator"], f"{arm} {qid}: separator drift")
            require(sha256_text(control) == addition["control_text_sha256"], f"{arm} {qid}: control hash drift")
            require(sha256_text(altered) == addition["altered_text_sha256"], f"{arm} {qid}: alteration hash drift")
            item = canonical["items"][qid]
            records.append({
                **source_row,
                "base_question_id": qid,
                "condition": arm,
                "variant_id": addition["variant_id"],
                "fabricated_entity": addition["fabricated_entity"],
                "inserted_sentence": addition["inserted_sentence"],
                "control_question_text": control,
                "altered_question_text": altered,
                "text_delta_qa": "PASS — insertion only",
                "inserted_character_offset": addition["boundary_end"],
                "inserted_segment_repr": repr(addition["inserted_sentence"] + separator),
                "control_text_sha256": addition["control_text_sha256"],
                "altered_text_sha256": addition["altered_text_sha256"],
                "strict_tier_member": "yes" if qid in canonical["tiers"]["strict"][arm] else "no",
                "relaxed_pair_member": "yes" if qid in canonical["tiers"]["relaxed"]["PAIR"] else "no",
                "cluster": item["cluster"],
                "pool_role": addition["pool_role"],
                "pool_rank": addition["pool_rank"],
                "candidate_origin": "NEW_REVIEWED",
                "reviewer_id": review["review"].get("reviewer_id", ""),
                "semantic_risk": addition["semantic_risk"],
                "review_rationale": addition["review_rationale"],
                "qa_status": "PASS_3_OF_3",
            })
        primary = [row for row in records if row["pool_role"] == "PRIMARY"]
        reserve = [row for row in records if row["pool_role"] == "RESERVE"]
        require(len(primary) == 100 and len(reserve) == 30, f"{arm}: role counts are not 100+30")
        require(len({row['base_question_id'] for row in records}) == 130, f"{arm}: duplicate base ID")
        require(len({normalized_question(row['control_question_text']) for row in records}) == 130, f"{arm}: duplicate normalized question text")
        records_by_arm[arm] = primary + reserve
    return source_header, records_by_arm, {"selection": selection, "review_ledger": review_ledger}


def add_readme(workbook: Workbook, arm: str, source: Path, canonical: Path, qa_dir: Path):
    ws = workbook.create_sheet("README")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 31
    ws.column_dimensions["B"].width = 105
    label = "biomarker" if arm == "BM" else "anatomy"
    selection = read_json(qa_dir / "selection_manifest.json")
    legacy_flags = [
        item["question_id"] for item in selection.get("existing_qa_flags", []) if item["arm"] == arm
    ]
    legacy_note = (
        f"Locked legacy rows with a non-unanimous QA interpretation: {', '.join(legacy_flags)}. "
        "They remain unchanged and are marked in qa_status and review_rationale."
        if legacy_flags
        else "None; all locked legacy rows received unanimous mechanical QA approval."
    )
    rows = [
        (f"Experiment C — {label} mechanical pool 100 + 30 reserve", ""),
        ("Contents", f"100 primary and 30 reserve {label} control-versus-alteration pairs."),
        ("Text guarantee", "Each control is exact source text. Each alteration is the control plus one assigned literal sentence and an exact copied source whitespace separator; no source character is deleted, replaced, reordered, normalized, or trimmed."),
        ("Existing rows", "The prior mechanical-pool membership, order, altered text, and rotated variant assignment are preserved exactly."),
        ("New rows", "New candidates come only from the same 474-row balanced-flat-A.xlsx bank and passed one blinded sourcing review plus three independent adversarial QA checks."),
        ("Placement rule", "Insertion after the patient description and immediately before the final interrogative/directive is acceptable, including after a declarative question premise."),
        ("Interpretation", "Exploratory mechanical pool; not clinically adjudicated as answer-key-preserving. J–N and other semantic risks are audit labels, not release claims."),
        ("Primary/reserve", "pool_role and pool_rank distinguish the 100 primary rows from the ordered 30-row replacement reserve."),
        ("Legacy QA flags", legacy_note),
        ("Source workbook", source.name),
        ("Source SHA-256", file_sha(source)),
        ("Canonical JSON", canonical.name),
        ("Canonical SHA-256", file_sha(canonical)),
        ("Selection manifest SHA-256", file_sha(qa_dir / "selection_manifest.json")),
    ]
    for row_number, (name, value) in enumerate(rows, 1):
        ws.cell(row_number, 1, name)
        ws.cell(row_number, 2, value)
        set_font(ws.cell(row_number, 1), bold=True, color=NAVY, size=15 if row_number == 1 else 10)
        set_font(ws.cell(row_number, 2), size=10)
        ws.cell(row_number, 1).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row_number, 2).alignment = Alignment(vertical="top", wrap_text=True)
        if row_number == 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
            ws.cell(1, 1).fill = PatternFill("solid", fgColor=BLUE)
            ws.row_dimensions[1].height = 28
        else:
            ws.row_dimensions[row_number].height = 42 if len(str(value)) > 120 else 27


def add_cards(workbook: Workbook, records: list[dict], arm: str):
    ws = workbook.create_sheet("paired_cards")
    ws.sheet_view.showGridLines = False
    for column, width in zip("ABCDEFGH", [5, 20, 20, 20, 20, 20, 20, 20]):
        ws.column_dimensions[column].width = width
    row_number = 1
    for index, record in enumerate(records):
        if index:
            ws.row_breaks.append(Break(id=row_number - 1))
        ws.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=8)
        title = ws.cell(
            row_number,
            1,
            f"{record['pool_role']} {record['pool_rank']} | {record['base_question_id']} | "
            f"{record['region']} {record['year']} | {record['variant_id']} | key {str(record['correct_letter']).upper()}",
        )
        set_font(title, bold=True, color=WHITE, size=11)
        title.fill = PatternFill("solid", fgColor=NAVY)
        row_number += 1
        ws.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=8)
        sentence = ws.cell(row_number, 1, f"Inserted string: {record['inserted_sentence']}")
        set_font(sentence, bold=True, color="375623", size=10)
        sentence.fill = PatternFill("solid", fgColor=GREEN)
        sentence.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row_number].height = 30
        row_number += 1
        ws.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=4)
        ws.merge_cells(start_row=row_number, start_column=5, end_row=row_number, end_column=8)
        for column, value in ((1, "CONTROL — exact source"), (5, f"ALTERATION — {arm} insertion only")):
            cell = ws.cell(row_number, column, value)
            set_font(cell, bold=True, color=NAVY, size=10)
            cell.fill = PatternFill("solid", fgColor=BLUE if column == 1 else AMBER)
            cell.alignment = Alignment(horizontal="center")
        row_number += 1
        for left, right in zip_longest(chunks(record["control_question_text"]), chunks(record["altered_question_text"]), fillvalue=""):
            ws.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=4)
            ws.merge_cells(start_row=row_number, start_column=5, end_row=row_number, end_column=8)
            for column, value in ((1, left), (5, right)):
                cell = ws.cell(row_number, column, value)
                set_font(cell, size=9)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            longest = max(len(left), len(right))
            ws.row_dimensions[row_number].height = min(390, max(30, 18 + 13 * ((longest // 90) + 1)))
            row_number += 1
        ws.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=8)
        qa = ws.cell(row_number, 1, f"QA: {record['qa_status']} — exactly one insertion; source fields and answer unchanged")
        set_font(qa, bold=True, color="375623", size=9)
        qa.fill = PatternFill("solid", fgColor=GREEN)
        row_number += 2
    ws.print_area = f"A1:H{row_number - 1}"
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True, autoPageBreaks=False)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.35, bottom=0.35, header=0.15, footer=0.15)


def add_source_rows(workbook: Workbook, name: str, source_header: list[str], records: list[dict], text_column: str):
    ws = workbook.create_sheet(name)
    columns = source_header + [
        "condition", "variant_id", "fabricated_entity", "inserted_sentence", "text_delta_qa",
        "pool_role", "pool_rank", "candidate_origin", "semantic_risk", "qa_status",
    ]
    output = []
    for record in records:
        row = {column: record.get(column, "") for column in source_header}
        row["question_text"] = record[text_column]
        row.update({column: record.get(column, "") for column in columns[len(source_header):]})
        output.append(row)
    append_rows(ws, columns, output)
    style_table(
        ws,
        {"A": 14, "B": 18, "G": 90, "H": 36, "I": 36, "J": 36, "K": 36, "M": 40, "Y": 60},
        {"question_text", "option_a", "option_b", "option_c", "option_d", "correct_option_text", "inserted_sentence"},
    )


def build_one(path: Path, arm: str, source_header: list[str], records: list[dict], source: Path, canonical: Path, qa_dir: Path, review_ledger: list[dict]):
    workbook = Workbook()
    comparison = workbook.active
    comparison.title = "control_vs_alteration"
    columns = [
        "pool_role", "pool_rank", "candidate_origin", "base_question_id", "region", "year", "specialty",
        "exam_part", "question_number", "condition", "variant_id", "fabricated_entity", "inserted_sentence",
        "control_question_text", "altered_question_text", "text_delta_qa", "inserted_character_offset",
        "inserted_segment_repr", "strict_tier_member", "relaxed_pair_member", "cluster", "semantic_risk",
        "reviewer_id", "qa_status", "review_rationale", "option_a", "option_b", "option_c", "option_d",
        "correct_letter", "correct_option_text", "flags", "page_in_exam_pdf", "source_exam_pdf",
        "source_answer_key_pdf", "source_key", "content_sha256", "control_text_sha256", "altered_text_sha256",
    ]
    append_rows(comparison, columns, records)
    style_table(
        comparison,
        {"A": 12, "B": 10, "D": 14, "E": 18, "K": 10, "L": 24, "M": 62, "N": 90, "O": 90,
         "Y": 60, "Z": 36, "AA": 36, "AB": 36, "AC": 36, "AE": 40, "AH": 45, "AI": 45},
        {"inserted_sentence", "control_question_text", "altered_question_text", "inserted_segment_repr", "review_rationale",
         "option_a", "option_b", "option_c", "option_d", "correct_option_text", "source_exam_pdf", "source_answer_key_pdf"},
    )
    for row_number in range(2, comparison.max_row + 1):
        comparison.cell(row_number, 16).fill = PatternFill("solid", fgColor=GREEN)
    add_source_rows(workbook, "control_rows", source_header, records, "control_question_text")
    add_source_rows(workbook, "alteration_rows", source_header, records, "altered_question_text")
    primary = [row for row in records if row["pool_role"] == "PRIMARY"]
    reserve = [row for row in records if row["pool_role"] == "RESERVE"]
    add_source_rows(workbook, "primary_control_rows", source_header, primary, "control_question_text")
    add_source_rows(workbook, "primary_alteration_rows", source_header, primary, "altered_question_text")
    add_source_rows(workbook, "reserve_control_rows", source_header, reserve, "control_question_text")
    add_source_rows(workbook, "reserve_alteration_rows", source_header, reserve, "altered_question_text")
    add_readme(workbook, arm, source, canonical, qa_dir)
    add_cards(workbook, records, arm)
    audit = workbook.create_sheet("filter_audit")
    audit_columns = [
        "question_id", "arm", "review_verdict", "semantic_risk", "mechanical_guard", "strict_guard", "relaxed_guard",
        "arm_gate_ok", "boundary_start", "boundary_end", "failure_codes", "rationale",
    ]
    audit_rows = []
    for item in review_ledger:
        if item["arm"] != arm:
            continue
        review = item["review"]
        prior = item["previous_filter"]
        audit_rows.append({
            "question_id": item["question_id"], "arm": arm, "review_verdict": review["verdict"],
            "semantic_risk": review["semantic_risk"], "mechanical_guard": prior["mechanical_guard"],
            "strict_guard": prior["strict_guard"], "relaxed_guard": prior["relaxed_guard"],
            "arm_gate_ok": prior["arm_gate_ok"], "boundary_start": review.get("boundary_start"),
            "boundary_end": review.get("boundary_end"), "failure_codes": ";".join(review.get("failure_codes", [])),
            "rationale": review.get("rationale", ""),
        })
    append_rows(audit, audit_columns, audit_rows)
    style_table(audit, {"A": 14, "C": 14, "D": 14, "K": 35, "L": 90}, {"failure_codes", "rationale"})
    workbook.active = 0
    atomic_save(workbook, path)


def verify_workbook(path: Path, arm: str, source: Path, existing_path: Path):
    _, source_rows = read_source(source)
    source_by_id = {row["question_id"]: row for row in source_rows}
    existing = load_workbook(existing_path, read_only=True, data_only=True)
    old_rows = list(existing["control_vs_alteration"].iter_rows(values_only=True))
    existing.close()
    old_header = list(old_rows[0])
    old = {row[0]: dict(zip(old_header, row)) for row in old_rows[1:] if row[0] is not None}
    workbook = load_workbook(path, read_only=True, data_only=False)
    require(workbook.sheetnames[:3] == ["control_vs_alteration", "control_rows", "alteration_rows"], f"{arm}: sheet order drift")
    rows = list(workbook["control_vs_alteration"].iter_rows(values_only=True))
    header = list(rows[0])
    records = [dict(zip(header, row)) for row in rows[1:] if row[0] is not None]
    require(len(records) == 130, f"{arm}: expected 130 comparison rows")
    require(sum(row["pool_role"] == "PRIMARY" for row in records) == 100, f"{arm}: primary count")
    require(sum(row["pool_role"] == "RESERVE" for row in records) == 30, f"{arm}: reserve count")
    require(len({normalized_question(row["control_question_text"]) for row in records}) == 130, f"{arm}: normalized-text duplicates")
    formulas = sum(1 for ws in workbook.worksheets for row in ws.iter_rows() for cell in row if cell.data_type == "f")
    require(formulas == 0, f"{arm}: formulas present")
    require(not workbook._external_links, f"{arm}: external links present")
    for record in records:
        qid = record["base_question_id"]
        source_row = source_by_id[qid]
        require(record["control_question_text"] == source_row["question_text"], f"{arm} {qid}: control mismatch")
        for field in ("option_a", "option_b", "option_c", "option_d", "correct_letter", "correct_option_text"):
            require(record[field] == source_row[field], f"{arm} {qid}: {field} mismatch")
        if qid in old:
            require(record["altered_question_text"] == old[qid]["altered_question_text"], f"{arm} {qid}: existing alteration changed")
    workbook.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, required=True)
    parser.add_argument("--canonical", type=Path, required=True)
    parser.add_argument("--qa-dir", type=Path, required=True)
    parser.add_argument("--out-dir", type=Path, required=True)
    parser.add_argument("--old-biomarker", type=Path, required=True)
    parser.add_argument("--old-anatomy", type=Path, required=True)
    args = parser.parse_args()
    source_header, records, evidence = selected_records(args.source, args.canonical, args.qa_dir)
    names = {
        "BM": "expC-biomarker-mechanical-100-plus-30-reserve-control-vs-alteration.xlsx",
        "AN": "expC-anatomy-mechanical-100-plus-30-reserve-control-vs-alteration.xlsx",
    }
    args.out_dir.mkdir(parents=True, exist_ok=True)
    outputs = {}
    for arm in ARMS:
        path = args.out_dir / names[arm]
        build_one(path, arm, source_header, records[arm], args.source, args.canonical, args.qa_dir, evidence["review_ledger"])
        old = args.old_biomarker if arm == "BM" else args.old_anatomy
        verify_workbook(path, arm, args.source, old)
        outputs[path.name] = {"arm": arm, "rows": 130, "primary": 100, "reserve": 30, "sha256": file_sha(path)}
    manifest = {
        "protocol_version": PROTOCOL,
        "source": args.source.name,
        "source_sha256": file_sha(args.source),
        "canonical": args.canonical.name,
        "canonical_sha256": file_sha(args.canonical),
        "selection_manifest_sha256": file_sha(args.qa_dir / "selection_manifest.json"),
        "outputs": outputs,
    }
    manifest_path = args.out_dir / "mechanical-130-manifest.json"
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(manifest, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
