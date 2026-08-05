from __future__ import annotations

import argparse
import collections
import difflib
import hashlib
import json
import re
from pathlib import Path

from openpyxl import load_workbook


PROTOCOL = "expc-mechanical-130-v1"
ARMS = ("BM", "AN")
ROTATIONS = {
    "BM": ("BM07", "BM02", "BM04", "BM08"),
    "AN": ("AN04", "AN10"),
}
NEEDED_NEW = {"BM": 31, "AN": 44}
NEW_PRIMARY = {"BM": 1, "AN": 14}
REQUIRED_CHECKS = (
    "patient_anchor",
    "narrative_stem_seam",
    "insertion_before_question",
    "no_source_edit_required",
    "grammatical_fit",
)
PROMPT = re.compile(
    r"^(?:¿|señale|señala|indique|indica|describa|describe|marque|elija|seleccione|"
    r"escoja|diga|conteste|responda|identifique|cuál|cuáles|qué|cómo|cuándo|"
    r"dónde|cuánt|quién|de los siguientes|de las siguientes|en relación|con respecto|"
    r"respecto|según|teniendo en cuenta|atendiendo|considerando|ante |si |pregunta de reserva)",
    re.IGNORECASE,
)


def require(condition: bool, message: str):
    if not condition:
        raise ValueError(message)


def sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def sha256_text(value: str) -> str:
    return sha256_bytes(value.encode("utf-8"))


def read_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, value):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def read_source(path: Path) -> tuple[list[str], list[dict]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows = list(workbook["questions"].iter_rows(values_only=True))
    workbook.close()
    header = [str(value) for value in rows[0]]
    records = [dict(zip(header, row)) for row in rows[1:] if row[0] is not None]
    require(len(records) == 474, f"expected 474 source rows, found {len(records)}")
    require(len({row['question_id'] for row in records}) == 474, "duplicate source question ID")
    return header, records


def read_existing(path: Path) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows = list(workbook["control_vs_alteration"].iter_rows(values_only=True))
    workbook.close()
    header = list(rows[0])
    return [dict(zip(header, row)) for row in rows[1:] if row[0] is not None]


def normalize(value: str) -> str:
    import unicodedata

    value = unicodedata.normalize("NFD", value or "")
    value = "".join(char for char in value if unicodedata.category(char) != "Mn")
    return re.sub(r"\s+", " ", value).casefold()


def cluster_ranks(source_rows: list[dict], canonical: dict) -> dict[str, int]:
    counts = collections.Counter()
    ranks = {}
    for row in source_rows:
        qid = row["question_id"]
        cluster = canonical["items"][qid]["cluster"]
        ranks[qid] = counts[cluster]
        counts[cluster] += 1
    return ranks


def sentence_table(canonical: dict, arm: str) -> dict[str, tuple[str, str, str]]:
    key = "biomarker" if arm == "BM" else "anatomy"
    return {variant: tuple(values) for variant, values in canonical["sentences"][key].items()}


def assign_variant(row: dict, arm: str, rank: int, canonical: dict) -> dict | None:
    table = sentence_table(canonical, arm)
    rotation = ROTATIONS[arm]
    whole = normalize(str(row["question_text"]))
    for step in range(len(rotation)):
        variant = rotation[(rank + step) % len(rotation)]
        entity, sentence, root = table[variant]
        if root not in whole:
            return {"variant_id": variant, "fabricated_entity": entity, "inserted_sentence": sentence}
    return None


def suggested_seams(text: str) -> list[dict]:
    candidates: dict[tuple[int, int], dict] = {}
    whitespace = list(re.finditer(r"\s+", text))
    for match in whitespace:
        suffix = text[match.end() :]
        stripped = suffix.lstrip()
        if stripped in {"", ":", ";"}:
            continue
        previous = text[match.start() - 1] if match.start() else ""
        prompt = bool(PROMPT.match(stripped))
        sentence_boundary = previous in ".!?:;" or "\n" in match.group()
        if not (prompt or sentence_boundary):
            continue
        start, end = match.span()
        candidates[(start, end)] = {
            "boundary_start": start,
            "boundary_end": end,
            "separator_repr": repr(text[start:end]),
            "prompt_like": prompt,
            "left_context": text[max(0, start - 220) : start],
            "right_context": text[end : min(len(text), end + 280)],
        }
    ordered = sorted(candidates.values(), key=lambda item: item["boundary_end"])
    return ordered[-30:]


def insert_only(control: str, sentence: str, start: int, end: int) -> tuple[str, str]:
    require(0 <= start < end <= len(control), "invalid boundary")
    separator = control[start:end]
    require(separator.isspace(), "boundary is not whitespace")
    altered = control[:end] + sentence + separator + control[end:]
    inserted = sentence + separator
    matcher = difflib.SequenceMatcher(a=control, b=altered, autojunk=False)
    changes = [item for item in matcher.get_opcodes() if item[0] != "equal"]
    require(len(changes) == 1 and changes[0][0] == "insert", "delta is not exactly one insertion")
    _, left_start, left_end, right_start, right_end = changes[0]
    require(left_start == left_end, "diff reports a source replacement")
    require(right_end - right_start == len(inserted), "inserted length mismatch")
    require(altered[:end] == control[:end], "source prefix changed")
    require(altered[end + len(inserted) :] == control[end:], "source suffix changed")
    require(altered.count(sentence) == 1, "sentence count is not exactly one")
    return altered, separator


def make_shards(source_rows: list[dict], canonical: dict) -> list[list[str]]:
    clusters: dict[str, list[str]] = collections.defaultdict(list)
    for row in source_rows:
        qid = row["question_id"]
        clusters[canonical["items"][qid]["cluster"]].append(qid)
    shards = [[] for _ in range(20)]
    sizes = [0] * 20
    for cluster, ids in sorted(clusters.items(), key=lambda item: (-len(item[1]), item[0])):
        index = min(range(20), key=lambda value: (sizes[value], value))
        shards[index].extend(ids)
        sizes[index] += len(ids)
    require(sorted(sizes) == [23] * 6 + [24] * 14, f"unexpected shard sizes: {sizes}")
    return shards


def prepare(args):
    source_header, source_rows = read_source(args.source)
    canonical = read_json(args.canonical)
    require(canonical["n_source_rows"] == 474, "canonical source count mismatch")
    source_by_id = {row["question_id"]: row for row in source_rows}
    existing = {
        "BM": read_existing(args.biomarker),
        "AN": read_existing(args.anatomy),
    }
    expected = {"BM": 99, "AN": 86}
    existing_ids = {}
    for arm in ARMS:
        existing_ids[arm] = [row["base_question_id"] for row in existing[arm]]
        require(len(existing_ids[arm]) == expected[arm], f"{arm}: existing count mismatch")
        require(existing_ids[arm] == canonical["tiers"]["relaxed"][arm], f"{arm}: existing order differs from canonical")
        for record in existing[arm]:
            qid = record["base_question_id"]
            require(record["control_question_text"] == source_by_id[qid]["question_text"], f"{arm} {qid}: existing control drift")
            sentence = record["inserted_sentence"]
            altered = record["altered_question_text"]
            offset = altered.index(sentence)
            require(altered[:offset] + altered[offset + len(sentence) :] != record["control_question_text"], f"{arm} {qid}: missing copied separator")
    ranks = cluster_ranks(source_rows, canonical)
    source_sha = sha256_bytes(args.source.read_bytes())
    canonical_sha = sha256_bytes(args.canonical.read_bytes())
    run = {
        "protocol_version": PROTOCOL,
        "source_path": str(args.source.resolve()),
        "source_sha256": source_sha,
        "canonical_path": str(args.canonical.resolve()),
        "canonical_sha256": canonical_sha,
        "source_header": source_header,
        "existing_workbooks": {
            "BM": {"path": str(args.biomarker.resolve()), "sha256": sha256_bytes(args.biomarker.read_bytes()), "rows": 99},
            "AN": {"path": str(args.anatomy.resolve()), "sha256": sha256_bytes(args.anatomy.read_bytes()), "rows": 86},
        },
        "required_new": NEEDED_NEW,
        "new_primary": NEW_PRIMARY,
        "existing_ids": existing_ids,
    }
    args.qa_dir.mkdir(parents=True, exist_ok=True)
    write_json(args.qa_dir / "run_manifest.json", run)
    write_json(args.qa_dir / "existing_locked.json", existing)
    shards = make_shards(source_rows, canonical)
    packets_dir = args.qa_dir / "review_packets"
    packets_dir.mkdir(parents=True, exist_ok=True)
    total = collections.Counter()
    for shard_index, shard_ids in enumerate(shards, 1):
        candidates = []
        for qid in shard_ids:
            source = source_by_id[qid]
            for arm in ARMS:
                if qid in set(existing_ids[arm]):
                    continue
                assignment = assign_variant(source, arm, ranks[qid], canonical)
                candidate = {
                    "question_id": qid,
                    "arm": arm,
                    "region": source["region"],
                    "year": source["year"],
                    "exam_part": source["exam_part"],
                    "cluster": canonical["items"][qid]["cluster"],
                    "question_text": source["question_text"],
                    "option_a": source["option_a"],
                    "option_b": source["option_b"],
                    "option_c": source["option_c"],
                    "option_d": source["option_d"],
                    "correct_letter": source["correct_letter"],
                    "correct_option_text": source["correct_option_text"],
                    "source_text_sha256": sha256_text(str(source["question_text"])),
                    "assignment": assignment,
                    "suggested_seams": suggested_seams(str(source["question_text"])),
                }
                candidates.append(candidate)
                total[arm] += 1
        packet = {
            "protocol_version": PROTOCOL,
            "reviewer_id": f"review_{shard_index:02d}",
            "shard_id": f"{shard_index:02d}",
            "source_sha256": source_sha,
            "canonical_sha256": canonical_sha,
            "instructions": {
                "scope": "Mechanical audit only; do not claim clinical key preservation.",
                "placement": "The sentence may be inserted after the patient description and immediately before the final interrogative/directive, including after a declarative premise. It must attach to a patient/context, read grammatically, and not contradict the source explicitly.",
                "blindness": "Do not inspect canonical guard outcomes, target counts, rankings, or other reviewer files.",
                "pass": "Choose one suggested whitespace seam and mark PASS only if every hard check passes and explicit_contradiction is NONE.",
            },
            "candidates": candidates,
        }
        write_json(packets_dir / f"review_{shard_index:02d}.json", packet)
    require(total == {"BM": 375, "AN": 388}, f"unexpected review coverage: {dict(total)}")
    print(json.dumps({"status": "prepared", "shards": 20, "coverage": dict(total)}, indent=2))


def validate_review(packet: dict, output: dict) -> list[dict]:
    require(output.get("protocol_version") == PROTOCOL, "review protocol mismatch")
    require(output.get("reviewer_id") == packet["reviewer_id"], "reviewer ID mismatch")
    require(output.get("shard_id") == packet["shard_id"], "shard ID mismatch")
    require(output.get("source_sha256") == packet["source_sha256"], "review source hash mismatch")
    require(output.get("canonical_sha256") == packet["canonical_sha256"], "review canonical hash mismatch")
    expected = {(item["question_id"], item["arm"]): item for item in packet["candidates"]}
    decisions = output.get("decisions", [])
    require(len(decisions) == len(expected), f"{packet['reviewer_id']}: decision count mismatch")
    require(len({(item['question_id'], item['arm']) for item in decisions}) == len(decisions), "duplicate review decision")
    validated = []
    for raw_decision in decisions:
        decision = dict(raw_decision)
        offered = decision.get("offered_boundary") or decision.get("selected_seam") or {}
        if decision.get("boundary_start") is None and offered:
            decision["boundary_start"] = offered.get("boundary_start")
            decision["boundary_end"] = offered.get("boundary_end")
        if decision.get("semantic_risk") is None:
            decision["semantic_risk"] = decision.get("risk")
        checks = dict(decision.get("checks", {}))
        legacy_checks = decision.get("required_checks", {})
        if legacy_checks:
            mapping = {
                "patient_anchor": "patient_context_attachment",
                "narrative_stem_seam": "placement_after_patient_description",
                "insertion_before_question": "before_final_interrogative_or_directive",
                "no_source_edit_required": "offered_whitespace_seam",
                "grammatical_fit": "grammatical_fit",
            }
            for target, source in mapping.items():
                if checks.get(target) is None and source in legacy_checks:
                    checks[target] = "PASS" if legacy_checks[source] is True else "FAIL"
            if checks.get("explicit_contradiction") is None and legacy_checks.get("explicit_contradiction") is not None:
                checks["explicit_contradiction"] = legacy_checks["explicit_contradiction"]
        if checks.get("explicit_contradiction") is None and decision.get("explicit_contradiction") is not None:
            checks["explicit_contradiction"] = decision["explicit_contradiction"]
        decision["checks"] = checks
        key = (decision.get("question_id"), decision.get("arm"))
        require(key in expected, f"unexpected review decision {key}")
        candidate = expected[key]
        require(decision.get("verdict") in {"PASS", "FAIL", "UNCERTAIN"}, f"{key}: invalid verdict")
        require(decision.get("semantic_risk") in {"LOW", "MEDIUM", "HIGH"}, f"{key}: invalid risk")
        require(all(checks.get(name) in {"PASS", "FAIL"} for name in REQUIRED_CHECKS), f"{key}: incomplete checks")
        require(checks.get("explicit_contradiction") in {"NONE", "PRESENT"}, f"{key}: bad contradiction check")
        record = {**candidate, "review": decision}
        if decision["verdict"] == "PASS":
            require(all(checks[name] == "PASS" for name in REQUIRED_CHECKS), f"{key}: PASS with failed check")
            require(checks["explicit_contradiction"] == "NONE", f"{key}: PASS with contradiction")
            require(candidate["assignment"] is not None, f"{key}: PASS without sentence assignment")
            pair = (decision.get("boundary_start"), decision.get("boundary_end"))
            seams = {(item["boundary_start"], item["boundary_end"]) for item in candidate["suggested_seams"]}
            require(pair in seams, f"{key}: selected seam was not offered")
            altered, separator = insert_only(
                str(candidate["question_text"]),
                candidate["assignment"]["inserted_sentence"],
                pair[0],
                pair[1],
            )
            record["altered_question_text"] = altered
            record["separator"] = separator
            record["altered_text_sha256"] = sha256_text(altered)
        validated.append(record)
    return validated


def aggregate_reviews(args):
    run = read_json(args.qa_dir / "run_manifest.json")
    canonical = read_json(Path(run["canonical_path"]))
    all_records = []
    for index in range(1, 21):
        packet = read_json(args.qa_dir / "review_packets" / f"review_{index:02d}.json")
        output = read_json(args.qa_dir / "reviewer_outputs" / f"review_{index:02d}.json")
        all_records.extend(validate_review(packet, output))
    require(len(all_records) == 763, f"review coverage mismatch: {len(all_records)}")
    eligible = []
    source_order = {qid: index for index, qid in enumerate(canonical["items"])}
    for record in all_records:
        qid, arm = record["question_id"], record["arm"]
        item = canonical["items"][qid]
        old_mechanical = item.get("mechanical_guard")
        old_relaxed = item.get("relaxed_guard")
        old_arm = item.get(f"{arm}_gate_ok") if old_mechanical is None else None
        record["previous_filter"] = {
            "mechanical_guard": old_mechanical,
            "strict_guard": item.get("strict_guard"),
            "relaxed_guard": old_relaxed,
            "arm_gate_ok": old_arm,
        }
        record["source_order"] = source_order[qid]
        if record["review"]["verdict"] == "PASS":
            eligible.append(record)
    counts = collections.Counter(record["arm"] for record in eligible)
    write_json(args.qa_dir / "review_ledger.json", {"protocol_version": PROTOCOL, "records": all_records})
    write_json(args.qa_dir / "eligible_candidates.json", {"protocol_version": PROTOCOL, "counts": counts, "records": eligible})
    existing = read_json(args.qa_dir / "existing_locked.json")
    qa_records = []
    for arm in ARMS:
        for record in existing[arm]:
            qa_records.append({
                "question_id": record["base_question_id"],
                "arm": arm,
                "candidate_type": "EXISTING_LOCKED",
                "variant_id": record["variant_id"],
                "inserted_sentence": record["inserted_sentence"],
                "control_question_text": record["control_question_text"],
                "altered_question_text": record["altered_question_text"],
                "control_text_sha256": record["control_text_sha256"],
                "altered_text_sha256": record["altered_text_sha256"],
                "seam_context": record["altered_question_text"][max(0, record["inserted_character_offset"] - 220):record["inserted_character_offset"] + len(record["inserted_sentence"]) + 280],
            })
    for record in eligible:
        start, end = record["review"]["boundary_start"], record["review"]["boundary_end"]
        qa_records.append({
            "question_id": record["question_id"],
            "arm": record["arm"],
            "candidate_type": "NEW_REVIEW_PASS",
            "variant_id": record["assignment"]["variant_id"],
            "inserted_sentence": record["assignment"]["inserted_sentence"],
            "boundary_start": start,
            "boundary_end": end,
            "control_question_text": record["question_text"],
            "altered_question_text": record["altered_question_text"],
            "control_text_sha256": record["source_text_sha256"],
            "altered_text_sha256": record["altered_text_sha256"],
            "seam_context": record["question_text"][max(0, start - 220):min(len(record["question_text"]), end + 280)],
        })
    qa_packet = {
        "protocol_version": PROTOCOL,
        "source_path": run["source_path"],
        "source_sha256": run["source_sha256"],
        "canonical_sha256": run["canonical_sha256"],
        "instructions": {
            "independence": "Do not inspect reviewer outputs, rankings, target counts, or other QA outputs.",
            "scope": "Audit every record. Existing text is intentionally locked; placement immediately before a final directive after a declarative premise is acceptable.",
            "hard_pass": "Exact single insertion, unchanged source fields/options/key/provenance, acceptable patient-attached seam, and no explicit contradiction.",
        },
        "records": qa_records,
    }
    write_json(args.qa_dir / "qa_packet.json", qa_packet)
    print(json.dumps({"status": "reviews_aggregated", "eligible": dict(counts), "qa_records": len(qa_records)}, indent=2))


def rank_key(record: dict):
    risk = {"LOW": 0, "MEDIUM": 1, "HIGH": 2}[record["review"]["semantic_risk"]]
    prior = record["previous_filter"]
    if prior["mechanical_guard"] is None and prior["arm_gate_ok"]:
        distance = 0
    elif prior["mechanical_guard"] is None:
        distance = 1
    else:
        distance = 2
    strict_risk = 0 if prior["strict_guard"] is None else 1
    return risk, distance, strict_risk, record["source_order"], record["question_id"]


def make_selection(args):
    run = read_json(args.qa_dir / "run_manifest.json")
    eligible_doc = read_json(args.qa_dir / "eligible_candidates.json")
    eligible = {(item["question_id"], item["arm"]): item for item in eligible_doc["records"]}
    qa_packet_path = args.qa_dir / "qa_packet.json"
    qa_packet_sha256 = sha256_bytes(qa_packet_path.read_bytes())
    qa_docs = []
    for index in range(1, 4):
        doc = read_json(args.qa_dir / "qa_outputs" / f"qa_{index:02d}.json")
        require(doc.get("protocol_version") == PROTOCOL, f"qa_{index:02d}: protocol mismatch")
        require(doc.get("qa_id") == f"qa_{index:02d}", f"qa_{index:02d}: ID mismatch")
        require(doc.get("source_sha256") == run["source_sha256"], f"qa_{index:02d}: source hash mismatch")
        require(doc.get("qa_packet_sha256") == qa_packet_sha256, f"qa_{index:02d}: QA packet hash mismatch")
        qa_docs.append(doc)
    packet = read_json(qa_packet_path)
    expected = {(item["question_id"], item["arm"], item["candidate_type"]): item for item in packet["records"]}
    qa_maps = []
    for doc in qa_docs:
        decisions = doc.get("decisions", [])
        require(len(decisions) == len(expected), f"{doc['qa_id']}: QA coverage mismatch")
        mapping = {(item.get("question_id"), item.get("arm"), item.get("candidate_type")): item for item in decisions}
        require(set(mapping) == set(expected), f"{doc['qa_id']}: QA record keys mismatch")
        for key, decision in mapping.items():
            require(decision.get("verdict") in {"PASS", "FAIL"}, f"{doc['qa_id']} {key}: invalid verdict")
            require(decision.get("exact_delta") in {"PASS", "FAIL"}, f"{doc['qa_id']} {key}: exact delta missing")
            require(decision.get("placement") in {"PASS", "FAIL"}, f"{doc['qa_id']} {key}: placement missing")
            require(decision.get("contradiction") in {"NONE", "PRESENT"}, f"{doc['qa_id']} {key}: contradiction missing")
        qa_maps.append(mapping)
    existing_failures = []
    approved = collections.defaultdict(list)
    for key in expected:
        passes = all(
            mapping[key]["verdict"] == "PASS"
            and mapping[key]["exact_delta"] == "PASS"
            and mapping[key]["placement"] == "PASS"
            and mapping[key]["contradiction"] == "NONE"
            for mapping in qa_maps
        )
        qid, arm, kind = key
        if kind == "EXISTING_LOCKED" and not passes:
            existing_failures.append({"question_id": qid, "arm": arm, "qa": [mapping[key] for mapping in qa_maps]})
        if kind == "NEW_REVIEW_PASS" and passes:
            approved[arm].append(eligible[(qid, arm)])
    existing = read_json(args.qa_dir / "existing_locked.json")
    duplicate_exclusions = collections.defaultdict(list)
    unique_approved = {}
    for arm in ARMS:
        seen = {normalize(str(item["control_question_text"])) for item in existing[arm]}
        require(len(seen) == len(existing[arm]), f"{arm}: existing pool contains normalized-text duplicates")
        unique_approved[arm] = []
        for record in sorted(approved[arm], key=rank_key):
            key = normalize(str(record["question_text"]))
            if key in seen:
                duplicate_exclusions[arm].append(record["question_id"])
                continue
            seen.add(key)
            unique_approved[arm].append(record)
    shortfall = {arm: max(0, NEEDED_NEW[arm] - len(unique_approved[arm])) for arm in ARMS}
    # Existing rows are a locked legacy baseline: report QA disagreements, but do
    # not silently replace or rewrite them. The fail-closed sufficiency gate
    # applies to the new, unanimously approved, duplicate-free candidate pool.
    if any(shortfall.values()):
        report = {
            "status": "FAILED_CLOSED",
            "existing_qa_flags": existing_failures,
            "approved_new": {arm: len(unique_approved[arm]) for arm in ARMS},
            "required_new": NEEDED_NEW,
            "shortfall": shortfall,
            "duplicate_exclusions": dict(duplicate_exclusions),
        }
        write_json(args.qa_dir / "shortfall_report.json", report)
        print(json.dumps(report, ensure_ascii=False, indent=2))
        raise SystemExit(2)
    shortfall_path = args.qa_dir / "shortfall_report.json"
    if shortfall_path.exists():
        shortfall_path.unlink()
    selection = {
        "protocol_version": PROTOCOL,
        "source": run,
        "arms": {},
        "qa_output_sha256": {},
        "existing_qa_flags": existing_failures,
    }
    for index in range(1, 4):
        path = args.qa_dir / "qa_outputs" / f"qa_{index:02d}.json"
        selection["qa_output_sha256"][path.name] = sha256_bytes(path.read_bytes())
    for arm in ARMS:
        ranked = unique_approved[arm]
        chosen = ranked[: NEEDED_NEW[arm]]
        primary_count = NEW_PRIMARY[arm]
        additions = []
        for index, record in enumerate(chosen):
            role = "PRIMARY" if index < primary_count else "RESERVE"
            additions.append({
                "question_id": record["question_id"],
                "arm": arm,
                "pool_role": role,
                "pool_rank": (len(run["existing_ids"][arm]) + index + 1) if role == "PRIMARY" else (index - primary_count + 1),
                "variant_id": record["assignment"]["variant_id"],
                "fabricated_entity": record["assignment"]["fabricated_entity"],
                "inserted_sentence": record["assignment"]["inserted_sentence"],
                "boundary_start": record["review"]["boundary_start"],
                "boundary_end": record["review"]["boundary_end"],
                "separator": record["separator"],
                "control_text_sha256": record["source_text_sha256"],
                "altered_text_sha256": record["altered_text_sha256"],
                "semantic_risk": record["review"]["semantic_risk"],
                "reviewer_id": record["review"].get("reviewer_id", ""),
                "review_rationale": record["review"].get("rationale", ""),
                "previous_filter": record["previous_filter"],
            })
        selection["arms"][arm] = {
            "existing_primary_ids": run["existing_ids"][arm],
            "additions": additions,
            "overflow_ids": [item["question_id"] for item in ranked[NEEDED_NEW[arm] :]],
            "duplicate_excluded_ids": duplicate_exclusions[arm],
        }
    write_json(args.qa_dir / "selection_manifest.json", selection)
    print(json.dumps({"status": "selection_ready", "BM": 130, "AN": 130}, indent=2))


def main():
    parser = argparse.ArgumentParser()
    subparsers = parser.add_subparsers(dest="command", required=True)
    prepare_parser = subparsers.add_parser("prepare")
    prepare_parser.add_argument("--source", type=Path, required=True)
    prepare_parser.add_argument("--canonical", type=Path, required=True)
    prepare_parser.add_argument("--biomarker", type=Path, required=True)
    prepare_parser.add_argument("--anatomy", type=Path, required=True)
    prepare_parser.add_argument("--qa-dir", type=Path, required=True)
    aggregate_parser = subparsers.add_parser("aggregate-reviews")
    aggregate_parser.add_argument("--qa-dir", type=Path, required=True)
    selection_parser = subparsers.add_parser("make-selection")
    selection_parser.add_argument("--qa-dir", type=Path, required=True)
    args = parser.parse_args()
    if args.command == "prepare":
        prepare(args)
    elif args.command == "aggregate-reviews":
        aggregate_reviews(args)
    else:
        make_selection(args)


if __name__ == "__main__":
    main()
