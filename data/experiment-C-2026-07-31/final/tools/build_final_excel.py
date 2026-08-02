from __future__ import annotations

import argparse
import hashlib
import json
import math
import os
import re
import tempfile
import unicodedata
import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.properties import PageSetupProperties


REQUIRED = [
    "question_id", "region", "year", "specialty", "exam_part",
    "question_number", "question_text", "option_a", "option_b", "option_c",
    "option_d", "correct_letter", "correct_option_text", "flags",
    "page_in_exam_pdf", "source_exam_pdf", "source_answer_key_pdf",
]
AUDIT = [
    "base_question_id", "condition", "validation_tier", "variant_id",
    "fabricated_entity", "inserted_sentence", "cluster", "source_key",
    "source_content_sha256", "question_text_sha256", "content_sha256",
    "selection_score", "context_ids", "clinical_qa_status", "clinical_qa_note",
]
BOUND = re.compile(r"(?:(?<=[.!?])\s+|\n+)")
PRIMARY = {
    "BM": (
        "BM07",
        "fibroquelina-X3",
        "La fibroquelina-X3 sérica se encuentra aumentada.",
        "fibro",
    ),
    "AN": (
        "AN04",
        "saco orfalónico",
        "La exploración revela dolor a la palpación del saco orfalónico.",
        "saco",
    ),
}

FONT = "Arial"
NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "E2F0D9"
AMBER = "FFF2CC"
RED = "F4CCCC"
GREY = "E7E6E6"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="B7C9D6")


def text_hash(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def require(condition: bool, message: str):
    if not condition:
        raise ValueError(message)


def item_hash(row: dict) -> str:
    fields = [row["question_text"]] + [row[f"option_{x}"] for x in "abcd"]
    return text_hash("\x1f".join(str(x) for x in fields))


def normal(value: str) -> str:
    value = unicodedata.normalize("NFD", value or "")
    value = "".join(c for c in value if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", value).casefold()


def insert_primary(value: str, sentence: str) -> str:
    source = unicodedata.normalize("NFC", value).rstrip()
    matches = list(BOUND.finditer(source))
    if not matches:
        raise ValueError("no insertion seam")
    match = matches[-1]
    prefix = source[: match.end()]
    separator = source[match.start() : match.end()]
    suffix = source[match.end() :]
    result = prefix + sentence + separator + suffix
    offset = len(prefix)
    require(result[:offset] == source[:offset], "insertion changed the source prefix")
    require(
        result[offset + len(sentence) + len(separator) :] == source[offset:],
        "insertion changed the source suffix",
    )
    require(
        len(result) == len(source) + len(sentence) + len(separator),
        "insertion length mismatch",
    )
    require(result.count(sentence) == 1, "inserted sentence count is not one")
    return result


def read_source(path: Path) -> tuple[list[str], list[dict]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows = list(workbook["questions"].iter_rows(values_only=True))
    header = [str(value) for value in rows[0]]
    records = [dict(zip(header, row)) for row in rows[1:] if row[0] is not None]
    require(not [column for column in REQUIRED if column not in header], "source workbook is missing required columns")
    ids = [record["question_id"] for record in records]
    require(len(ids) == len(set(ids)), "source workbook has duplicate question_id values")
    for record in records:
        qid = record["question_id"]
        letter = str(record.get("correct_letter") or "").casefold()
        require(letter in "abcd", f"{qid}: invalid correct_letter")
        require(record.get("correct_option_text") == record.get(f"option_{letter}"), f"{qid}: answer text does not match keyed option")
        require(all(record.get(f"option_{item}") for item in "abcd"), f"{qid}: empty option")
    return header, records


def add_flag(value, token: str) -> str:
    parts = [part.strip() for part in str(value or "").replace(";", ",").split(",") if part.strip()]
    if token not in parts:
        parts.append(token)
    return ";".join(parts)


def variant_row(source: dict, condition: str, decision: dict) -> dict:
    base = source["question_id"]
    row = dict(source)
    row["base_question_id"] = base
    row["condition"] = condition
    row["validation_tier"] = "strict_adversarial_pass"
    row["cluster"] = decision["cluster"]
    row["source_key"] = source.get("source_key") or ""
    row["source_content_sha256"] = source.get("content_sha256") or ""
    row["selection_score"] = source.get("selection_score") or ""
    row["context_ids"] = source.get("context_ids") or ""
    row["clinical_qa_status"] = "PASS"
    row["clinical_qa_note"] = decision.get("reason", "")
    original = unicodedata.normalize("NFC", str(source["question_text"])).rstrip()
    if condition == "CTRL":
        row["variant_id"] = "CTRL"
        row["fabricated_entity"] = ""
        row["inserted_sentence"] = ""
        row["question_text"] = original
    else:
        variant, entity, sentence, root = PRIMARY[condition]
        if root in normal(original):
            raise ValueError(f"{base}: primary {condition} root collision")
        row["variant_id"] = variant
        row["fabricated_entity"] = entity
        row["inserted_sentence"] = sentence
        row["question_text"] = insert_primary(original, sentence)
    row["question_id"] = f"c_{base}_{condition.lower()}"
    row["flags"] = add_flag(source.get("flags"), f"experiment_c:{condition.lower()}")
    row["question_text_sha256"] = text_hash(row["question_text"])
    row["content_sha256"] = item_hash(row)
    return row


def set_font(cell, *, bold=False, color="000000", size=9):
    cell.font = Font(name=FONT, size=size, bold=bold, color=color)


def table_style(ws, widths: dict[str, float], *, freeze="A2", filter_ref=None):
    for cell in ws[1]:
        set_font(cell, bold=True, color=WHITE, size=9)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN)
    ws.row_dimensions[1].height = 36
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            set_font(cell)
            cell.alignment = Alignment(vertical="top", wrap_text=False)
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = freeze
    ws.auto_filter.ref = filter_ref or ws.dimensions
    ws.sheet_view.zoomScale = 85
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True, autoPageBreaks=False)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:1"
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.4, bottom=0.4, header=0.2, footer=0.2)


def write_table(ws, columns: list[str], rows: list[dict], *, text_columns=()):
    ws.append(columns)
    for record in rows:
        ws.append([record.get(column, "") for column in columns])
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                cell.data_type = "s"
    index = {column: i + 1 for i, column in enumerate(columns)}
    for column in text_columns:
        if column not in index:
            continue
        for row in range(2, ws.max_row + 1):
            ws.cell(row, index[column]).alignment = Alignment(vertical="top", wrap_text=True)


def add_readme(workbook: Workbook, meta: dict):
    ws = workbook.create_sheet("README")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 100
    rows = [
        ("Experiment C — final adversarial-QA package", ""),
        ("Purpose", "One fabricated biomedical entity is inserted at the narrative-to-stem seam. Generation is deterministic; no language model authors any question."),
        ("Release set", f"{meta['n_base']} base questions × CTRL/BM/AN = {meta['n_rows']} importable rows."),
        ("Primary biomarker", f"{PRIMARY['BM'][0]} — {PRIMARY['BM'][2]}"),
        ("Primary anatomy", f"{PRIMARY['AN'][0]} — {PRIMARY['AN'][2]}"),
        ("Why these two", "They implement the user's requested one-sentence-per-arm design. The result supports inference about these two fixed stimuli, not every possible fictional biomarker or anatomical entity."),
        ("Validation policy", "Only strict paired items that passed item-by-item adversarial clinical adjudication in both arms appear in the runner-ready questions sheet."),
        ("Relaxed tier", "Not released as key-preserving. It remains in candidate_audit with explicit exploratory status."),
        ("Runner compatibility", "The first sheet is named questions and begins with the 17 columns required by medrag_eval.excel_io. Condition and base ID are also encoded in question_id and flags."),
        ("Print/readability", "Use review_cards for complete printable text. The raw data sheets are intended for filtering and machine import."),
        ("Control against carry-over", "Run the three condition-specific companion workbooks as separate datasets; do not expose the same model to adjacent CTRL/BM/AN copies in one session."),
        ("Outcome limitation", "Accuracy can measure whether the manipulation changes a selected answer, but it cannot identify fabricated-entity recognition or endorsement. The database also retains raw responses; a recognition endpoint still needs a requested, normalized rationale or explicit entity probe."),
        ("Source workbook", meta["source"]),
        ("Source SHA-256", meta["source_sha256"]),
        ("Canonical draft SHA-256", meta["canonical_sha256"]),
        ("Five-agent QA", "Lineage PASS; workbook usability FAIL then remediated; methodology FAIL then conservatively gated; report cross-check and reproducibility findings retained in qa_findings."),
        ("Security scan", meta["security"]),
    ]
    for row, (label, value) in enumerate(rows, 1):
        ws.cell(row, 1, label)
        ws.cell(row, 2, value)
        set_font(ws.cell(row, 1), bold=True, color=NAVY, size=10 if row > 1 else 15)
        set_font(ws.cell(row, 2), size=10)
        ws.cell(row, 1).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row, 2).alignment = Alignment(vertical="top", wrap_text=True)
        if row == 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
            ws.cell(1, 1).fill = PatternFill("solid", fgColor=BLUE)
            ws.row_dimensions[1].height = 28
        else:
            ws.cell(row, 1).fill = PatternFill("solid", fgColor=GREY)
            ws.row_dimensions[row].height = max(24, min(90, 15 * math.ceil(max(len(str(value)), 1) / 105)))
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True, autoPageBreaks=False)
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_area = f"A1:B{len(rows)}"
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5, header=0.2, footer=0.2)


def chunks(value: str, limit=700) -> list[str]:
    """Split printable text into paragraph-preserving, Excel-safe blocks."""
    out = []
    for paragraph in re.split(r"\n+", str(value)):
        paragraph = paragraph.strip()
        if not paragraph:
            continue
        while len(paragraph) > limit:
            cut = max(
                paragraph.rfind(". ", 0, limit),
                paragraph.rfind("; ", 0, limit),
                paragraph.rfind(" ", 0, limit),
            )
            if cut < limit // 2:
                cut = limit
            else:
                cut += 1
            out.append(paragraph[:cut].strip())
            paragraph = paragraph[cut:].strip()
        if paragraph:
            out.append(paragraph)
    return out


def print_height(value: str, *, chars_per_line=88, line_height=17, minimum=26, maximum=390) -> float:
    lines = sum(max(1, math.ceil(len(line) / chars_per_line)) for line in str(value).splitlines() or [""])
    return min(maximum, max(minimum, 8 + line_height * lines))


def add_review_cards(workbook: Workbook, rows: list[dict]):
    ws = workbook.create_sheet("review_cards")
    ws.sheet_view.showGridLines = False
    for column, width in zip("ABCDEFGH", [5, 17, 17, 17, 17, 17, 17, 17]):
        ws.column_dimensions[column].width = width
    row = 1
    for index, record in enumerate(rows):
        if index:
            ws.row_breaks.append(Break(id=row - 1))
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        title = ws.cell(row, 1, f"{record['question_id']}  |  {record['region']} {record['year']}  |  {record['condition']}")
        set_font(title, bold=True, color=WHITE, size=12)
        title.fill = PatternFill("solid", fgColor=NAVY)
        title.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 24
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        meta = ws.cell(row, 1, f"Base {record['base_question_id']} · {record['exam_part']} · source page {record['page_in_exam_pdf']} · {record['variant_id']}")
        set_font(meta, color=NAVY, size=9)
        meta.fill = PatternFill("solid", fgColor=BLUE)
        row += 1
        for part in chunks(record["question_text"]):
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            cell = ws.cell(row, 1, part)
            set_font(cell, size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.row_dimensions[row].height = print_height(part)
            row += 1
        for letter in "abcd":
            ws.cell(row, 1, letter.upper())
            set_font(ws.cell(row, 1), bold=True, color=NAVY, size=10)
            ws.cell(row, 1).alignment = Alignment(horizontal="center", vertical="top")
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
            cell = ws.cell(row, 2, record[f"option_{letter}"])
            set_font(cell, size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.row_dimensions[row].height = print_height(cell.value, chars_per_line=84, maximum=190)
            row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        answer = ws.cell(row, 1, f"Key: {record['correct_letter'].upper()} — {record['correct_option_text']}")
        set_font(answer, bold=True, color="375623", size=10)
        answer.fill = PatternFill("solid", fgColor=GREEN)
        answer.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[row].height = print_height(answer.value, chars_per_line=84, maximum=140)
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        source = ws.cell(row, 1, f"Source key: {record['source_key']} | exam PDF: {record['source_exam_pdf']}")
        set_font(source, color="666666", size=8)
        source.alignment = Alignment(vertical="top", wrap_text=True)
        row += 2
    ws.print_area = f"A1:H{row - 1}"
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True, autoPageBreaks=False)
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.4, bottom=0.4, header=0.2, footer=0.2)


def save_import(path: Path, rows: list[dict], condition: str):
    workbook = Workbook()
    ws = workbook.active
    ws.title = "questions"
    selected = [row for row in rows if row["condition"] == condition]
    columns = REQUIRED + AUDIT
    write_table(ws, columns, selected, text_columns=("question_text", "option_a", "option_b", "option_c", "option_d", "correct_option_text"))
    table_style(ws, {"A": 18, "B": 18, "G": 90, "H": 38, "I": 38, "J": 38, "K": 38, "M": 38})
    ws.print_area = "A1:A1"
    workbook.active = 0
    atomic_save_workbook(workbook, path)


def atomic_save_workbook(workbook: Workbook, path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    handle = tempfile.NamedTemporaryFile(prefix=f".{path.stem}.", suffix=".xlsx", dir=path.parent, delete=False)
    temporary = Path(handle.name)
    handle.close()
    try:
        workbook.save(temporary)
        with zipfile.ZipFile(temporary) as archive:
            require(archive.testzip() is None, f"{path.name}: corrupt OOXML archive")
        os.replace(temporary, path)
    finally:
        if temporary.exists():
            temporary.unlink()


def atomic_write_text(path: Path, value: str):
    path.parent.mkdir(parents=True, exist_ok=True)
    handle = tempfile.NamedTemporaryFile(prefix=f".{path.name}.", suffix=".tmp", dir=path.parent, delete=False, mode="w", encoding="utf-8")
    temporary = Path(handle.name)
    try:
        handle.write(value)
        handle.flush()
        os.fsync(handle.fileno())
        handle.close()
        os.replace(temporary, path)
    finally:
        if not handle.closed:
            handle.close()
        if temporary.exists():
            temporary.unlink()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, required=True)
    parser.add_argument("--canonical", type=Path, required=True)
    parser.add_argument("--adjudication", type=Path, required=True)
    parser.add_argument("--out-dir", type=Path, required=True)
    args = parser.parse_args()

    canonical = json.loads(args.canonical.read_text(encoding="utf-8"))
    adjudication = json.loads(args.adjudication.read_text(encoding="utf-8"))
    _, source_rows = read_source(args.source)
    source_by_id = {row["question_id"]: row for row in source_rows}
    release_ids = adjudication["paired_keep"]
    require(bool(release_ids), "adjudication released no questions")
    require(len(release_ids) == len(set(release_ids)), "adjudication has duplicate release IDs")
    require(set(release_ids) <= set(canonical["tiers"]["strict"]["PAIR"]), "release contains a non-strict-paired ID")
    require(len(source_rows) == canonical["n_source_rows"] == 474, "source/canonical row-count mismatch")
    source_sha256 = hashlib.sha256(args.source.read_bytes()).hexdigest()
    canonical_sha256 = hashlib.sha256(args.canonical.read_bytes()).hexdigest()
    adjudication_sha256 = hashlib.sha256(args.adjudication.read_bytes()).hexdigest()
    builder_sha256 = hashlib.sha256(Path(__file__).read_bytes()).hexdigest()
    require(source_sha256 == adjudication["expected_source_sha256"], "source workbook hash mismatch")
    require(canonical_sha256 == adjudication["expected_canonical_sha256"], "canonical JSON hash mismatch")

    rows = []
    for qid in release_ids:
        decision = dict(adjudication["decisions"][qid])
        decision["cluster"] = canonical["items"][qid]["cluster"]
        for condition in ("CTRL", "BM", "AN"):
            rows.append(variant_row(source_by_id[qid], condition, decision))

    for row in rows:
        letter = row["correct_letter"]
        require(letter in "abcd", f"{row['question_id']}: invalid answer letter")
        require(row["correct_option_text"] == row[f"option_{letter}"], f"{row['question_id']}: keyed answer mismatch")
        require(len(row["question_text"]) <= 32767, f"{row['question_id']}: Excel cell limit exceeded")
    require(len({row["question_id"] for row in rows}) == len(rows), "output question IDs are not unique")

    args.out_dir.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    ws = workbook.active
    ws.title = "questions"
    columns = REQUIRED + AUDIT
    write_table(ws, columns, rows, text_columns=("question_text", "option_a", "option_b", "option_c", "option_d", "correct_option_text", "clinical_qa_note"))
    table_style(ws, {"A": 18, "B": 18, "G": 90, "H": 38, "I": 38, "J": 38, "K": 38, "M": 38})
    ws.print_area = "A1:A1"
    for column in ("question_id", "question_text", "correct_letter", "flags"):
        cell = ws.cell(1, columns.index(column) + 1)
        cell.comment = Comment("Required by tier1_mcq/code/medrag_eval/excel_io.py", "Codex")

    meta = {
        "n_base": len(release_ids),
        "n_rows": len(rows),
        "source": args.source.name,
        "source_sha256": source_sha256,
        "canonical_sha256": canonical_sha256,
        "security": adjudication["security_scan"],
    }
    add_readme(workbook, meta)

    review_columns = [
        "base_question_id", "region", "year", "exam_part", "cluster", "correct_letter",
        "option_a", "option_b", "option_c", "option_d", "correct_option_text",
        "CTRL_text", "BM_text", "AN_text", "source_key", "page_in_exam_pdf",
        "source_exam_pdf", "source_answer_key_pdf", "clinical_qa_note",
    ]
    review_rows = []
    for qid in release_ids:
        src = source_by_id[qid]
        decision = adjudication["decisions"][qid]
        row_by_condition = {row["condition"]: row for row in rows if row["base_question_id"] == qid}
        review_rows.append({
            **src,
            "base_question_id": qid,
            "cluster": canonical["items"][qid]["cluster"],
            "CTRL_text": row_by_condition["CTRL"]["question_text"],
            "BM_text": row_by_condition["BM"]["question_text"],
            "AN_text": row_by_condition["AN"]["question_text"],
            "clinical_qa_note": decision.get("reason", ""),
        })
    ws = workbook.create_sheet("validated_paired_review")
    write_table(ws, review_columns, review_rows, text_columns=("CTRL_text", "BM_text", "AN_text", "option_a", "option_b", "option_c", "option_d", "correct_option_text", "clinical_qa_note"))
    table_style(ws, {"A": 15, "B": 18, "G": 35, "H": 35, "I": 35, "J": 35, "L": 90, "M": 90, "N": 90})
    ws.print_area = "A1:A1"

    candidate_columns = [
        "question_id", "region", "year", "exam_part", "cluster", "correct_letter",
        "strict_BM", "strict_AN", "strict_PAIR", "relaxed_BM", "relaxed_AN",
        "relaxed_PAIR", "release_status", "qa_note", "question_text", "option_a",
        "option_b", "option_c", "option_d", "correct_option_text", "source_key",
        "page_in_exam_pdf", "source_exam_pdf", "source_answer_key_pdf",
    ]
    strict = canonical["tiers"]["strict"]
    relaxed = canonical["tiers"]["relaxed"]
    generated = set(canonical["generated"])
    candidate_rows = []
    for source in source_rows:
        qid = source["question_id"]
        if qid not in generated:
            continue
        decision = adjudication["decisions"].get(qid, {})
        candidate_rows.append({
            **source,
            "cluster": canonical["items"][qid]["cluster"],
            "strict_BM": "yes" if qid in strict["BM"] else "no",
            "strict_AN": "yes" if qid in strict["AN"] else "no",
            "strict_PAIR": "yes" if qid in strict["PAIR"] else "no",
            "relaxed_BM": "yes" if qid in relaxed["BM"] else "no",
            "relaxed_AN": "yes" if qid in relaxed["AN"] else "no",
            "relaxed_PAIR": "yes" if qid in relaxed["PAIR"] else "no",
            "release_status": "RELEASED" if qid in release_ids else "NOT_RELEASED",
            "qa_note": decision.get("reason", "Exploratory candidate; not item-by-item cleared for release"),
        })
    ws = workbook.create_sheet("candidate_audit")
    write_table(ws, candidate_columns, candidate_rows, text_columns=("qa_note", "question_text", "option_a", "option_b", "option_c", "option_d", "correct_option_text"))
    table_style(ws, {"A": 14, "B": 18, "N": 70, "O": 90, "P": 35, "Q": 35, "R": 35, "S": 35, "T": 35})
    ws.print_area = "A1:A1"

    verdict_columns = [
        "question_id", "region", "year", "exam_part", "mechanical_guard",
        "strict_guard", "relaxed_guard", "BM_gate_ok", "AN_gate_ok", "cluster",
        "question_text", "option_a", "option_b", "option_c", "option_d",
        "correct_letter", "correct_option_text", "source_key",
    ]
    verdict_rows = []
    for source in source_rows:
        qid = source["question_id"]
        item = canonical["items"][qid]
        verdict_rows.append({
            **source,
            "mechanical_guard": item.get("mechanical_guard", ""),
            "strict_guard": item.get("strict_guard", ""),
            "relaxed_guard": item.get("relaxed_guard", ""),
            "BM_gate_ok": item.get("BM_gate_ok", ""),
            "AN_gate_ok": item.get("AN_gate_ok", ""),
            "cluster": item.get("cluster", ""),
        })
    ws = workbook.create_sheet("all_474_verdict")
    write_table(ws, verdict_columns, verdict_rows, text_columns=("question_text", "option_a", "option_b", "option_c", "option_d", "correct_option_text"))
    table_style(ws, {"A": 14, "B": 18, "K": 90, "L": 35, "M": 35, "N": 35, "O": 35, "Q": 35})
    ws.print_area = "A1:A1"

    ws = workbook.create_sheet("qa_findings")
    qa_columns = ["severity", "scope", "finding", "affected_ids", "resolution"]
    write_table(ws, qa_columns, adjudication["qa_findings"], text_columns=("finding", "affected_ids", "resolution"))
    qa_widths = {"A": 10, "B": 16, "C": 52, "D": 32, "E": 54}
    table_style(ws, qa_widths)
    for row_number in range(2, ws.max_row + 1):
        heights = []
        for column_number, column_letter in enumerate(("A", "B", "C", "D", "E"), 1):
            cell = ws.cell(row_number, column_number)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            heights.append(print_height(cell.value, chars_per_line=max(12, int(qa_widths[column_letter] * 0.9)), line_height=13, maximum=220))
        ws.row_dimensions[row_number].height = max(heights)
    ws.print_area = ws.dimensions

    sentence_rows = []
    for arm_key, arm_name in (("biomarker", "BM"), ("anatomy", "AN")):
        for variant, values in canonical["sentences"][arm_key].items():
            sentence_rows.append({
                "id": variant,
                "arm": arm_name,
                "entity": values[0],
                "sentence": values[1],
                "root": values[2],
                "primary": "yes" if variant == PRIMARY[arm_name][0] else "no",
            })
    ws = workbook.create_sheet("sentences")
    write_table(ws, ["id", "arm", "entity", "sentence", "root", "primary"], sentence_rows, text_columns=("sentence",))
    table_style(ws, {"A": 9, "B": 8, "C": 24, "D": 85, "E": 15, "F": 10})
    ws.print_area = ws.dimensions

    add_review_cards(workbook, rows)
    workbook.active = 0
    master = args.out_dir / "expC-items-final.xlsx"
    atomic_save_workbook(workbook, master)
    for condition, label in (("CTRL", "control"), ("BM", "biomarker"), ("AN", "anatomy")):
        save_import(args.out_dir / f"expC-{label}-final.xlsx", rows, condition)

    outputs = sorted(set(args.out_dir.glob("expC-*-final.xlsx")))
    manifest = {
        "source": args.source.name,
        "source_sha256": source_sha256,
        "canonical": args.canonical.name,
        "canonical_sha256": canonical_sha256,
        "adjudication": args.adjudication.name,
        "adjudication_sha256": adjudication_sha256,
        "builder": Path(__file__).name,
        "builder_sha256": builder_sha256,
        "base_questions": release_ids,
        "counts": {"base": len(release_ids), "conditions": 3, "master_rows": len(rows)},
        "primary_sentences": {key: {"variant": value[0], "entity": value[1], "sentence": value[2]} for key, value in PRIMARY.items()},
        "outputs": {path.name: hashlib.sha256(path.read_bytes()).hexdigest() for path in outputs},
    }
    atomic_write_text(args.out_dir / "final-manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))
    print(json.dumps(manifest, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
