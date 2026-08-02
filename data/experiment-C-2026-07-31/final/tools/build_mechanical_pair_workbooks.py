from __future__ import annotations

import argparse
import hashlib
import json
import os
import tempfile
import zipfile
from itertools import zip_longest
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.properties import PageSetupProperties


FONT = "Arial"
NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "E2F0D9"
AMBER = "FFF2CC"
GREY = "E7E6E6"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="B7C9D6")


def require(condition: bool, message: str):
    if not condition:
        raise ValueError(message)


def sha256_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def read_source(path: Path) -> tuple[list[str], list[dict]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows = list(workbook["questions"].iter_rows(values_only=True))
    header = [str(value) for value in rows[0]]
    records = [dict(zip(header, row)) for row in rows[1:] if row[0] is not None]
    require(len(records) == 474, f"expected 474 source rows, found {len(records)}")
    require(len({row['question_id'] for row in records}) == 474, "source question IDs are not unique")
    return header, records


def insertion_delta(control: str, altered: str, sentence: str) -> tuple[int, str]:
    require(control != altered, "control and alteration are identical")
    require(altered.count(sentence) == 1, "inserted sentence does not occur exactly once")
    start = altered.index(sentence)
    end = start + len(sentence)
    candidates = []
    for stop in range(end, min(len(altered), end + 12) + 1):
        suffix = altered[end:stop]
        if suffix and not suffix.isspace():
            break
        if altered[:start] + altered[stop:] == control:
            candidates.append((start, altered[start:stop]))
    require(len(candidates) == 1, "alteration is not exactly one sentence-plus-spacing insertion")
    offset, inserted = candidates[0]
    require(inserted.startswith(sentence), "inserted segment does not start with the assigned sentence")
    require(not inserted[len(sentence):] or inserted[len(sentence):].isspace(), "non-whitespace text follows inserted sentence")
    return offset, inserted


def set_font(cell, *, bold=False, color="000000", size=9):
    cell.font = Font(name=FONT, size=size, bold=bold, color=color)


def style_table(ws, widths: dict[str, float], text_headers: set[str]):
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        set_font(cell, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN)
    ws.row_dimensions[1].height = 36
    header_index = {cell.value: cell.column for cell in ws[1]}
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            set_font(cell)
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column in {header_index.get(name) for name in text_headers})
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.zoomScale = 80
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True, autoPageBreaks=False)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:1"
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.4, bottom=0.4, header=0.2, footer=0.2)


def append_rows(ws, columns: list[str], records: list[dict]):
    ws.append(columns)
    for record in records:
        ws.append([record.get(column, "") for column in columns])
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                cell.data_type = "s"


def chunks(value: str, limit=650) -> list[str]:
    parts = []
    for paragraph in str(value).splitlines():
        paragraph = paragraph.strip()
        if not paragraph:
            continue
        while len(paragraph) > limit:
            cut = max(paragraph.rfind(". ", 0, limit), paragraph.rfind("; ", 0, limit), paragraph.rfind(" ", 0, limit))
            cut = limit if cut < limit // 2 else cut + 1
            parts.append(paragraph[:cut].strip())
            paragraph = paragraph[cut:].strip()
        if paragraph:
            parts.append(paragraph)
    return parts or [""]


def add_readme(workbook: Workbook, arm: str, count: int, source: Path, canonical: Path):
    ws = workbook.create_sheet("README")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 31
    ws.column_dimensions["B"].width = 105
    label = "biomarker" if arm == "BM" else "anatomy"
    rows = [
        (f"Experiment C — {label} mechanical pool", ""),
        ("Contents", f"{count} mechanically accepted {label} candidates, each shown as exact control text beside its deterministic alteration."),
        ("Text guarantee", "The control is byte-for-byte identical to balanced-flat-A.xlsx. The alteration is the control plus exactly one assigned literal sentence and a whitespace separator. No original character is deleted, replaced, or reordered."),
        ("Unchanged fields", "Options A–D, correct letter, correct option text, source key, PDF provenance, and all other source metadata remain unchanged."),
        ("Selection status", "Exploratory mechanical pool. These rows passed deterministic seam and arm-specific guards; they were not all clinically adjudicated as answer-key-preserving."),
        ("Relation to final 16", "The conservative final release remains 16 base questions across CTRL/BM/AN. This companion workbook deliberately restores the wider mechanical candidate pool for transparent review."),
        ("Sentence assignment", "The draft's deterministic rank-within-cluster rotation is preserved. The assigned variant, fabricated entity, and exact inserted sentence appear in every comparison row."),
        ("Primary review sheet", "control_vs_alteration shows the two texts side by side. paired_cards provides complete untruncated reading blocks for long vignettes."),
        ("Data sheets", "control_rows and alteration_rows preserve every source column. In alteration_rows, question_text is the only source column whose value differs."),
        ("Source workbook", source.name),
        ("Source SHA-256", hashlib.sha256(source.read_bytes()).hexdigest()),
        ("Canonical JSON", canonical.name),
        ("Canonical SHA-256", hashlib.sha256(canonical.read_bytes()).hexdigest()),
    ]
    for row_number, (name, value) in enumerate(rows, 1):
        ws.cell(row_number, 1, name)
        ws.cell(row_number, 2, value)
        set_font(ws.cell(row_number, 1), bold=True, color=NAVY, size=15 if row_number == 1 else 10)
        set_font(ws.cell(row_number, 2), size=10)
        ws.cell(row_number, 1).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row_number, 2).alignment = Alignment(vertical="top", wrap_text=True)
        if row_number == 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
            ws.cell(1, 1).fill = PatternFill("solid", fgColor=BLUE)
            ws.row_dimensions[1].height = 28
        else:
            ws.cell(row_number, 1).fill = PatternFill("solid", fgColor=GREY)
            ws.row_dimensions[row_number].height = 42 if len(str(value)) > 120 else 27


def add_cards(workbook: Workbook, records: list[dict], arm: str):
    ws = workbook.create_sheet("paired_cards")
    ws.sheet_view.showGridLines = False
    for column, width in zip("ABCDEFGH", [5, 20, 20, 20, 20, 20, 20, 20]):
        ws.column_dimensions[column].width = width
    row_number = 1
    for index, record in enumerate(records):
        if index:
            ws.row_breaks.append(Break(id=row_number - 1))
        ws.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=8)
        title = ws.cell(row_number, 1, f"{record['base_question_id']} | {record['region']} {record['year']} | {record['variant_id']} | key {str(record['correct_letter']).upper()}")
        set_font(title, bold=True, color=WHITE, size=11)
        title.fill = PatternFill("solid", fgColor=NAVY)
        row_number += 1
        ws.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=8)
        sentence = ws.cell(row_number, 1, f"Inserted string: {record['inserted_sentence']}")
        set_font(sentence, bold=True, color="375623", size=10)
        sentence.fill = PatternFill("solid", fgColor=GREEN)
        sentence.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row_number].height = 30
        row_number += 1
        ws.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=4)
        ws.merge_cells(start_row=row_number, start_column=5, end_row=row_number, end_column=8)
        for column, value in ((1, "CONTROL — exact source"), (5, f"ALTERATION — {arm} insertion only")):
            cell = ws.cell(row_number, column, value)
            set_font(cell, bold=True, color=NAVY, size=10)
            cell.fill = PatternFill("solid", fgColor=BLUE if column == 1 else AMBER)
            cell.alignment = Alignment(horizontal="center")
        row_number += 1
        control_chunks = chunks(record["control_question_text"])
        altered_chunks = chunks(record["altered_question_text"])
        for left, right in zip_longest(control_chunks, altered_chunks, fillvalue=""):
            ws.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=4)
            ws.merge_cells(start_row=row_number, start_column=5, end_row=row_number, end_column=8)
            for column, value in ((1, left), (5, right)):
                cell = ws.cell(row_number, column, value)
                set_font(cell, size=9)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            longest = max(len(left), len(right))
            ws.row_dimensions[row_number].height = min(390, max(30, 18 + 13 * ((longest // 90) + 1)))
            row_number += 1
        ws.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=8)
        qa = ws.cell(row_number, 1, "QA: PASS — exactly one insertion; no deletion/replacement; options and answer key unchanged")
        set_font(qa, bold=True, color="375623", size=9)
        qa.fill = PatternFill("solid", fgColor=GREEN)
        row_number += 2
    ws.print_area = f"A1:H{row_number - 1}"
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True, autoPageBreaks=False)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.35, bottom=0.35, header=0.15, footer=0.15)


def atomic_save(workbook: Workbook, path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    handle = tempfile.NamedTemporaryFile(prefix=f".{path.stem}.", suffix=".xlsx", dir=path.parent, delete=False)
    temporary = Path(handle.name)
    handle.close()
    try:
        workbook.save(temporary)
        with zipfile.ZipFile(temporary) as archive:
            require(archive.testzip() is None, f"{path.name}: corrupt OOXML archive")
        os.replace(temporary, path)
    finally:
        if temporary.exists():
            temporary.unlink()


def build_workbook(path: Path, arm: str, source_header: list[str], source_rows: list[dict], canonical: dict, source: Path, canonical_path: Path):
    tier_ids = canonical["tiers"]["relaxed"][arm]
    expected = 99 if arm == "BM" else 86
    require(len(tier_ids) == expected, f"{arm}: expected {expected} candidates, found {len(tier_ids)}")
    source_by_id = {row["question_id"]: row for row in source_rows}
    records = []
    for qid in tier_ids:
        source_row = source_by_id[qid]
        generated = canonical["generated"][qid]
        control = str(source_row["question_text"])
        altered = str(generated[f"{arm}_text"])
        sentence = str(generated[f"{arm}_sentence"])
        require(generated["CTRL"] == control, f"{qid}: canonical control differs from source")
        offset, inserted = insertion_delta(control, altered, sentence)
        for letter in "abcd":
            require(source_row[f"option_{letter}"] is not None, f"{qid}: empty option {letter}")
        correct_letter = str(source_row["correct_letter"]).lower()
        require(source_row["correct_option_text"] == source_row[f"option_{correct_letter}"], f"{qid}: answer key mismatch")
        item = canonical["items"][qid]
        records.append({
            **source_row,
            "base_question_id": qid,
            "condition": arm,
            "variant_id": generated[f"{arm}_variant"],
            "fabricated_entity": generated[f"{arm}_entity"],
            "inserted_sentence": sentence,
            "control_question_text": control,
            "altered_question_text": altered,
            "text_delta_qa": "PASS — insertion only",
            "inserted_character_offset": offset,
            "inserted_segment_repr": repr(inserted),
            "control_text_sha256": sha256_text(control),
            "altered_text_sha256": sha256_text(altered),
            "strict_tier_member": "yes" if qid in canonical["tiers"]["strict"][arm] else "no",
            "relaxed_pair_member": "yes" if qid in canonical["tiers"]["relaxed"]["PAIR"] else "no",
            "cluster": item["cluster"],
        })

    workbook = Workbook()
    comparison = workbook.active
    comparison.title = "control_vs_alteration"
    comparison_columns = [
        "base_question_id", "region", "year", "specialty", "exam_part", "question_number",
        "condition", "variant_id", "fabricated_entity", "inserted_sentence", "control_question_text",
        "altered_question_text", "text_delta_qa", "inserted_character_offset", "inserted_segment_repr",
        "strict_tier_member", "relaxed_pair_member", "cluster", "option_a", "option_b", "option_c",
        "option_d", "correct_letter", "correct_option_text", "flags", "page_in_exam_pdf",
        "source_exam_pdf", "source_answer_key_pdf", "source_key", "content_sha256",
        "control_text_sha256", "altered_text_sha256",
    ]
    append_rows(comparison, comparison_columns, records)
    style_table(
        comparison,
        {"A": 14, "B": 18, "H": 10, "I": 24, "J": 62, "K": 90, "L": 90, "M": 22,
         "O": 30, "R": 28, "S": 36, "T": 36, "U": 36, "V": 36, "X": 40, "AA": 45,
         "AB": 45, "AC": 40, "AD": 67, "AE": 67, "AF": 67},
        {"inserted_sentence", "control_question_text", "altered_question_text", "inserted_segment_repr",
         "option_a", "option_b", "option_c", "option_d", "correct_option_text", "source_exam_pdf",
         "source_answer_key_pdf"},
    )
    for row_number in range(2, comparison.max_row + 1):
        comparison.cell(row_number, 13).fill = PatternFill("solid", fgColor=GREEN)

    for sheet_name, text_column in (("control_rows", "control_question_text"), ("alteration_rows", "altered_question_text")):
        ws = workbook.create_sheet(sheet_name)
        columns = source_header + ["condition", "variant_id", "fabricated_entity", "inserted_sentence", "text_delta_qa"]
        sheet_records = []
        for record in records:
            row = {column: record.get(column, "") for column in source_header}
            row["question_text"] = record[text_column]
            row.update({key: record[key] for key in ("condition", "variant_id", "fabricated_entity", "inserted_sentence", "text_delta_qa")})
            sheet_records.append(row)
        append_rows(ws, columns, sheet_records)
        style_table(ws, {"A": 14, "B": 18, "G": 90, "H": 36, "I": 36, "J": 36, "K": 36, "M": 40, "Y": 60}, {"question_text", "option_a", "option_b", "option_c", "option_d", "correct_option_text", "inserted_sentence"})

    add_readme(workbook, arm, len(records), source, canonical_path)
    add_cards(workbook, records, arm)
    workbook.active = 0
    atomic_save(workbook, path)
    return records


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, required=True)
    parser.add_argument("--canonical", type=Path, required=True)
    parser.add_argument("--out-dir", type=Path, required=True)
    args = parser.parse_args()
    canonical = json.loads(args.canonical.read_text(encoding="utf-8"))
    source_header, source_rows = read_source(args.source)
    require(canonical["n_source_rows"] == len(source_rows), "source/canonical row count mismatch")
    outputs = {}
    names = {
        "BM": "expC-biomarker-mechanical-99-control-vs-alteration.xlsx",
        "AN": "expC-anatomy-mechanical-86-control-vs-alteration.xlsx",
    }
    for arm, name in names.items():
        records = build_workbook(args.out_dir / name, arm, source_header, source_rows, canonical, args.source, args.canonical)
        outputs[name] = {
            "arm": arm,
            "rows": len(records),
            "sha256": hashlib.sha256((args.out_dir / name).read_bytes()).hexdigest(),
            "all_insert_only": all(record["text_delta_qa"].startswith("PASS") for record in records),
        }
    print(json.dumps(outputs, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
