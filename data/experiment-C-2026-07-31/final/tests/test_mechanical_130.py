from __future__ import annotations

import json
import sys
import unittest
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[3]
TOOLS = ROOT / "experiment-C-2026-07-31" / "final" / "tools"
sys.path.insert(0, str(TOOLS))

from mechanical_130_pipeline import insert_only, normalize, suggested_seams  # noqa: E402


class Mechanical130Tests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        source = ROOT / "experiment-31-07-26" / "balanced-flat-A.xlsx"
        workbook = load_workbook(source, read_only=True, data_only=True)
        rows = list(workbook["questions"].iter_rows(values_only=True))
        workbook.close()
        header = list(rows[0])
        cls.source = {row[0]: dict(zip(header, row)) for row in rows[1:] if row[0] is not None}

    def test_insert_only_preserves_source(self):
        control = "Paciente estable. \u00bfCu\u00e1l es la respuesta?"
        sentence = "La fibroquelina-X3 s\u00e9rica se encuentra aumentada."
        start, end = 17, 18
        altered, separator = insert_only(control, sentence, start, end)
        self.assertEqual(separator, " ")
        self.assertEqual(altered, "Paciente estable. La fibroquelina-X3 s\u00e9rica se encuentra aumentada. \u00bfCu\u00e1l es la respuesta?")

    def test_b360_offers_prior_question_seam_not_terminal_colon(self):
        text = self.source["b360"]["question_text"]
        seams = suggested_seams(text)
        suffixes = [text[item["boundary_end"] :] for item in seams]
        self.assertTrue(any(value.startswith("¿Cuál") for value in suffixes))
        self.assertFalse(any(value == ":" for value in suffixes))

    def test_b298_offers_question_seam(self):
        text = self.source["b298"]["question_text"]
        seams = suggested_seams(text)
        self.assertTrue(any(text[item["boundary_end"] :].startswith("¿Cuál") for item in seams))

    def test_review_packets_cover_missing_pairs_once(self):
        qa_dir = ROOT / "experiment-C-2026-07-31" / "final" / "qa" / "mechanical-130"
        records = []
        for path in sorted((qa_dir / "review_packets").glob("review_*.json")):
            records.extend(json.loads(path.read_text(encoding="utf-8"))["candidates"])
        keys = [(item["question_id"], item["arm"]) for item in records]
        self.assertEqual(len(keys), 763)
        self.assertEqual(len(set(keys)), 763)

    def test_existing_pools_have_no_normalized_text_duplicates(self):
        base = ROOT / "experiment-C-2026-07-31" / "final" / "outputs"
        for name in (
            "expC-biomarker-mechanical-99-control-vs-alteration.xlsx",
            "expC-anatomy-mechanical-86-control-vs-alteration.xlsx",
        ):
            workbook = load_workbook(base / name, read_only=True, data_only=True)
            rows = list(workbook["control_vs_alteration"].iter_rows(values_only=True))
            workbook.close()
            header = list(rows[0])
            records = [dict(zip(header, row)) for row in rows[1:] if row[0] is not None]
            texts = [normalize(item["control_question_text"]) for item in records]
            self.assertEqual(len(texts), len(set(texts)))


if __name__ == "__main__":
    unittest.main()
