#!/usr/bin/env python3
"""Builder D -- independent regression tests for the Experiment C
2-fake / 50-50 rebalanced baseline.

This suite deliberately does NOT import or call anything from
``build_baseline.py`` (or any other builder script). Every normalization
rule, camouflage check, insertion-contract recomputation, and hash is
re-derived here directly from the task specification and cross-checked
against the raw artifacts:

  - baseline.json                                  (canonical machine-readable output)
  - expC-*-2fake-5050-baseline.xlsx                 (produced Excel, one per arm)
  - balanced-flat-A.xlsx                            (ultimate ground truth: control/options/key)
  - the two LOCKED mechanical-130 input workbooks   (used only to independently
                                                      re-derive each row's insertion
                                                      point/separator -- never modified)

Run with real pytest (preferred):
    /Users/ernestsaenz/Programming/gift-project-compile/tier1_mcq/.venv/bin/python \
        -m pytest \
        /Users/ernestsaenz/Programming/gift-project-compile/tier1_mcq/data/experiment-C-2026-07-31/final/baseline-2fake-5050/tests/ \
        -q

Every test below is a bare ``test_*`` function using plain ``assert`` --
no fixtures, no ``pytest.mark``, no ``pytest`` import. That means this file
needs nothing beyond the standard library + openpyxl to be *executed*
directly too (``python test_baseline_2fake_5050.py``), which the
``__main__`` block at the bottom does with a tiny pytest-alike runner, for
environments where the ``pytest`` package itself is not installed. Pytest,
when available, will collect and run these exact same functions unchanged.
"""

from __future__ import annotations

import ast
import collections
import difflib
import functools
import glob
import hashlib
import json
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Paths (hardcoded from the task spec -- independent of build_baseline.py)
# ---------------------------------------------------------------------------

BASE_DIR = Path("/Users/ernestsaenz/Programming/gift-project-compile/tier1_mcq")
BASELINE_DIR = BASE_DIR / "data/experiment-C-2026-07-31/final/baseline-2fake-5050"
BASELINE_JSON_PATH = BASELINE_DIR / "baseline.json"
EXCEL_GLOB_PATTERN = str(BASELINE_DIR / "expC-*-2fake-5050-baseline.xlsx")

SOURCE_PATH = BASE_DIR / "data/experiment-31-07-26/balanced-flat-A.xlsx"

LOCKED_INPUT_PATH = {
    "BM": BASE_DIR
    / "data/experiment-C-2026-07-31/final/outputs"
    / "expC-biomarker-mechanical-100-plus-30-reserve-control-vs-alteration.xlsx",
    "AN": BASE_DIR
    / "data/experiment-C-2026-07-31/final/outputs"
    / "expC-anatomy-mechanical-100-plus-30-reserve-control-vs-alteration.xlsx",
}

ARMS = ("BM", "AN")

# The two kept fabricated entities per arm, per root, per sentence -- copied
# verbatim from the task specification (NOT read back out of baseline.json
# or build_baseline.py, so this suite can actually catch either of them
# getting it wrong).
KEPT_FAKE_SPEC = {
    "BM": {
        "fibroquelina-X3": {
            "root": "fibro",
            "sentence": "La fibroquelina-X3 sérica se encuentra aumentada.",
        },
        "colangiomirina-8": {
            "root": "colangi",
            "sentence": "La colangiomirina-8 plasmática se encuentra por encima del intervalo de referencia.",
        },
    },
    "AN": {
        "saco orfalónico": {
            "root": "saco",
            "sentence": "La exploración revela dolor a la palpación del saco orfalónico.",
        },
        "órgano liradónico": {
            "root": "organo",
            "sentence": "La exploración muestra sensibilidad localizada sobre el órgano liradónico.",
        },
    },
}
ROOTS = {arm: {e: v["root"] for e, v in ents.items()} for arm, ents in KEPT_FAKE_SPEC.items()}
SENTENCES = {arm: {e: v["sentence"] for e, v in ents.items()} for arm, ents in KEPT_FAKE_SPEC.items()}

PRIMARY_TARGET = 100
RESERVE_TARGET = 30
POOL_TARGET = PRIMARY_TARGET + RESERVE_TARGET
PRIMARY_PER_FAKE_TARGET = 50

PROVENANCE_FIELDS = (
    "region",
    "year",
    "specialty",
    "exam_part",
    "question_number",
    "flags",
    "page_in_exam_pdf",
    "source_exam_pdf",
    "source_answer_key_pdf",
    "source_key",
)

# The full column list the task spec declares "relevant" for the
# control_vs_alteration sheet -- the produced Excel must carry at least
# these (extra columns are fine and not checked).
EXCEL_REQUIRED_COLUMNS = (
    "pool_role",
    "pool_rank",
    "candidate_origin",
    "base_question_id",
    "region",
    "year",
    "specialty",
    "exam_part",
    "question_number",
    "condition",
    "variant_id",
    "fabricated_entity",
    "inserted_sentence",
    "control_question_text",
    "altered_question_text",
    "inserted_character_offset",
    "inserted_segment_repr",
    "cluster",
    "option_a",
    "option_b",
    "option_c",
    "option_d",
    "correct_letter",
    "correct_option_text",
    "flags",
    "page_in_exam_pdf",
    "source_exam_pdf",
    "source_answer_key_pdf",
    "source_key",
    "content_sha256",
    "control_text_sha256",
    "altered_text_sha256",
)

# Excel column name -> baseline.json field name, for the row-for-row check.
TEXT_FIELDS_EXCEL_TO_JSON = {
    "base_question_id": "base_question_id",
    "pool_role": "pool_role",
    "variant_id": "variant_id",
    "fabricated_entity": "fabricated_entity",
    "inserted_sentence": "inserted_sentence",
    "control_question_text": "control_question_text",
    "altered_question_text": "altered_question_text",
    "control_text_sha256": "control_text_sha256",
    "altered_text_sha256": "altered_text_sha256",
    "option_a": "option_a",
    "option_b": "option_b",
    "option_c": "option_c",
    "option_d": "option_d",
    "correct_letter": "correct_letter",
    "correct_option_text": "correct_option_text",
    "cluster": "cluster",
    "region": "region",
    "specialty": "specialty",
    "exam_part": "exam_part",
    "flags": "flags",
    "page_in_exam_pdf": "page_in_exam_pdf",
    "source_exam_pdf": "source_exam_pdf",
    "source_answer_key_pdf": "source_answer_key_pdf",
    "source_key": "source_key",
}
NUMERIC_FIELDS_EXCEL_TO_JSON = {
    "pool_rank": "pool_rank",
    "year": "year",
    "question_number": "question_number",
    "inserted_character_offset": "insertion_offset",
}


# ---------------------------------------------------------------------------
# Exact normalization / hashing helpers -- re-derived from the spec text,
# not imported from build_baseline.py.
# ---------------------------------------------------------------------------


def normalize(value: str) -> str:
    value = unicodedata.normalize("NFD", value or "")
    value = "".join(ch for ch in value if unicodedata.category(ch) != "Mn")
    return re.sub(r"\s+", " ", value).casefold()


def sha256_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def is_camouflage_legal(root: str, control_text: str) -> bool:
    return root not in normalize(control_text)


def _single_insert_span(control: str, altered: str):
    """Return (i, inserted_text) for the unique 'insert' opcode turning
    control into altered. Raises ValueError with a diagnostic message if the
    pair is not exactly one pure-insertion diff op."""
    matcher = difflib.SequenceMatcher(a=control, b=altered, autojunk=False)
    ops = [op for op in matcher.get_opcodes() if op[0] != "equal"]
    if len(ops) != 1:
        raise ValueError(f"expected exactly 1 non-equal diff op between control and altered, found {len(ops)}")
    tag, i1, i2, j1, j2 = ops[0]
    if tag != "insert":
        raise ValueError(f"expected the single diff op to be 'insert', found {tag!r}")
    if i1 != i2:
        raise ValueError(f"insert op has a nonzero source span ({i1},{i2}), expected a pure insertion")
    return i1, altered[j1:j2]


def _format_violations(violations: list, cap: int = 30) -> str:
    shown = violations[:cap]
    body = "\n".join(f"  - {v}" for v in shown)
    if len(violations) > cap:
        body += f"\n  ... and {len(violations) - cap} more"
    return f"{len(violations)} violation(s):\n{body}"


def _streq(a, b) -> bool:
    """String-equality with None/'' treated as equivalent, to absorb the
    well-known openpyxl round-trip quirk (empty string written -> None read
    back) without weakening real-content comparisons."""
    return (a if a is not None else "") == (b if b is not None else "")


def _as_int(v) -> int:
    return int(v)


# ---------------------------------------------------------------------------
# Loaders (cached -- pure I/O, no assertions with test semantics beyond
# "the file/sheet/column exists", which is itself something a test should be
# able to fail on with a clear message).
# ---------------------------------------------------------------------------


def _iter_data_rows(ws, key_column: str):
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(h) for h in next(rows_iter)]
    hidx = {h: i for i, h in enumerate(header)}
    if key_column not in hidx:
        raise AssertionError(f"sheet is missing expected key column {key_column!r} (found {header})")
    key_i = hidx[key_column]
    for row in rows_iter:
        if row[key_i] is None:
            continue
        yield dict(zip(header, row))


@functools.lru_cache(maxsize=None)
def _load_baseline_json() -> dict:
    with open(BASELINE_JSON_PATH, encoding="utf-8") as f:
        return json.load(f)


@functools.lru_cache(maxsize=None)
def _load_source_records() -> dict:
    wb = load_workbook(SOURCE_PATH, read_only=True, data_only=True)
    try:
        ws = wb["questions"]
        return {rec["question_id"]: rec for rec in _iter_data_rows(ws, "question_id")}
    finally:
        wb.close()


@functools.lru_cache(maxsize=None)
def _load_locked_input_by_id(arm: str) -> dict:
    wb = load_workbook(LOCKED_INPUT_PATH[arm], read_only=True, data_only=True)
    try:
        ws = wb["control_vs_alteration"]
        return {rec["base_question_id"]: rec for rec in _iter_data_rows(ws, "base_question_id")}
    finally:
        wb.close()


def _detect_arm_from_workbook(path: Path) -> str:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if "control_vs_alteration" not in wb.sheetnames:
            raise AssertionError(
                f"{path.name}: missing required sheet 'control_vs_alteration' (found {wb.sheetnames})"
            )
        ws = wb["control_vs_alteration"]
        records = list(_iter_data_rows(ws, "base_question_id"))
        if not records:
            raise AssertionError(f"{path.name}: 'control_vs_alteration' sheet has a header but no data rows")
        if "condition" not in records[0]:
            raise AssertionError(f"{path.name}: sheet has no 'condition' column to self-identify its arm")
        conditions = {rec["condition"] for rec in records}
        if len(conditions) != 1:
            raise AssertionError(
                f"{path.name}: mixed 'condition' values {conditions}, expected exactly one arm per workbook"
            )
        arm = next(iter(conditions))
        if arm not in ARMS:
            raise AssertionError(f"{path.name}: condition value {arm!r} is not one of {ARMS}")
        return arm
    finally:
        wb.close()


@functools.lru_cache(maxsize=None)
def _find_produced_excel_files() -> dict:
    paths = sorted(Path(p) for p in glob.glob(EXCEL_GLOB_PATTERN))
    if not paths:
        raise AssertionError(
            f"No produced Excel workbook found matching {EXCEL_GLOB_PATTERN!r}. Expected one "
            f"workbook per arm (e.g. expC-biomarker-2fake-5050-baseline.xlsx, "
            f"expC-anatomy-2fake-5050-baseline.xlsx) under {BASELINE_DIR}."
        )
    by_arm: dict = {}
    for path in paths:
        arm = _detect_arm_from_workbook(path)
        if arm in by_arm:
            raise AssertionError(
                f"Both {by_arm[arm]} and {path} self-identify (via their 'condition' column) as "
                f"arm {arm!r} -- expected exactly one produced workbook per arm."
            )
        by_arm[arm] = path
    missing = [a for a in ARMS if a not in by_arm]
    if missing:
        raise AssertionError(
            f"Produced Excel workbook(s) missing for arm(s) {missing}. Files matched by "
            f"{EXCEL_GLOB_PATTERN!r}: {[str(p) for p in paths]}; self-identified arms: "
            f"{ {a: str(p) for a, p in by_arm.items()} }."
        )
    return by_arm


@functools.lru_cache(maxsize=None)
def _load_excel_header(arm: str) -> tuple:
    path = _find_produced_excel_files()[arm]
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["control_vs_alteration"]
        header = next(ws.iter_rows(values_only=True))
        return tuple(str(h) for h in header)
    finally:
        wb.close()


@functools.lru_cache(maxsize=None)
def _load_excel_arm_rows(arm: str) -> tuple:
    path = _find_produced_excel_files()[arm]
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["control_vs_alteration"]
        return tuple(_iter_data_rows(ws, "base_question_id"))
    finally:
        wb.close()


def _iter_json_rows(baseline: dict, arm: str):
    for role in ("primary", "reserve"):
        for row in baseline["arms"][arm][role]:
            yield role, row


# ===========================================================================
# Tests -- baseline.json vs the task spec / balanced-flat-A.xlsx
# ===========================================================================


def test_source_workbook_has_expected_row_count():
    source = _load_source_records()
    assert len(source) == 474, (
        f"balanced-flat-A.xlsx 'questions' sheet has {len(source)} rows, expected 474"
    )


def test_baseline_json_declared_kept_fakes_match_the_task_spec():
    baseline = _load_baseline_json()
    violations = []
    for arm in ARMS:
        declared = {e["entity"]: e for e in baseline["kept_fakes"][arm]}
        if set(declared) != set(KEPT_FAKE_SPEC[arm]):
            violations.append(
                f"{arm}: baseline.json kept_fakes entities {sorted(declared)} != spec "
                f"{sorted(KEPT_FAKE_SPEC[arm])}"
            )
            continue
        for entity, spec in KEPT_FAKE_SPEC[arm].items():
            if declared[entity]["root"] != spec["root"]:
                violations.append(
                    f"{arm}/{entity}: baseline.json declared root {declared[entity]['root']!r} != "
                    f"spec root {spec['root']!r}"
                )
            if declared[entity]["sentence"] != spec["sentence"]:
                violations.append(f"{arm}/{entity}: baseline.json declared sentence differs from spec sentence")
    assert not violations, _format_violations(violations)


def test_each_arm_uses_exactly_the_two_kept_fabricated_entities():
    """Each arm uses EXACTLY the 2 specified fabricated_entity and nothing
    else, anywhere (primary+reserve)."""
    baseline = _load_baseline_json()
    violations = []
    for arm in ARMS:
        expected = set(KEPT_FAKE_SPEC[arm])
        seen = {row["fabricated_entity"] for _, row in _iter_json_rows(baseline, arm)}
        if seen != expected:
            violations.append(f"{arm}: fabricated_entity set is {sorted(seen)}, expected exactly {sorted(expected)}")
    assert not violations, _format_violations(violations)


def test_inserted_sentence_matches_the_kept_fake_sentence_exactly():
    baseline = _load_baseline_json()
    violations = []
    for arm in ARMS:
        for role, row in _iter_json_rows(baseline, arm):
            entity = row["fabricated_entity"]
            expected = SENTENCES[arm].get(entity)
            if expected is None:
                continue  # reported by test_each_arm_uses_exactly_the_two_kept_fabricated_entities
            if row["inserted_sentence"] != expected:
                violations.append(
                    f"{arm}/{role} {row['base_question_id']}: inserted_sentence != the spec sentence for "
                    f"{entity!r}"
                )
    assert not violations, _format_violations(violations)


def test_row_counts_are_130_per_arm_100_primary_30_reserve():
    baseline = _load_baseline_json()
    violations = []
    for arm in ARMS:
        primary = baseline["arms"][arm]["primary"]
        reserve = baseline["arms"][arm]["reserve"]
        if len(primary) != PRIMARY_TARGET:
            violations.append(f"{arm}: primary has {len(primary)} rows, expected {PRIMARY_TARGET}")
        if len(reserve) != RESERVE_TARGET:
            violations.append(f"{arm}: reserve has {len(reserve)} rows, expected {RESERVE_TARGET}")
        if len(primary) + len(reserve) != POOL_TARGET:
            violations.append(f"{arm}: total rows {len(primary) + len(reserve)} != {POOL_TARGET}")
        primary_ranks = sorted(r["pool_rank"] for r in primary)
        if primary_ranks != list(range(1, PRIMARY_TARGET + 1)):
            violations.append(f"{arm}: primary pool_rank values are not exactly 1..{PRIMARY_TARGET}")
        reserve_ranks = sorted(r["pool_rank"] for r in reserve)
        if reserve_ranks != list(range(1, RESERVE_TARGET + 1)):
            violations.append(f"{arm}: reserve pool_rank values are not exactly 1..{RESERVE_TARGET}")
        for row in primary:
            if row["pool_role"] != "PRIMARY":
                violations.append(f"{arm} {row['base_question_id']}: listed under primary[] but pool_role={row['pool_role']!r}")
        for row in reserve:
            if row["pool_role"] != "RESERVE":
                violations.append(f"{arm} {row['base_question_id']}: listed under reserve[] but pool_role={row['pool_role']!r}")
    assert not violations, _format_violations(violations)


def test_no_duplicate_base_question_id_within_arm():
    baseline = _load_baseline_json()
    violations = []
    for arm in ARMS:
        ids = [row["base_question_id"] for _, row in _iter_json_rows(baseline, arm)]
        counts = collections.Counter(ids)
        dupes = sorted(qid for qid, c in counts.items() if c > 1)
        if dupes:
            violations.append(f"{arm}: duplicate base_question_id(s): {dupes}")
    assert not violations, _format_violations(violations)


def test_no_duplicate_normalized_control_text_within_arm():
    baseline = _load_baseline_json()
    violations = []
    for arm in ARMS:
        texts = [normalize(row["control_question_text"]) for _, row in _iter_json_rows(baseline, arm)]
        counts = collections.Counter(texts)
        dupes = [t for t, c in counts.items() if c > 1]
        if dupes:
            violations.append(f"{arm}: {len(dupes)} normalized control-text value(s) appear more than once within the arm")
    assert not violations, _format_violations(violations)


def test_primary_is_50_50_or_gap_fully_explained_by_camouflage_forcing():
    """PRIMARY is EXACTLY 50/50 between the two fakes; if not, the gap must
    be fully explained by camouflage-forced rows (rows legal for only one of
    the two kept fakes), independently recomputed here."""
    baseline = _load_baseline_json()
    violations = []
    for arm in ARMS:
        entities = sorted(KEPT_FAKE_SPEC[arm])
        primary = baseline["arms"][arm]["primary"]
        actual_counts = collections.Counter(row["fabricated_entity"] for row in primary)

        declared_counts = baseline["arms"][arm]["primary_counts"]
        for e in entities:
            if declared_counts.get(e, 0) != actual_counts.get(e, 0):
                violations.append(
                    f"{arm}: declared primary_counts[{e!r}]={declared_counts.get(e, 0)} != actual tally "
                    f"{actual_counts.get(e, 0)}"
                )

        c0, c1 = actual_counts.get(entities[0], 0), actual_counts.get(entities[1], 0)
        if c0 == PRIMARY_PER_FAKE_TARGET and c1 == PRIMARY_PER_FAKE_TARGET:
            continue  # exact 50/50 achieved -- nothing further to justify

        # Not exact 50/50: independently recompute, from each row's own
        # control text, which rows are camouflage-legal for only one fake
        # ("forced"), and verify any excess above 50 is fully covered by
        # forced rows for that fake.
        forced_to = collections.Counter()
        for row in primary:
            norm_control = normalize(row["control_question_text"])
            legal = [e for e in entities if ROOTS[arm][e] not in norm_control]
            if len(legal) == 0:
                violations.append(
                    f"{arm} {row['base_question_id']}: camouflage-illegal for BOTH kept fakes -- should "
                    f"never happen in a valid PRIMARY row"
                )
            elif len(legal) == 1:
                forced_to[legal[0]] += 1

        for e in entities:
            over = actual_counts.get(e, 0) - PRIMARY_PER_FAKE_TARGET
            if over > 0 and forced_to[e] < over:
                violations.append(
                    f"{arm}: primary is not 50/50 ({dict(actual_counts)}) and the {over}-row excess for "
                    f"{e!r} is NOT fully explained by camouflage forcing (only {forced_to[e]} primary rows "
                    f"are camouflage-legal for {e!r} ALONE)"
                )
    assert not violations, _format_violations(violations)


def test_camouflage_rule_holds_for_every_row():
    """A fake may be assigned to a row only if its root is NOT a substring
    of normalize(control_question_text) -- checked for every row, primary
    and reserve, both arms, independently of the row's own declared
    camouflage_legal_entities (which is cross-checked too, when present)."""
    baseline = _load_baseline_json()
    violations = []
    for arm in ARMS:
        for role, row in _iter_json_rows(baseline, arm):
            label = f"{arm}/{role} {row['base_question_id']}"
            entity = row["fabricated_entity"]
            if entity not in ROOTS[arm]:
                violations.append(f"{label}: fabricated_entity {entity!r} is not one of the two kept fakes for {arm}")
                continue
            root = ROOTS[arm][entity]
            norm_control = normalize(row["control_question_text"])
            if root in norm_control:
                violations.append(
                    f"{label}: camouflage rule VIOLATED -- root {root!r} (of assigned fake {entity!r}) found "
                    f"in normalize(control_question_text)"
                )
            recomputed_legal = {e for e, r in ROOTS[arm].items() if r not in norm_control}
            declared_legal = set(row.get("camouflage_legal_entities") or [])
            if declared_legal and declared_legal != recomputed_legal:
                violations.append(
                    f"{label}: declared camouflage_legal_entities {sorted(declared_legal)} != independently "
                    f"recomputed {sorted(recomputed_legal)}"
                )
            if entity not in recomputed_legal:
                violations.append(f"{label}: assigned fake is not camouflage-legal per independent recomputation")
    assert not violations, _format_violations(violations)


def test_stored_control_altered_pair_self_consistent_and_hashes_match():
    """Lightweight, locked-input-independent re-verification of the exact
    insertion contract directly on the FINAL stored (control, altered)
    pair: exactly one insertion; prefix/suffix preserved; sentence occurs
    once; separator_repr agrees; both sha256 hashes match."""
    baseline = _load_baseline_json()
    violations = []
    for arm in ARMS:
        for role, row in _iter_json_rows(baseline, arm):
            label = f"{arm}/{role} {row['base_question_id']}"
            control = row["control_question_text"]
            altered = row["altered_question_text"]
            sentence = row["inserted_sentence"]
            try:
                i1, inserted_text = _single_insert_span(control, altered)
            except ValueError as exc:
                violations.append(f"{label}: {exc}")
                continue
            if i1 != row["insertion_offset"]:
                violations.append(
                    f"{label}: recomputed insertion index {i1} != stored insertion_offset {row['insertion_offset']}"
                )
            if not inserted_text.startswith(sentence):
                violations.append(f"{label}: inserted text does not start with inserted_sentence")
            else:
                sep = inserted_text[len(sentence):]
                if not sep.isspace():
                    violations.append(f"{label}: text after inserted_sentence, {sep!r}, is not pure whitespace")
                try:
                    declared_sep = ast.literal_eval(row["separator_repr"])
                except (ValueError, SyntaxError) as exc:
                    violations.append(f"{label}: separator_repr {row['separator_repr']!r} not a valid literal: {exc}")
                else:
                    if declared_sep != sep:
                        violations.append(
                            f"{label}: separator_repr decodes to {declared_sep!r}, recomputed SEP is {sep!r}"
                        )
            if altered.count(sentence) != 1:
                violations.append(
                    f"{label}: inserted_sentence occurs {altered.count(sentence)} time(s) in altered text, "
                    f"expected exactly 1"
                )
            if control[:i1] != altered[:i1]:
                violations.append(f"{label}: altered text prefix before the insertion point differs from control")
            if control[i1:] != altered[i1 + len(inserted_text):]:
                violations.append(f"{label}: altered text suffix after the insertion differs from control")
            if sha256_text(control) != row["control_text_sha256"]:
                violations.append(f"{label}: sha256(control_question_text) != control_text_sha256")
            if sha256_text(altered) != row["altered_text_sha256"]:
                violations.append(f"{label}: sha256(altered_question_text) != altered_text_sha256")
    assert not violations, _format_violations(violations)


def test_exact_insertion_contract_recomputed_from_locked_inputs():
    """Deep, fully first-principles re-derivation: recover (i, SEP) from
    the LOCKED INPUT workbook's own (control, old_altered) pair -- entirely
    independent of build_baseline.py or baseline.json's separator_repr --
    then rebuild the altered text from scratch using the row's NEW sentence
    and confirm it matches baseline.json byte-for-byte, offset-for-offset,
    hash-for-hash."""
    baseline = _load_baseline_json()
    violations = []
    for arm in ARMS:
        input_by_id = _load_locked_input_by_id(arm)
        for role, row in _iter_json_rows(baseline, arm):
            qid = row["base_question_id"]
            label = f"{arm}/{role} {qid}"
            inp = input_by_id.get(qid)
            if inp is None:
                violations.append(f"{label}: base_question_id not found in the locked input workbook")
                continue

            control = row["control_question_text"]
            if control != inp["control_question_text"]:
                violations.append(f"{label}: baseline.json control text differs from the locked input's control text")
                continue

            old_sentence = inp["inserted_sentence"]
            try:
                i_raw, inserted_text_input = _single_insert_span(control, inp["altered_question_text"])
            except ValueError as exc:
                violations.append(f"{label}: could not recover insertion from the locked input's own pair: {exc}")
                continue
            if not inserted_text_input.startswith(old_sentence):
                violations.append(
                    f"{label}: locked input's inserted text does not start with its own recorded inserted_sentence"
                )
                continue
            sep = inserted_text_input[len(old_sentence):]
            if not sep.isspace():
                violations.append(f"{label}: locked-input-recovered separator {sep!r} is not whitespace")
                continue

            if i_raw != inp["inserted_character_offset"]:
                violations.append(
                    f"{label}: locked-input-recovered insertion index {i_raw} != locked input's own "
                    f"inserted_character_offset {inp['inserted_character_offset']}"
                )
            if i_raw != row["insertion_offset"]:
                violations.append(
                    f"{label}: locked-input-recovered insertion index {i_raw} != baseline.json's "
                    f"insertion_offset {row['insertion_offset']} (the offset is a property of the control "
                    f"text and must be stable across which fake is inserted)"
                )

            new_sentence = row["inserted_sentence"]
            expected_altered = control[:i_raw] + new_sentence + sep + control[i_raw:]
            if expected_altered != row["altered_question_text"]:
                violations.append(
                    f"{label}: altered text independently rebuilt from the locked input's (i, SEP) + the "
                    f"row's NEW sentence does not match baseline.json's altered_question_text"
                )
                continue
            if expected_altered.count(new_sentence) != 1:
                violations.append(f"{label}: rebuilt altered text does not contain the new sentence exactly once")
            if sha256_text(control) != row["control_text_sha256"]:
                violations.append(f"{label}: recomputed sha256(control) != control_text_sha256")
            if sha256_text(expected_altered) != row["altered_text_sha256"]:
                violations.append(f"{label}: recomputed sha256(rebuilt altered) != altered_text_sha256")

            ops = [
                op
                for op in difflib.SequenceMatcher(a=control, b=expected_altered, autojunk=False).get_opcodes()
                if op[0] != "equal"
            ]
            if len(ops) != 1 or ops[0][0] != "insert":
                violations.append(f"{label}: rebuilt (control, altered) pair is not exactly one 'insert' op")
    assert not violations, _format_violations(violations, cap=40)


def test_control_options_and_key_byte_identical_to_source():
    baseline = _load_baseline_json()
    source = _load_source_records()
    violations = []
    for arm in ARMS:
        for role, row in _iter_json_rows(baseline, arm):
            qid = row["base_question_id"]
            label = f"{arm}/{role} {qid}"
            src = source.get(qid)
            if src is None:
                violations.append(f"{label}: base_question_id not present in balanced-flat-A.xlsx")
                continue
            if row["control_question_text"] != src["question_text"]:
                violations.append(f"{label}: control_question_text not byte-identical to source question_text")
            for opt in ("option_a", "option_b", "option_c", "option_d"):
                if row[opt] != src[opt]:
                    violations.append(f"{label}: {opt} not byte-identical to source")
            if row["correct_letter"] != src["correct_letter"]:
                violations.append(f"{label}: correct_letter not byte-identical to source")
            if row["correct_option_text"] != src["correct_option_text"]:
                violations.append(f"{label}: correct_option_text not byte-identical to source")
            if row[f"option_{row['correct_letter']}"] != row["correct_option_text"]:
                violations.append(f"{label}: row's correct_option_text != option_[correct_letter]")
            if src[f"option_{src['correct_letter']}"] != src["correct_option_text"]:
                violations.append(f"{label}: SOURCE's own correct_option_text != option_[correct_letter] (source data bug)")
    assert not violations, _format_violations(violations)


def test_provenance_fields_byte_identical_to_source():
    baseline = _load_baseline_json()
    source = _load_source_records()
    violations = []
    for arm in ARMS:
        for role, row in _iter_json_rows(baseline, arm):
            qid = row["base_question_id"]
            src = source.get(qid)
            if src is None:
                continue  # already reported by test_control_options_and_key_byte_identical_to_source
            for field in PROVENANCE_FIELDS:
                if row[field] != src[field]:
                    violations.append(
                        f"{arm}/{role} {qid}: provenance field {field!r} differs from source "
                        f"({row[field]!r} vs {src[field]!r})"
                    )
    assert not violations, _format_violations(violations)


# ===========================================================================
# Tests -- the produced Excel workbook(s)
# ===========================================================================


def test_produced_excel_files_exist_one_per_arm():
    by_arm = _find_produced_excel_files()
    assert set(by_arm) == set(ARMS), f"self-identified arms {sorted(by_arm)} != expected {sorted(ARMS)}"


def test_excel_has_required_columns():
    violations = []
    for arm in ARMS:
        header = set(_load_excel_header(arm))
        missing = [c for c in EXCEL_REQUIRED_COLUMNS if c not in header]
        if missing:
            violations.append(f"{arm}: produced Excel is missing required column(s): {missing}")
    assert not violations, _format_violations(violations)


def test_excel_row_counts_and_roles():
    violations = []
    for arm in ARMS:
        rows = _load_excel_arm_rows(arm)
        if len(rows) != POOL_TARGET:
            violations.append(f"{arm}: produced Excel has {len(rows)} data rows, expected {POOL_TARGET}")
        primary = [r for r in rows if r.get("pool_role") == "PRIMARY"]
        reserve = [r for r in rows if r.get("pool_role") == "RESERVE"]
        other = [r for r in rows if r.get("pool_role") not in ("PRIMARY", "RESERVE")]
        if len(primary) != PRIMARY_TARGET:
            violations.append(f"{arm}: produced Excel has {len(primary)} PRIMARY rows, expected {PRIMARY_TARGET}")
        if len(reserve) != RESERVE_TARGET:
            violations.append(f"{arm}: produced Excel has {len(reserve)} RESERVE rows, expected {RESERVE_TARGET}")
        if other:
            violations.append(f"{arm}: produced Excel has {len(other)} row(s) with an unexpected pool_role value")
        primary_ranks = sorted(_as_int(r["pool_rank"]) for r in primary)
        if primary_ranks != list(range(1, PRIMARY_TARGET + 1)):
            violations.append(f"{arm}: produced Excel PRIMARY pool_rank values are not exactly 1..{PRIMARY_TARGET}")
        reserve_ranks = sorted(_as_int(r["pool_rank"]) for r in reserve)
        if reserve_ranks != list(range(1, RESERVE_TARGET + 1)):
            violations.append(f"{arm}: produced Excel RESERVE pool_rank values are not exactly 1..{RESERVE_TARGET}")
        ids = [r["base_question_id"] for r in rows]
        id_counts = collections.Counter(ids)
        dupes = sorted(qid for qid, c in id_counts.items() if c > 1)
        if dupes:
            violations.append(f"{arm}: produced Excel has duplicate base_question_id(s): {dupes}")
    assert not violations, _format_violations(violations)


def test_excel_matches_baseline_json_row_for_row():
    """Excel control_vs_alteration matches baseline.json row-for-row.

    Matched by base_question_id (the unique key within an arm, verified
    elsewhere) rather than by raw physical row position, since physical
    ordering is not part of the documented contract -- but pool_role /
    pool_rank ARE compared as ordinary fields below, so any reordering that
    changes a row's role/rank is still caught.
    """
    baseline = _load_baseline_json()
    violations = []
    for arm in ARMS:
        json_by_id = {row["base_question_id"]: row for _, row in _iter_json_rows(baseline, arm)}
        excel_rows = _load_excel_arm_rows(arm)
        excel_by_id: dict = {}
        for erow in excel_rows:
            qid = erow.get("base_question_id")
            if qid in excel_by_id:
                continue  # duplicate -- already reported by test_excel_row_counts_and_roles
            excel_by_id[qid] = erow

        missing_in_excel = sorted(set(json_by_id) - set(excel_by_id))
        extra_in_excel = sorted(set(excel_by_id) - set(json_by_id))
        if missing_in_excel:
            violations.append(
                f"{arm}: {len(missing_in_excel)} baseline.json row(s) missing from the produced Excel: "
                f"{missing_in_excel[:10]}"
            )
        if extra_in_excel:
            violations.append(
                f"{arm}: {len(extra_in_excel)} produced-Excel row(s) not present in baseline.json: "
                f"{extra_in_excel[:10]}"
            )

        for qid in sorted(set(json_by_id) & set(excel_by_id)):
            jrow = json_by_id[qid]
            erow = excel_by_id[qid]
            label = f"{arm} {qid}"
            for ecol, jkey in TEXT_FIELDS_EXCEL_TO_JSON.items():
                if not _streq(jrow.get(jkey), erow.get(ecol)):
                    violations.append(
                        f"{label}: Excel column {ecol!r} != baseline.json {jkey!r} "
                        f"({erow.get(ecol)!r} vs {jrow.get(jkey)!r})"
                    )
            for ecol, jkey in NUMERIC_FIELDS_EXCEL_TO_JSON.items():
                jval, eval_ = jrow.get(jkey), erow.get(ecol)
                try:
                    match = int(eval_) == int(jval)
                except (TypeError, ValueError):
                    match = False
                if not match:
                    violations.append(f"{label}: Excel column {ecol!r}={eval_!r} != baseline.json {jkey!r}={jval!r}")
    assert not violations, _format_violations(violations, cap=40)


def test_excel_inserted_segment_repr_matches_its_own_control_altered_pair():
    """Bonus, fully first-principles check of a field baseline.json does
    not itself carry: inserted_segment_repr must decode to exactly the
    span the Excel row's own (control, altered) pair actually inserted,
    and inserted_character_offset must match the independently recomputed
    insertion index."""
    violations = []
    for arm in ARMS:
        for erow in _load_excel_arm_rows(arm):
            qid = erow.get("base_question_id")
            label = f"{arm} {qid}"
            repr_val = erow.get("inserted_segment_repr")
            control = erow.get("control_question_text") or ""
            altered = erow.get("altered_question_text") or ""
            try:
                i1, inserted_text = _single_insert_span(control, altered)
            except ValueError as exc:
                violations.append(f"{label}: {exc}")
                continue
            if repr_val is None:
                violations.append(f"{label}: inserted_segment_repr is empty")
                continue
            try:
                decoded = ast.literal_eval(repr_val)
            except (ValueError, SyntaxError) as exc:
                violations.append(f"{label}: inserted_segment_repr {repr_val!r} is not a valid Python literal: {exc}")
                continue
            if decoded != inserted_text:
                violations.append(
                    f"{label}: inserted_segment_repr decodes to {decoded!r}, expected the actual inserted "
                    f"span {inserted_text!r}"
                )
            offset = erow.get("inserted_character_offset")
            try:
                offset_ok = int(offset) == i1
            except (TypeError, ValueError):
                offset_ok = False
            if not offset_ok:
                violations.append(f"{label}: inserted_character_offset={offset!r} != recomputed insertion index {i1}")
    assert not violations, _format_violations(violations)


def test_excel_camouflage_rule_holds_for_every_row():
    """Same camouflage check as test_camouflage_rule_holds_for_every_row,
    run independently against the Excel's own stored text (defense in
    depth -- does not rely on the row-for-row equality test having caught
    a text-mangling bug introduced only during the xlsx write/read)."""
    violations = []
    for arm in ARMS:
        for erow in _load_excel_arm_rows(arm):
            qid = erow.get("base_question_id")
            label = f"{arm} {qid}"
            entity = erow.get("fabricated_entity")
            if entity not in ROOTS[arm]:
                violations.append(f"{label}: fabricated_entity {entity!r} is not one of the two kept fakes for {arm}")
                continue
            root = ROOTS[arm][entity]
            norm_control = normalize(erow.get("control_question_text") or "")
            if root in norm_control:
                violations.append(
                    f"{label}: camouflage rule VIOLATED in produced Excel -- root {root!r} (of {entity!r}) "
                    f"found in normalize(control_question_text)"
                )
    assert not violations, _format_violations(violations)


# ===========================================================================
# Minimal pytest-alike runner, for environments where the `pytest` package
# itself is not installed. Every function above is a normal callable with
# no pytest-specific decorators, so real pytest collects and runs them
# identically -- this block only exists to make `python thisfile.py` work
# too.
# ===========================================================================


def _all_test_functions():
    return [
        (name, obj)
        for name, obj in sorted(globals().items())
        if name.startswith("test_") and callable(obj)
    ]


def _main() -> int:
    tests = _all_test_functions()
    passed = 0
    failed = []
    errored = []
    for name, fn in tests:
        try:
            fn()
        except AssertionError as exc:
            failed.append((name, str(exc)))
        except Exception as exc:  # noqa: BLE001 - surface any error as a reported failure
            errored.append((name, f"{type(exc).__name__}: {exc}"))
        else:
            passed += 1
            print(f"PASSED  {name}")

    for name, msg in failed:
        print(f"\nFAILED  {name}")
        for line in (msg.splitlines() or [msg]):
            print(f"    {line}")
    for name, msg in errored:
        print(f"\nERROR   {name}")
        print(f"    {msg}")

    total = len(tests)
    total_failed = len(failed) + len(errored)
    print()
    print("=" * 78)
    print(
        f"{passed} passed, {total_failed} failed out of {total} total "
        f"(run via built-in fallback runner -- the `pytest` package is not installed in this "
        f"venv; every test_* function in this module was still executed and its assertions "
        f"checked exactly as real pytest would)."
    )
    return 0 if total_failed == 0 else 1


if __name__ == "__main__":
    raise SystemExit(_main())
