#!/usr/bin/env python3
"""Experiment C -- 2-fake / 50-50 rebalanced baseline builder.

Deterministic, reproducible assignment engine. Reads the two LOCKED
mechanical-130 workbooks (biomarker, anatomy) plus the ultimate source-of-truth
workbook (balanced-flat-A.xlsx), re-designates which 100 of each arm's 130
rows are PRIMARY vs RESERVE, assigns each row one of exactly two "kept"
fabricated entities per arm so that the 100 PRIMARY rows split EXACTLY 50/50,
recomputes the altered text for any row whose fabricated entity changed
(reusing the row's own previously-reviewed insertion point and whitespace
separator), self-asserts every invariant, and emits:

  - baseline.json          canonical machine-readable baseline
  - assignment_report.md   human-readable summary of what changed and why

This script performs NO network/model calls and never writes outside its own
output directory. It never modifies the locked input workbooks.

Run with:
  /Users/ernestsaenz/Programming/gift-project-compile/tier1_mcq/.venv/bin/python \
    /Users/ernestsaenz/Programming/gift-project-compile/tier1_mcq/data/experiment-C-2026-07-31/final/baseline-2fake-5050/build_baseline.py
"""

from __future__ import annotations

import ast
import collections
import difflib
import hashlib
import json
import re
import sys
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

# --------------------------------------------------------------------------
# Constants / paths
# --------------------------------------------------------------------------

PROTOCOL = "expc-2fake-5050-baseline-v1"
ARMS = ("BM", "AN")

BASE_DIR = Path("/Users/ernestsaenz/Programming/gift-project-compile/tier1_mcq")
INPUT_PATH = {
    "BM": BASE_DIR
    / "data/experiment-C-2026-07-31/final/outputs"
    / "expC-biomarker-mechanical-100-plus-30-reserve-control-vs-alteration.xlsx",
    "AN": BASE_DIR
    / "data/experiment-C-2026-07-31/final/outputs"
    / "expC-anatomy-mechanical-100-plus-30-reserve-control-vs-alteration.xlsx",
}
SOURCE_PATH = BASE_DIR / "data/experiment-31-07-26/balanced-flat-A.xlsx"

OUTPUT_DIR = BASE_DIR / "data/experiment-C-2026-07-31/final/baseline-2fake-5050"
OUTPUT_JSON = OUTPUT_DIR / "baseline.json"
OUTPUT_REPORT = OUTPUT_DIR / "assignment_report.md"

# The two kept fabricated entities per arm (every other fake is dropped).
KEPT_FAKES = {
    "BM": {
        "A": {
            "entity": "fibroquelina-X3",
            "root": "fibro",
            "sentence": "La fibroquelina-X3 sérica se encuentra aumentada.",
            "variant_id": "BM07",
        },
        "B": {
            "entity": "colangiomirina-8",
            "root": "colangi",
            "sentence": "La colangiomirina-8 plasmática se encuentra por encima del intervalo de referencia.",
            "variant_id": "BM02",
        },
    },
    "AN": {
        "A": {
            "entity": "saco orfalónico",
            "root": "saco",
            "sentence": "La exploración revela dolor a la palpación del saco orfalónico.",
            "variant_id": "AN04",
        },
        "B": {
            "entity": "órgano liradónico",
            "root": "organo",
            "sentence": "La exploración muestra sensibilidad localizada sobre el órgano liradónico.",
            "variant_id": "AN10",
        },
    },
}
LETTERS = ("A", "B")

PRIMARY_TARGET = 100
RESERVE_TARGET = 30
POOL_TARGET = PRIMARY_TARGET + RESERVE_TARGET
PRIMARY_PER_FAKE_TARGET = 50

# Provenance / byte-identity fields that MUST match balanced-flat-A.xlsx exactly.
PROVENANCE_FIELDS = (
    ("region", "region"),
    ("year", "year"),
    ("specialty", "specialty"),
    ("exam_part", "exam_part"),
    ("question_number", "question_number"),
    ("flags", "flags"),
    ("page_in_exam_pdf", "page_in_exam_pdf"),
    ("source_exam_pdf", "source_exam_pdf"),
    ("source_answer_key_pdf", "source_answer_key_pdf"),
    ("source_key", "source_key"),
)


class InvariantViolation(RuntimeError):
    """Raised (and fatal) when a hard invariant fails during construction."""


def require(condition: bool, message: str) -> None:
    if not condition:
        raise InvariantViolation(message)


# --------------------------------------------------------------------------
# Exact normalization / hashing helpers (must match spec EXACTLY)
# --------------------------------------------------------------------------


def normalize(value: str) -> str:
    value = unicodedata.normalize("NFD", value or "")
    value = "".join(ch for ch in value if unicodedata.category(ch) != "Mn")
    return re.sub(r"\s+", " ", value).casefold()


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def sha256_text(value: str) -> str:
    return sha256_bytes(value.encode("utf-8"))


def is_camouflage_legal(root: str, control_text: str) -> bool:
    return root not in normalize(control_text)


# --------------------------------------------------------------------------
# Workbook readers
# --------------------------------------------------------------------------


def read_source(path: Path) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["questions"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    header = [str(c) for c in rows[0]]
    records = {}
    for row in rows[1:]:
        if row[0] is None:
            continue
        rec = dict(zip(header, row))
        records[rec["question_id"]] = rec
    require(len(records) == 474, f"source: expected 474 rows, found {len(records)}")
    return records


def read_pool(path: Path) -> list:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["control_vs_alteration"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    header = [str(c) for c in rows[0]]
    records = []
    for row in rows[1:]:
        if row[0] is None:
            continue
        records.append(dict(zip(header, row)))
    return records


# --------------------------------------------------------------------------
# Exact-insertion contract: recovery from the existing pair + rebuild for a
# new sentence, per spec.
# --------------------------------------------------------------------------


def recover_insertion(control: str, altered: str, sentence: str, expected_offset, label: str):
    matcher = difflib.SequenceMatcher(a=control, b=altered, autojunk=False)
    ops = [op for op in matcher.get_opcodes() if op[0] != "equal"]
    require(len(ops) == 1, f"{label}: existing (control, altered) pair has {len(ops)} non-equal ops, expected 1")
    tag, i1, i2, j1, j2 = ops[0]
    require(tag == "insert", f"{label}: existing pair's single diff op is '{tag}', expected 'insert'")
    require(i1 == i2, f"{label}: existing pair insert op has nonzero source span ({i1},{i2})")
    inserted_text = altered[j1:j2]
    require(
        inserted_text.startswith(sentence),
        f"{label}: existing inserted text does not start with recorded inserted_sentence",
    )
    sep = inserted_text[len(sentence):]
    require(sep.isspace(), f"{label}: recovered separator {sep!r} is not whitespace")
    require(
        i1 == expected_offset,
        f"{label}: recovered insertion index {i1} != inserted_character_offset {expected_offset}",
    )
    return i1, sep


def build_altered(control: str, new_sentence: str, i: int, sep: str, label: str) -> str:
    require(0 <= i <= len(control), f"{label}: insertion index {i} out of bounds")
    altered = control[:i] + new_sentence + sep + control[i:]
    matcher = difflib.SequenceMatcher(a=control, b=altered, autojunk=False)
    ops = [op for op in matcher.get_opcodes() if op[0] != "equal"]
    require(len(ops) == 1, f"{label}: new (control, altered) pair has {len(ops)} non-equal ops, expected 1")
    tag, i1, i2, j1, j2 = ops[0]
    require(tag == "insert", f"{label}: new pair's single diff op is '{tag}', expected 'insert'")
    require(i1 == i2 == i, f"{label}: new insert position drifted ({i1},{i2}) != {i}")
    inserted = new_sentence + sep
    require(j2 - j1 == len(inserted), f"{label}: new inserted length mismatch")
    require(altered[:i] == control[:i], f"{label}: prefix changed vs control")
    require(altered[i + len(inserted):] == control[i:], f"{label}: suffix changed vs control")
    require(altered.count(new_sentence) == 1, f"{label}: new sentence does not occur exactly once")
    return altered


# --------------------------------------------------------------------------
# Phase 0: load + validate the 130-row pool for one arm against source
# --------------------------------------------------------------------------


def load_arm_rows(arm: str, source_records: dict) -> list:
    kept = KEPT_FAKES[arm]
    pool_records = read_pool(INPUT_PATH[arm])
    require(len(pool_records) == POOL_TARGET, f"{arm}: expected {POOL_TARGET} pool rows, found {len(pool_records)}")

    rows = []
    seen_ids = set()
    for rec in pool_records:
        qid = rec["base_question_id"]
        label = f"{arm} {qid}"
        require(qid not in seen_ids, f"{label}: duplicate base_question_id in locked input")
        seen_ids.add(qid)

        src = source_records.get(qid)
        require(src is not None, f"{label}: missing from source workbook (balanced-flat-A.xlsx)")

        control = src["question_text"]
        require(control == rec["control_question_text"], f"{label}: locked control text drifted from source")

        for field, src_field in PROVENANCE_FIELDS:
            require(rec[field] == src[src_field], f"{label}: locked '{field}' drifted from source")
        require(rec["content_sha256"] == src["content_sha256"], f"{label}: locked content_sha256 drifted from source")
        require(rec["option_a"] == src["option_a"], f"{label}: locked option_a drifted from source")
        require(rec["option_b"] == src["option_b"], f"{label}: locked option_b drifted from source")
        require(rec["option_c"] == src["option_c"], f"{label}: locked option_c drifted from source")
        require(rec["option_d"] == src["option_d"], f"{label}: locked option_d drifted from source")
        require(rec["correct_letter"] == src["correct_letter"], f"{label}: locked correct_letter drifted from source")
        require(
            rec["correct_option_text"] == src["correct_option_text"],
            f"{label}: locked correct_option_text drifted from source",
        )
        require(
            src[f"option_{src['correct_letter']}"] == src["correct_option_text"],
            f"{label}: source correct_option_text != option_[correct_letter]",
        )

        # Recover this row's own previously-reviewed insertion point / separator.
        i, sep = recover_insertion(
            control, rec["altered_question_text"], rec["inserted_sentence"], rec["inserted_character_offset"], label
        )

        legal = {letter for letter, info in kept.items() if is_camouflage_legal(info["root"], control)}
        require(len(legal) >= 1, f"{label}: illegal for BOTH kept fakes (camouflage rule) -- cannot be used")

        rows.append(
            {
                "base_question_id": qid,
                "orig_role": rec["pool_role"],
                "orig_rank": rec["pool_rank"],
                "orig_entity": rec["fabricated_entity"],
                "insertion_offset": i,
                "separator": sep,
                "control": control,
                "option_a": src["option_a"],
                "option_b": src["option_b"],
                "option_c": src["option_c"],
                "option_d": src["option_d"],
                "correct_letter": src["correct_letter"],
                "correct_option_text": src["correct_option_text"],
                "cluster": rec["cluster"],
                "region": src["region"],
                "year": src["year"],
                "specialty": src["specialty"],
                "exam_part": src["exam_part"],
                "question_number": src["question_number"],
                "flags": src["flags"],
                "page_in_exam_pdf": src["page_in_exam_pdf"],
                "source_exam_pdf": src["source_exam_pdf"],
                "source_answer_key_pdf": src["source_answer_key_pdf"],
                "source_key": src["source_key"],
                "legal": legal,
            }
        )

    require(len(rows) == POOL_TARGET, f"{arm}: pool row count drifted during load")
    require(
        sum(1 for r in rows if r["orig_role"] == "PRIMARY") == PRIMARY_TARGET,
        f"{arm}: expected {PRIMARY_TARGET} PRIMARY rows in locked input",
    )
    require(
        sum(1 for r in rows if r["orig_role"] == "RESERVE") == RESERVE_TARGET,
        f"{arm}: expected {RESERVE_TARGET} RESERVE rows in locked input",
    )
    return rows


def current_letter(row: dict, arm: str):
    kept = KEPT_FAKES[arm]
    for letter, info in kept.items():
        if row["orig_entity"] == info["entity"]:
            return letter
    return None


def tie_key(row: dict):
    # (currently-PRIMARY first, then pool_rank ascending, then base_question_id)
    return (0 if row["orig_role"] == "PRIMARY" else 1, row["orig_rank"], row["base_question_id"])


# --------------------------------------------------------------------------
# Spread-aware deterministic selection helpers
# --------------------------------------------------------------------------


def choose_flip_subset(candidates: list, k: int) -> list:
    """Choose exactly k of `candidates` to flip, spreading the chosen subset
    across clusters/regions via a deterministic greedy (min running count,
    ties broken by the fixed tie_key order)."""
    if k <= 0:
        return []
    remaining = sorted(candidates, key=tie_key)
    require(k <= len(remaining), f"cannot flip {k} rows out of only {len(remaining)} candidates")
    cluster_sel: collections.Counter = collections.Counter()
    region_sel: collections.Counter = collections.Counter()
    chosen = []
    for _ in range(k):
        best_idx = min(
            range(len(remaining)),
            key=lambda idx: (cluster_sel[remaining[idx]["cluster"]], region_sel[remaining[idx]["region"]]),
        )
        best = remaining.pop(best_idx)
        chosen.append(best)
        cluster_sel[best["cluster"]] += 1
        region_sel[best["region"]] += 1
    return chosen


def assign_free_rows(candidates: list, need: dict) -> dict:
    """Assign each row in `candidates` a letter from `need` (dict letter->count,
    must sum to len(candidates)), respecting per-row legality, spreading each
    letter's picks across clusters/regions via deterministic greedy."""
    total_need = sum(need.values())
    require(total_need == len(candidates), f"assign_free_rows: need sums to {total_need}, expected {len(candidates)}")
    pool = sorted(candidates, key=tie_key)
    remaining = dict(need)
    cluster_sel = {letter: collections.Counter() for letter in need}
    region_sel = {letter: collections.Counter() for letter in need}
    assignment = {}
    for row in pool:
        options = [letter for letter in LETTERS if letter in need and letter in row["legal"] and remaining[letter] > 0]
        require(options, f"{row['base_question_id']}: no legal+available fake left for free assignment")
        if len(options) == 1:
            choice = options[0]
        else:
            choice = min(
                options,
                key=lambda letter: (
                    cluster_sel[letter][row["cluster"]],
                    region_sel[letter][row["region"]],
                    -remaining[letter],
                ),
            )
        assignment[row["base_question_id"]] = choice
        remaining[choice] -= 1
        cluster_sel[choice][row["cluster"]] += 1
        region_sel[choice][row["region"]] += 1
    require(all(v == 0 for v in remaining.values()), f"assign_free_rows: leftover quota {remaining}")
    return assignment


# --------------------------------------------------------------------------
# Phase 1: determine final PRIMARY (100) / RESERVE (30) membership
# --------------------------------------------------------------------------


def determine_primary_set(rows: list):
    """Prefer the current PRIMARY/RESERVE partition untouched (0 role churn).
    Only swap rows in from RESERVE if the current PRIMARY set cannot possibly
    reach exact 50/50 (i.e. a forced-single-legal-fake side exceeds 50 within
    the current PRIMARY set)."""
    by_id = {r["base_question_id"]: r for r in rows}
    primary_ids = {r["base_question_id"] for r in rows if r["orig_role"] == "PRIMARY"}
    reserve_ids = {r["base_question_id"] for r in rows if r["orig_role"] == "RESERVE"}
    swap_count = 0
    infeasible_reason = None

    for letter, other_letter in (("A", "B"), ("B", "A")):
        group = [by_id[i] for i in primary_ids if by_id[i]["legal"] == {letter}]
        if len(group) > PRIMARY_PER_FAKE_TARGET:
            excess = len(group) - PRIMARY_PER_FAKE_TARGET
            evict = sorted(group, key=tie_key, reverse=True)[:excess]
            bring_in_pool = [by_id[i] for i in reserve_ids if other_letter in by_id[i]["legal"]]
            bring_in = sorted(bring_in_pool, key=tie_key)[:excess]
            n = min(excess, len(bring_in))
            for r in evict[:n]:
                primary_ids.discard(r["base_question_id"])
                reserve_ids.add(r["base_question_id"])
            for r in bring_in[:n]:
                reserve_ids.discard(r["base_question_id"])
                primary_ids.add(r["base_question_id"])
            swap_count += 2 * n
            if n < excess:
                infeasible_reason = (
                    f"forced-{letter}-only PRIMARY rows ({len(group)}) exceed the 50-row target and only "
                    f"{n} of {excess} needed RESERVE rows legal-for-{other_letter} were available to swap in"
                )

    require(len(primary_ids) == PRIMARY_TARGET, "primary size drifted during swap resolution")
    require(len(reserve_ids) == RESERVE_TARGET, "reserve size drifted during swap resolution")
    primary_rows = [by_id[i] for i in primary_ids]
    reserve_rows = [by_id[i] for i in reserve_ids]
    return primary_rows, reserve_rows, swap_count, infeasible_reason


# --------------------------------------------------------------------------
# Phase 2: fake assignment
# --------------------------------------------------------------------------


def assign_primary_fakes(primary_rows: list, arm: str, swap_infeasible_reason):
    keepA, keepB, forcedA, forcedB, fresh = [], [], [], [], []
    for r in primary_rows:
        cur = current_letter(r, arm)
        if r["legal"] == {"A"}:
            forcedA.append(r)
        elif r["legal"] == {"B"}:
            forcedB.append(r)
        elif cur == "A":
            keepA.append(r)
        elif cur == "B":
            keepB.append(r)
        else:
            fresh.append(r)

    target_a = PRIMARY_PER_FAKE_TARGET - len(forcedA)
    target_b = PRIMARY_PER_FAKE_TARGET - len(forcedB)
    exact_feasible = swap_infeasible_reason is None and target_a >= 0 and target_b >= 0
    reason = swap_infeasible_reason
    if target_a < 0:
        reason = reason or f"forced-A-only PRIMARY rows ({len(forcedA)}) alone exceed the 50-row target"
        target_a = 0
    if target_b < 0:
        reason = reason or f"forced-B-only PRIMARY rows ({len(forcedB)}) alone exceed the 50-row target"
        target_b = 0

    result = {}
    for r in forcedA:
        result[r["base_question_id"]] = "A"
    for r in forcedB:
        result[r["base_question_id"]] = "B"

    overflow_a = max(0, len(keepA) - target_a)
    overflow_b = max(0, len(keepB) - target_b)
    flipped_a_to_b = choose_flip_subset(keepA, overflow_a)
    flipped_b_to_a = choose_flip_subset(keepB, overflow_b)
    flipped_ids = {r["base_question_id"] for r in flipped_a_to_b} | {r["base_question_id"] for r in flipped_b_to_a}

    for r in keepA:
        if r["base_question_id"] not in flipped_ids:
            result[r["base_question_id"]] = "A"
    for r in keepB:
        if r["base_question_id"] not in flipped_ids:
            result[r["base_question_id"]] = "B"
    for r in flipped_a_to_b:
        result[r["base_question_id"]] = "B"
    for r in flipped_b_to_a:
        result[r["base_question_id"]] = "A"

    kept_a_final = len(keepA) - overflow_a
    kept_b_final = len(keepB) - overflow_b
    fresh_a_need = target_a - kept_a_final - len(flipped_b_to_a)
    fresh_b_need = target_b - kept_b_final - len(flipped_a_to_b)
    if fresh_a_need < 0 or fresh_b_need < 0 or fresh_a_need + fresh_b_need != len(fresh):
        exact_feasible = False
        reason = reason or "fresh-row quota arithmetic could not land on an exact 50/50 split"
        fresh_a_need = max(0, min(len(fresh), fresh_a_need))
        fresh_b_need = len(fresh) - fresh_a_need

    if fresh:
        result.update(assign_free_rows(fresh, {"A": fresh_a_need, "B": fresh_b_need}))

    require(len(result) == len(primary_rows), f"{arm}: primary fake assignment coverage mismatch")
    counts = collections.Counter(result.values())
    if counts["A"] != PRIMARY_PER_FAKE_TARGET or counts["B"] != PRIMARY_PER_FAKE_TARGET:
        exact_feasible = False
        reason = reason or f"resulting primary split was A={counts['A']} B={counts['B']}, not 50/50"

    diagnostics = {
        "forcedA": len(forcedA),
        "forcedB": len(forcedB),
        "keepA": len(keepA),
        "keepB": len(keepB),
        "fresh": len(fresh),
        "flipped_a_to_b": len(flipped_a_to_b),
        "flipped_b_to_a": len(flipped_b_to_a),
        "fresh_a_need": fresh_a_need,
        "fresh_b_need": fresh_b_need,
    }
    return result, exact_feasible, reason, diagnostics


def assign_reserve_fakes(reserve_rows: list, arm: str):
    keepA, keepB, forcedA, forcedB, fresh = [], [], [], [], []
    for r in reserve_rows:
        cur = current_letter(r, arm)
        if r["legal"] == {"A"}:
            forcedA.append(r)
        elif r["legal"] == {"B"}:
            forcedB.append(r)
        elif cur == "A":
            keepA.append(r)
        elif cur == "B":
            keepB.append(r)
        else:
            fresh.append(r)

    result = {}
    for r in forcedA:
        result[r["base_question_id"]] = "A"
    for r in forcedB:
        result[r["base_question_id"]] = "B"
    for r in keepA:
        result[r["base_question_id"]] = "A"
    for r in keepB:
        result[r["base_question_id"]] = "B"

    total_a_so_far = len(forcedA) + len(keepA)
    total_b_so_far = len(forcedB) + len(keepB)
    ideal_a = RESERVE_TARGET // 2
    fresh_a_need = max(0, min(len(fresh), ideal_a - total_a_so_far))
    fresh_b_need = len(fresh) - fresh_a_need

    if fresh:
        result.update(assign_free_rows(fresh, {"A": fresh_a_need, "B": fresh_b_need}))

    require(len(result) == len(reserve_rows), f"{arm}: reserve fake assignment coverage mismatch")
    diagnostics = {
        "forcedA": len(forcedA),
        "forcedB": len(forcedB),
        "keepA": len(keepA),
        "keepB": len(keepB),
        "fresh": len(fresh),
        "fresh_a_need": fresh_a_need,
        "fresh_b_need": fresh_b_need,
    }
    return result, diagnostics


# --------------------------------------------------------------------------
# Phase 3: build final ROW records
# --------------------------------------------------------------------------


def assign_ranks(rows_in_role: list, role: str) -> dict:
    """Rows that keep their original role keep their original rank; rows that
    moved into this role (role change) get fresh ranks appended after, in
    deterministic tie order. When zero rows move, this reproduces the
    original 1..N numbering exactly."""
    kept = [r for r in rows_in_role if r["orig_role"] == role]
    moved_in = [r for r in rows_in_role if r["orig_role"] != role]
    ranks = {}
    used = set()
    for r in kept:
        ranks[r["base_question_id"]] = r["orig_rank"]
        used.add(r["orig_rank"])
    next_rank = 1
    moved_in_sorted = sorted(moved_in, key=tie_key)
    for r in moved_in_sorted:
        while next_rank in used:
            next_rank += 1
        ranks[r["base_question_id"]] = next_rank
        used.add(next_rank)
    return ranks


def build_final_row(row: dict, arm: str, role: str, rank: int, letter: str) -> dict:
    kept = KEPT_FAKES[arm][letter]
    label = f"{arm} {row['base_question_id']}"
    new_sentence = kept["sentence"]
    altered = build_altered(row["control"], new_sentence, row["insertion_offset"], row["separator"], label)

    require(letter in row["legal"], f"{label}: assigned fake '{letter}' is camouflage-ILLEGAL for this control text")

    reassigned_from = row["orig_entity"] if row["orig_entity"] != kept["entity"] else None
    role_changed = role != row["orig_role"]

    legal_entities = [KEPT_FAKES[arm][l]["entity"] for l in LETTERS if l in row["legal"]]

    return {
        "base_question_id": row["base_question_id"],
        "pool_role": role,
        "pool_rank": rank,
        "variant_id": kept["variant_id"],
        "fabricated_entity": kept["entity"],
        "inserted_sentence": new_sentence,
        "insertion_offset": row["insertion_offset"],
        "separator_repr": repr(row["separator"]),
        "control_question_text": row["control"],
        "altered_question_text": altered,
        "control_text_sha256": sha256_text(row["control"]),
        "altered_text_sha256": sha256_text(altered),
        "option_a": row["option_a"],
        "option_b": row["option_b"],
        "option_c": row["option_c"],
        "option_d": row["option_d"],
        "correct_letter": row["correct_letter"],
        "correct_option_text": row["correct_option_text"],
        "cluster": row["cluster"],
        "region": row["region"],
        "year": row["year"],
        "specialty": row["specialty"],
        "exam_part": row["exam_part"],
        "question_number": row["question_number"],
        "flags": row["flags"],
        "page_in_exam_pdf": row["page_in_exam_pdf"],
        "source_exam_pdf": row["source_exam_pdf"],
        "source_answer_key_pdf": row["source_answer_key_pdf"],
        "source_key": row["source_key"],
        "reassigned_from": reassigned_from,
        "role_changed": role_changed,
        "camouflage_legal_entities": legal_entities,
    }


# --------------------------------------------------------------------------
# Phase 4: self-assertion of every invariant over the FINAL built rows
# --------------------------------------------------------------------------


def self_assert_arm(arm: str, primary: list, reserve: list, source_records: dict, violations: list):
    kept = KEPT_FAKES[arm]
    kept_entities = {info["entity"] for info in kept.values()}
    all_rows = primary + reserve

    if len(primary) != PRIMARY_TARGET:
        violations.append(f"{arm}: primary count is {len(primary)}, expected {PRIMARY_TARGET}")
    if len(reserve) != RESERVE_TARGET:
        violations.append(f"{arm}: reserve count is {len(reserve)}, expected {RESERVE_TARGET}")

    ids = [r["base_question_id"] for r in all_rows]
    if len(ids) != len(set(ids)):
        dupes = [qid for qid, cnt in collections.Counter(ids).items() if cnt > 1]
        violations.append(f"{arm}: duplicate base_question_id(s): {dupes}")

    norm_texts = [normalize(r["control_question_text"]) for r in all_rows]
    if len(norm_texts) != len(set(norm_texts)):
        dupes = [t for t, cnt in collections.Counter(norm_texts).items() if cnt > 1]
        violations.append(f"{arm}: {len(dupes)} duplicate normalized control text group(s) within arm")

    distinct_entities = {r["fabricated_entity"] for r in all_rows}
    if distinct_entities != kept_entities:
        violations.append(
            f"{arm}: final fabricated_entity set is {sorted(distinct_entities)}, expected exactly {sorted(kept_entities)}"
        )

    for row in all_rows:
        label = f"{arm} {row['base_question_id']}"
        src = source_records.get(row["base_question_id"])
        if src is None:
            violations.append(f"{label}: base_question_id not found in source workbook")
            continue
        if row["control_question_text"] != src["question_text"]:
            violations.append(f"{label}: control_question_text not byte-identical to source")
        for opt in ("option_a", "option_b", "option_c", "option_d"):
            if row[opt] != src[opt]:
                violations.append(f"{label}: {opt} not byte-identical to source")
        if row["correct_letter"] != src["correct_letter"]:
            violations.append(f"{label}: correct_letter not byte-identical to source")
        if row["correct_option_text"] != src["correct_option_text"]:
            violations.append(f"{label}: correct_option_text not byte-identical to source")
        for field, src_field in PROVENANCE_FIELDS:
            if row[field] != src[src_field]:
                violations.append(f"{label}: provenance field '{field}' not byte-identical to source")

        if row[f"option_{row['correct_letter']}"] != row["correct_option_text"]:
            violations.append(f"{label}: correct_option_text != option_[correct_letter]")

        if row["fabricated_entity"] not in kept_entities:
            violations.append(f"{label}: fabricated_entity '{row['fabricated_entity']}' is not a kept fake")
        if row["fabricated_entity"] not in row["camouflage_legal_entities"]:
            violations.append(f"{label}: assigned fabricated_entity is not in its own camouflage_legal_entities")
        entity_root = next(info["root"] for info in kept.values() if info["entity"] == row["fabricated_entity"])
        if not is_camouflage_legal(entity_root, row["control_question_text"]):
            violations.append(f"{label}: assigned fake '{row['fabricated_entity']}' violates camouflage rule")

        # Independent re-verification of the exact-insertion contract on the
        # FINAL stored text (not just at construction time).
        matcher = difflib.SequenceMatcher(a=row["control_question_text"], b=row["altered_question_text"], autojunk=False)
        ops = [op for op in matcher.get_opcodes() if op[0] != "equal"]
        if len(ops) != 1 or ops[0][0] != "insert":
            violations.append(f"{label}: final (control, altered) pair is not exactly one insert op")
        else:
            _, i1, i2, j1, j2 = ops[0]
            inserted = row["altered_question_text"][j1:j2]
            # separator_repr is repr() of a plain str we generated ourselves (never
            # external input); ast.literal_eval safely parses it back without
            # executing arbitrary code, unlike eval().
            expected_inserted = row["inserted_sentence"] + ast.literal_eval(row["separator_repr"])
            if inserted != expected_inserted:
                violations.append(f"{label}: final inserted text does not match sentence+separator")
            if i1 != row["insertion_offset"]:
                violations.append(f"{label}: final insertion offset drifted")
        if row["altered_question_text"].count(row["inserted_sentence"]) != 1:
            violations.append(f"{label}: inserted_sentence does not occur exactly once in altered text")
        if sha256_text(row["control_question_text"]) != row["control_text_sha256"]:
            violations.append(f"{label}: control_text_sha256 mismatch")
        if sha256_text(row["altered_question_text"]) != row["altered_text_sha256"]:
            violations.append(f"{label}: altered_text_sha256 mismatch")

    for row in primary:
        if row["pool_role"] != "PRIMARY":
            violations.append(f"{arm} {row['base_question_id']}: listed under primary[] but pool_role={row['pool_role']}")
    for row in reserve:
        if row["pool_role"] != "RESERVE":
            violations.append(f"{arm} {row['base_question_id']}: listed under reserve[] but pool_role={row['pool_role']}")

    ranks = [r["pool_rank"] for r in primary]
    if sorted(ranks) != list(range(1, PRIMARY_TARGET + 1)):
        violations.append(f"{arm}: primary pool_rank set is not exactly 1..{PRIMARY_TARGET}")
    ranks = [r["pool_rank"] for r in reserve]
    if sorted(ranks) != list(range(1, RESERVE_TARGET + 1)):
        violations.append(f"{arm}: reserve pool_rank set is not exactly 1..{RESERVE_TARGET}")


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------


def process_arm(arm: str, source_records: dict, report: dict):
    rows = load_arm_rows(arm, source_records)

    primary_rows, reserve_rows, swap_count, swap_infeasible_reason = determine_primary_set(rows)
    primary_assignment, exact_feasible, infeasible_reason, primary_diag = assign_primary_fakes(
        primary_rows, arm, swap_infeasible_reason
    )
    reserve_assignment, reserve_diag = assign_reserve_fakes(reserve_rows, arm)

    primary_ranks = assign_ranks(primary_rows, "PRIMARY")
    reserve_ranks = assign_ranks(reserve_rows, "RESERVE")

    final_primary = [
        build_final_row(r, arm, "PRIMARY", primary_ranks[r["base_question_id"]], primary_assignment[r["base_question_id"]])
        for r in primary_rows
    ]
    final_reserve = [
        build_final_row(r, arm, "RESERVE", reserve_ranks[r["base_question_id"]], reserve_assignment[r["base_question_id"]])
        for r in reserve_rows
    ]
    final_primary.sort(key=lambda r: r["pool_rank"])
    final_reserve.sort(key=lambda r: r["pool_rank"])

    rows_reassigned = sum(1 for r in final_primary + final_reserve if r["reassigned_from"] is not None)
    rows_moved = sum(1 for r in final_primary + final_reserve if r["role_changed"])

    forced_a_rows = [r["base_question_id"] for r in rows if r["legal"] == {"A"}]
    forced_b_rows = [r["base_question_id"] for r in rows if r["legal"] == {"B"}]

    primary_counts = dict(collections.Counter(r["fabricated_entity"] for r in final_primary))
    reserve_counts = dict(collections.Counter(r["fabricated_entity"] for r in final_reserve))
    for info in KEPT_FAKES[arm].values():
        primary_counts.setdefault(info["entity"], 0)
        reserve_counts.setdefault(info["entity"], 0)

    report["arms"][arm] = {
        "primary_counts": primary_counts,
        "reserve_counts": reserve_counts,
        "rows_reassigned": rows_reassigned,
        "rows_moved_primary_reserve": rows_moved,
        "swap_count_raw": swap_count,
        "exact_5050_achieved": exact_feasible,
        "infeasible_reason": infeasible_reason,
        "camouflage_forced_a": forced_a_rows,
        "camouflage_forced_b": forced_b_rows,
        "primary_diagnostics": primary_diag,
        "reserve_diagnostics": reserve_diag,
        "kept_fake_a_entity": KEPT_FAKES[arm]["A"]["entity"],
        "kept_fake_b_entity": KEPT_FAKES[arm]["B"]["entity"],
    }

    return final_primary, final_reserve


def main() -> int:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    source_records = read_source(SOURCE_PATH)

    report = {"arms": {}}
    arm_results = {}
    try:
        for arm in ARMS:
            arm_results[arm] = process_arm(arm, source_records, report)
    except InvariantViolation as exc:
        print("FATAL during construction:", exc, file=sys.stderr)
        return 1

    violations: list = []
    for arm in ARMS:
        final_primary, final_reserve = arm_results[arm]
        self_assert_arm(arm, final_primary, final_reserve, source_records, violations)

    # Cross-arm duplicate check is NOT required (BM/AN share base_question_id
    # space by design -- 89/130 overlap observed in the locked inputs), so we
    # only assert no-duplicate-within-arm above.

    if violations:
        print(f"FATAL: {len(violations)} invariant violation(s) found:", file=sys.stderr)
        for v in violations:
            print(" -", v, file=sys.stderr)
        return 1

    bm_sha = sha256_bytes(INPUT_PATH["BM"].read_bytes())
    an_sha = sha256_bytes(INPUT_PATH["AN"].read_bytes())
    source_sha = sha256_bytes(SOURCE_PATH.read_bytes())

    baseline = {
        "protocol": PROTOCOL,
        "derived_from": {
            "bm_input_sha256": bm_sha,
            "an_input_sha256": an_sha,
            "source_sha256": source_sha,
        },
        "kept_fakes": {
            arm: [
                {
                    "letter": letter,
                    "entity": KEPT_FAKES[arm][letter]["entity"],
                    "root": KEPT_FAKES[arm][letter]["root"],
                    "sentence": KEPT_FAKES[arm][letter]["sentence"],
                    "variant_id": KEPT_FAKES[arm][letter]["variant_id"],
                }
                for letter in LETTERS
            ]
            for arm in ARMS
        },
        "arms": {},
    }
    for arm in ARMS:
        final_primary, final_reserve = arm_results[arm]
        baseline["arms"][arm] = {
            "primary": final_primary,
            "reserve": final_reserve,
            "primary_counts": report["arms"][arm]["primary_counts"],
            "reserve_counts": report["arms"][arm]["reserve_counts"],
        }

    OUTPUT_JSON.write_text(json.dumps(baseline, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    write_report(report)

    summary = {
        "status": "ok",
        "baseline_json": str(OUTPUT_JSON),
        "assignment_report": str(OUTPUT_REPORT),
        "BM": {
            "primary_counts": report["arms"]["BM"]["primary_counts"],
            "exact_5050_achieved": report["arms"]["BM"]["exact_5050_achieved"],
        },
        "AN": {
            "primary_counts": report["arms"]["AN"]["primary_counts"],
            "exact_5050_achieved": report["arms"]["AN"]["exact_5050_achieved"],
        },
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


def write_report(report: dict) -> None:
    lines = []
    lines.append("# Experiment C -- 2-Fake / 50-50 Rebalanced Baseline -- Assignment Report")
    lines.append("")
    lines.append(f"Protocol: `{PROTOCOL}`")
    lines.append("")
    lines.append("Built by `build_baseline.py` from the locked mechanical-130 workbooks and")
    lines.append("`balanced-flat-A.xlsx` (ultimate ground truth). No input files were modified.")
    lines.append("")

    overall_achieved = all(report["arms"][arm]["exact_5050_achieved"] for arm in ARMS)
    lines.append(f"## Overall: exact 50/50 achieved for both arms: **{overall_achieved}**")
    lines.append("")

    for arm in ARMS:
        a = report["arms"][arm]
        lines.append(f"## Arm {arm}")
        lines.append("")
        lines.append(f"- Kept fake A: `{a['kept_fake_a_entity']}`")
        lines.append(f"- Kept fake B: `{a['kept_fake_b_entity']}`")
        lines.append("")
        lines.append("### PRIMARY (100 rows)")
        lines.append("")
        lines.append("| entity | count |")
        lines.append("|---|---|")
        for entity, count in sorted(a["primary_counts"].items()):
            lines.append(f"| {entity} | {count} |")
        lines.append("")
        lines.append(f"- exact_5050_achieved: **{a['exact_5050_achieved']}**")
        if not a["exact_5050_achieved"]:
            lines.append(f"- reason: {a['infeasible_reason']}")
        lines.append("")
        lines.append("### RESERVE (30 rows)")
        lines.append("")
        lines.append("| entity | count |")
        lines.append("|---|---|")
        for entity, count in sorted(a["reserve_counts"].items()):
            lines.append(f"| {entity} | {count} |")
        lines.append("")
        lines.append("### Churn")
        lines.append("")
        lines.append(f"- rows reassigned (fabricated_entity changed): **{a['rows_reassigned']}**")
        lines.append(f"- rows moved PRIMARY<->RESERVE: **{a['rows_moved_primary_reserve']}**")
        lines.append("")
        lines.append("### Camouflage-forced rows")
        lines.append("")
        lines.append(
            f"- forced to A only (root of B is present in control text): "
            f"**{len(a['camouflage_forced_a'])}** -- {a['camouflage_forced_a']}"
        )
        lines.append(
            f"- forced to B only (root of A is present in control text): "
            f"**{len(a['camouflage_forced_b'])}** -- {a['camouflage_forced_b']}"
        )
        lines.append("")
        lines.append("### Assignment diagnostics")
        lines.append("")
        lines.append("PRIMARY:")
        lines.append("")
        lines.append("```json")
        lines.append(json.dumps(a["primary_diagnostics"], indent=2))
        lines.append("```")
        lines.append("")
        lines.append("RESERVE:")
        lines.append("")
        lines.append("```json")
        lines.append(json.dumps(a["reserve_diagnostics"], indent=2))
        lines.append("```")
        lines.append("")

    lines.append("## Methodology")
    lines.append("")
    lines.append(
        "- **Churn minimization**: the current PRIMARY/RESERVE partition is kept untouched unless "
        "a forced-single-legal-fake side would exceed 50 rows within the current PRIMARY set (not the "
        "case for either arm here -- 0 rows moved). Within PRIMARY, a row keeps its current fake "
        "whenever that fake is one of the two kept fakes and is still camouflage-legal for it; only "
        "rows using a dropped fake, or the minimum number needed to correct an over-50 side, change."
    )
    lines.append(
        "- **Deterministic tie-breaks**: (currently-PRIMARY first, then pool_rank ascending, then "
        "base_question_id) throughout."
    )
    lines.append(
        "- **Spread**: whenever a row's fake assignment is genuinely free (no current-fake preference, "
        "or forced by an overflow flip), the choice is made by a deterministic greedy that prefers the "
        "cluster/region with the fewest rows already carrying that fake, so each fake's 50 are not "
        "clumped in a handful of clusters."
    )
    lines.append("")

    OUTPUT_REPORT.write_text("\n".join(lines) + "\n", encoding="utf-8")


if __name__ == "__main__":
    sys.exit(main())
