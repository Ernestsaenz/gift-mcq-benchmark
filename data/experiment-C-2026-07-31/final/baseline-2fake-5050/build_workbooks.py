#!/usr/bin/env python3
"""Builder B -- render baseline.json (produced by Builder A) into the two
2-fake / 50-50 baseline Excel workbooks for Experiment C.

Consumes /baseline-2fake-5050/baseline.json AS-IS -- no assignment logic is
recomputed here. This script only:
  1. Loads each arm's locked mechanical-130 input workbook (read-only,
     never modified) to source every column baseline.json does not itself
     carry (candidate_origin, condition, text_delta_qa, strict_tier_member,
     relaxed_pair_member, semantic_risk, reviewer_id, qa_status,
     review_rationale, content_sha256, selection_score, context_ids).
  2. Re-derives inserted_segment_repr = repr(inserted_sentence + separator)
     and re-verifies the exact-insertion contract per row.
  3. Writes expC-<arm>-2fake-5050-baseline.xlsx with sheets:
     control_vs_alteration, primary_control_rows, primary_alteration_rows,
     reserve_control_rows, reserve_alteration_rows, README.
  4. Writes manifest.json (sha256 of produced workbooks, baseline.json,
     input workbooks, and the ultimate-ground-truth source workbook).
  5. Re-opens both produced workbooks and verifies them independently.

All outputs are written ONLY under this directory
(.../final/baseline-2fake-5050). The locked input workbooks and the source
workbook are opened read-only and are never written to.
"""
from __future__ import annotations

import ast
import datetime
import difflib
import hashlib
import json
import re
import unicodedata
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
# Paths
# --------------------------------------------------------------------------
ROOT = Path("/Users/ernestsaenz/Programming/gift-project-compile/tier1_mcq")
OUT_DIR = ROOT / "data/experiment-C-2026-07-31/final/baseline-2fake-5050"
INPUT_DIR = ROOT / "data/experiment-C-2026-07-31/final/outputs"

BASELINE_JSON_PATH = OUT_DIR / "baseline.json"

INPUT_PATHS = {
    "BM": INPUT_DIR / "expC-biomarker-mechanical-100-plus-30-reserve-control-vs-alteration.xlsx",
    "AN": INPUT_DIR / "expC-anatomy-mechanical-100-plus-30-reserve-control-vs-alteration.xlsx",
}
SOURCE_PATH = ROOT / "data/experiment-31-07-26/balanced-flat-A.xlsx"

OUTPUT_PATHS = {
    "BM": OUT_DIR / "expC-biomarker-2fake-5050-baseline.xlsx",
    "AN": OUT_DIR / "expC-anatomy-2fake-5050-baseline.xlsx",
}
MANIFEST_PATH = OUT_DIR / "manifest.json"

ARM_LABELS = {"BM": "biomarker", "AN": "anatomy"}

# Exact column order for the pass-through sheet, taken from the locked
# input workbooks (verified identical for both arms).
CVA_HEADER = (
    "pool_role", "pool_rank", "candidate_origin", "base_question_id",
    "region", "year", "specialty", "exam_part", "question_number",
    "condition", "variant_id", "fabricated_entity", "inserted_sentence",
    "control_question_text", "altered_question_text", "text_delta_qa",
    "inserted_character_offset", "inserted_segment_repr",
    "strict_tier_member", "relaxed_pair_member", "cluster", "semantic_risk",
    "reviewer_id", "qa_status", "review_rationale", "option_a", "option_b",
    "option_c", "option_d", "correct_letter", "correct_option_text",
    "flags", "page_in_exam_pdf", "source_exam_pdf", "source_answer_key_pdf",
    "source_key", "content_sha256", "control_text_sha256",
    "altered_text_sha256",
)

ROW_SCHEMA_HEADER = (
    "question_id", "region", "year", "specialty", "exam_part",
    "question_number", "question_text", "option_a", "option_b", "option_c",
    "option_d", "correct_letter", "correct_option_text", "flags",
    "page_in_exam_pdf", "source_exam_pdf", "source_answer_key_pdf",
    "content_sha256", "source_key", "selection_score", "context_ids",
    "condition", "variant_id", "fabricated_entity", "inserted_sentence",
    "text_delta_qa", "pool_role", "pool_rank", "candidate_origin",
    "semantic_risk", "qa_status",
)

# Fields baseline.json rows carry verbatim from the locked input row -- must
# equal the corresponding input-workbook column exactly (else abort: this
# would mean baseline.json is inconsistent with the locked input it claims
# to be derived from).
UNCHANGED_FIELDS = [
    "pool_role", "pool_rank", "region", "year", "specialty", "exam_part",
    "question_number", "control_question_text", "option_a", "option_b",
    "option_c", "option_d", "correct_letter", "correct_option_text",
    "cluster", "flags", "page_in_exam_pdf", "source_exam_pdf",
    "source_answer_key_pdf", "source_key", "control_text_sha256",
]

# Fields that ARE allowed to change vs the locked input row.
OVERRIDE_FIELDS = [
    "variant_id", "fabricated_entity", "inserted_sentence",
    "altered_question_text", "altered_text_sha256",
]


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def normalize(s: str) -> str:
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = re.sub(r"\s+", " ", s)
    return s.casefold()


def load_input_workbook(path: Path):
    """Return (orig_lookup, sel_lookup) keyed by base_question_id / question_id."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["control_vs_alteration"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    assert tuple(header) == CVA_HEADER, f"unexpected control_vs_alteration header in {path}"
    orig_lookup = {}
    for r in rows[1:]:
        d = dict(zip(header, r))
        orig_lookup[d["base_question_id"]] = d
    assert len(orig_lookup) == 130, f"expected 130 unique base_question_id in {path}, got {len(orig_lookup)}"

    ws2 = wb["control_rows"]
    rows2 = list(ws2.iter_rows(values_only=True))
    header2 = rows2[0]
    sel_lookup = {}
    for r in rows2[1:]:
        d = dict(zip(header2, r))
        sel_lookup[d["question_id"]] = {
            "selection_score": d["selection_score"],
            "context_ids": d["context_ids"],
        }
    assert len(sel_lookup) == 130
    wb.close()
    return orig_lookup, sel_lookup


def verify_and_build_row(arm: str, baseline_row: dict, orig: dict, root_by_entity: dict) -> dict:
    """Cross-check baseline_row against the locked input row, re-derive
    inserted_segment_repr, re-verify the exact-insertion contract and the
    camouflage rule, and return the full 39-field control_vs_alteration row
    (dict keyed by CVA_HEADER column names)."""
    qid = baseline_row["base_question_id"]

    # 1. Unchanged fields must be byte-identical to the locked input row.
    for f in UNCHANGED_FIELDS:
        if baseline_row[f] != orig[f]:
            raise AssertionError(
                f"[{arm}/{qid}] unchanged field {f!r} differs: "
                f"baseline={baseline_row[f]!r} input={orig[f]!r}"
            )
    if baseline_row["insertion_offset"] != orig["inserted_character_offset"]:
        raise AssertionError(
            f"[{arm}/{qid}] insertion_offset differs: "
            f"baseline={baseline_row['insertion_offset']!r} "
            f"input={orig['inserted_character_offset']!r}"
        )

    # 2. Exact-insertion contract, re-verified from scratch.
    control = baseline_row["control_question_text"]
    i = baseline_row["insertion_offset"]
    sentence = baseline_row["inserted_sentence"]
    sep = ast.literal_eval(baseline_row["separator_repr"])
    if not sep.isspace():
        raise AssertionError(f"[{arm}/{qid}] separator is not whitespace: {sep!r}")

    reconstructed = control[:i] + sentence + sep + control[i:]
    if reconstructed != baseline_row["altered_question_text"]:
        raise AssertionError(f"[{arm}/{qid}] reconstructed altered text mismatch")

    digest = hashlib.sha256(reconstructed.encode("utf-8")).hexdigest()
    if digest != baseline_row["altered_text_sha256"]:
        raise AssertionError(f"[{arm}/{qid}] altered_text_sha256 mismatch")

    sm = difflib.SequenceMatcher(None, control, reconstructed, autojunk=False)
    ops = [op for op in sm.get_opcodes() if op[0] != "equal"]
    if len(ops) != 1 or ops[0][0] != "insert":
        raise AssertionError(f"[{arm}/{qid}] expected exactly one insert opcode, got {ops}")
    tag, i1, i2, j1, j2 = ops[0]
    if i1 != i or reconstructed[j1:j2] != sentence + sep:
        raise AssertionError(f"[{arm}/{qid}] insert opcode does not match offset/segment")
    if reconstructed[:i] != control[:i]:
        raise AssertionError(f"[{arm}/{qid}] prefix mismatch")
    if reconstructed[i + len(sentence + sep):] != control[i:]:
        raise AssertionError(f"[{arm}/{qid}] suffix mismatch")
    if reconstructed.count(sentence) != 1:
        raise AssertionError(f"[{arm}/{qid}] inserted sentence does not occur exactly once")

    # 3. Camouflage rule.
    entity = baseline_row["fabricated_entity"]
    root = root_by_entity[entity]
    if root in normalize(control):
        raise AssertionError(f"[{arm}/{qid}] camouflage violation: root {root!r} in control text")

    # 4. correct_option_text must equal option_[correct_letter].
    letter = baseline_row["correct_letter"]
    if baseline_row[f"option_{letter}"] != baseline_row["correct_option_text"]:
        raise AssertionError(f"[{arm}/{qid}] correct_option_text does not equal option_{letter}")

    inserted_segment_repr = repr(sentence + sep)

    # 5. Build the full 39-column row: start from the locked input row,
    #    then overlay only the fields the contract allows to change.
    new_row = dict(orig)
    new_row["pool_role"] = baseline_row["pool_role"]
    new_row["pool_rank"] = baseline_row["pool_rank"]
    for f in OVERRIDE_FIELDS:
        new_row[f] = baseline_row[f]
    new_row["inserted_segment_repr"] = inserted_segment_repr
    # inserted_character_offset is invariant across reassignment; keep the
    # (already-verified-equal) baseline.json value for clarity.
    new_row["inserted_character_offset"] = baseline_row["insertion_offset"]
    new_row["base_question_id"] = qid
    return new_row


def build_row_schema_row(new_row: dict, sel: dict, variant: str) -> tuple:
    """variant: 'control' or 'alteration'."""
    question_text = new_row["control_question_text"] if variant == "control" else new_row["altered_question_text"]
    values = {
        "question_id": new_row["base_question_id"],
        "region": new_row["region"],
        "year": new_row["year"],
        "specialty": new_row["specialty"],
        "exam_part": new_row["exam_part"],
        "question_number": new_row["question_number"],
        "question_text": question_text,
        "option_a": new_row["option_a"],
        "option_b": new_row["option_b"],
        "option_c": new_row["option_c"],
        "option_d": new_row["option_d"],
        "correct_letter": new_row["correct_letter"],
        "correct_option_text": new_row["correct_option_text"],
        "flags": new_row["flags"],
        "page_in_exam_pdf": new_row["page_in_exam_pdf"],
        "source_exam_pdf": new_row["source_exam_pdf"],
        "source_answer_key_pdf": new_row["source_answer_key_pdf"],
        "content_sha256": new_row["content_sha256"],
        "source_key": new_row["source_key"],
        "selection_score": sel["selection_score"],
        "context_ids": sel["context_ids"],
        "condition": new_row["condition"],
        "variant_id": new_row["variant_id"],
        "fabricated_entity": new_row["fabricated_entity"],
        "inserted_sentence": new_row["inserted_sentence"],
        "text_delta_qa": new_row["text_delta_qa"],
        "pool_role": new_row["pool_role"],
        "pool_rank": new_row["pool_rank"],
        "candidate_origin": new_row["candidate_origin"],
        "semantic_risk": new_row["semantic_risk"],
        "qa_status": new_row["qa_status"],
    }
    return tuple(values[h] for h in ROW_SCHEMA_HEADER)


def style_header(ws, ncols: int) -> None:
    bold = Font(bold=True)
    for c in range(1, ncols + 1):
        ws.cell(row=1, column=c).font = bold
    ws.freeze_panes = "A2"


def write_data_sheet(wb, name: str, header: tuple, rows: list) -> None:
    ws = wb.create_sheet(name)
    ws.append(list(header))
    for row in rows:
        ws.append(list(row))
    style_header(ws, len(header))


def write_readme_sheet(wb, arm: str, arm_data: dict, kept_fakes: list, derived_from: dict,
                        churn: dict, generated_at: str) -> None:
    ws = wb.create_sheet("README")
    label = ARM_LABELS[arm]
    a, b = kept_fakes
    pc = arm_data["primary_counts"]
    rc = arm_data["reserve_counts"]
    rows = [
        (f"Experiment C -- {label} 2-fake / 50-50 rebalanced baseline", None),
        ("Protocol", "expc-2fake-5050-baseline-v1"),
        ("Status", "FRESH BASELINE ARTIFACT -- not a replacement of the committed files"),
        (
            "Provenance departure",
            "This workbook intentionally departs from the locked-legacy SHA "
            "provenance recorded in the input workbook's own README/manifest. "
            "It is a new derived artifact produced by reassigning "
            "fabricated_entity labels on top of the locked mechanical-130 "
            "pool -- it does not overwrite, and is not a replacement for, "
            "the committed workbook under final/outputs/, which remains the "
            "locked baseline of record.",
        ),
        ("Contents", "130 rows total: 100 primary + 30 reserve control-versus-alteration pairs."),
        (
            "Two-fake design",
            f"Exactly two fabricated findings are used in this arm: "
            f"A = {a['entity']} (root {a['root']!r}, variant {a['variant_id']}), "
            f"B = {b['entity']} (root {b['root']!r}, variant {b['variant_id']}). "
            "Every other fake used in the locked input pool was dropped and "
            "its rows reassigned onto A or B.",
        ),
        (
            "50/50 split",
            f"Primary (100 rows): {a['entity']} x {pc.get(a['entity'], 0)}, "
            f"{b['entity']} x {pc.get(b['entity'], 0)}. "
            f"Reserve (30 rows): {a['entity']} x {rc.get(a['entity'], 0)}, "
            f"{b['entity']} x {rc.get(b['entity'], 0)}.",
        ),
        (
            "Camouflage rule",
            "A fake may be assigned to a row only if its root is not a "
            "substring of NFD-normalized (accents stripped), whitespace-"
            "collapsed, casefolded control_question_text. Verified for "
            "every row in this workbook.",
        ),
        (
            "Text guarantee",
            "control_question_text is byte-identical to the locked input "
            "workbook and to balanced-flat-A.xlsx (ultimate ground truth). "
            "altered_question_text = control text with the assigned "
            "fake's sentence plus the row's original copied whitespace "
            "separator inserted at the row's original insertion offset -- "
            "no other character is added, removed, reordered, or "
            "normalized.",
        ),
        (
            "Unchanged vs locked input",
            "Only variant_id, fabricated_entity, inserted_sentence, "
            "altered_question_text, altered_text_sha256, and "
            "inserted_segment_repr change per reassigned row (plus "
            "pool_role/pool_rank if a row's primary/reserve role changed). "
            "Every other column -- including control_question_text, "
            "option_a..d, correct_letter, correct_option_text, and all "
            "provenance columns -- is byte-identical to the locked input "
            "workbook.",
        ),
        (
            "Churn vs locked input",
            f"{churn['reassigned']} of 130 rows had fabricated_entity "
            f"reassigned; {churn['role_changed']} rows moved between "
            "PRIMARY and RESERVE.",
        ),
        ("Primary/reserve", "pool_role and pool_rank distinguish the 100 primary rows from the 30 reserve rows."),
        ("Assignment source", "baseline.json (built by build_baseline.py, Builder A). Consumed as-is by this workbook builder; no assignment logic was recomputed here."),
        ("Locked input workbook", INPUT_PATHS[arm].name),
        ("Locked input workbook SHA-256", derived_from[f"{arm.lower()}_input_sha256"]),
        ("Ultimate ground-truth source", SOURCE_PATH.name),
        ("Ultimate ground-truth source SHA-256", derived_from["source_sha256"]),
        ("Generated by", "build_workbooks.py (Builder B)"),
        ("Generated at (UTC)", generated_at),
    ]
    for r in rows:
        ws.append(list(r))
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 100
    ws.cell(row=1, column=1).font = Font(bold=True, size=13)
    for r in range(2, len(rows) + 1):
        ws.cell(row=r, column=1).font = Font(bold=True)
        ws.cell(row=r, column=2).alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r, column=1).alignment = openpyxl.styles.Alignment(vertical="top")


def main():
    print("Loading baseline.json ...")
    with open(BASELINE_JSON_PATH, "r", encoding="utf-8") as f:
        baseline = json.load(f)
    assert baseline["protocol"] == "expc-2fake-5050-baseline-v1"

    generated_at = datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    manifest = {
        "protocol": baseline["protocol"],
        "generated_at": generated_at,
        "builder": "Builder B (build_workbooks.py) -- renders baseline.json into workbooks; assignment logic owned by Builder A (build_baseline.py)",
        "inputs": {},
        "outputs": {},
        "verification": {},
    }

    # sha256 of baseline.json itself
    baseline_json_sha256 = sha256_file(BASELINE_JSON_PATH)
    manifest["inputs"]["baseline_json"] = {
        "path": str(BASELINE_JSON_PATH),
        "sha256": baseline_json_sha256,
    }

    # sha256 of the ultimate ground-truth source
    source_sha256 = sha256_file(SOURCE_PATH)
    manifest["inputs"]["source_workbook"] = {
        "path": str(SOURCE_PATH),
        "sha256": source_sha256,
    }
    source_matches_claim = source_sha256 == baseline["derived_from"]["source_sha256"]
    print(f"  source_workbook sha256 matches baseline.json derived_from claim: {source_matches_claim}")
    assert source_matches_claim, "SOURCE workbook sha256 does not match baseline.json's recorded provenance"

    for_all_arms_ok = True

    for arm in ["BM", "AN"]:
        print(f"\n=== Arm {arm} ===")
        input_path = INPUT_PATHS[arm]
        input_sha256 = sha256_file(input_path)
        manifest["inputs"][f"{arm.lower()}_input_workbook"] = {
            "path": str(input_path),
            "sha256": input_sha256,
        }
        claim_key = f"{arm.lower()}_input_sha256"
        input_matches_claim = input_sha256 == baseline["derived_from"][claim_key]
        print(f"  input workbook sha256 matches baseline.json derived_from claim: {input_matches_claim}")
        assert input_matches_claim, f"{arm} input workbook sha256 does not match baseline.json's recorded provenance"

        orig_lookup, sel_lookup = load_input_workbook(input_path)
        print(f"  loaded {len(orig_lookup)} locked input rows")

        kept_fakes = baseline["kept_fakes"][arm]
        root_by_entity = {kf["entity"]: kf["root"] for kf in kept_fakes}
        entity_a, entity_b = kept_fakes[0]["entity"], kept_fakes[1]["entity"]

        arm_data = baseline["arms"][arm]
        primary_rows_json = sorted(arm_data["primary"], key=lambda r: r["pool_rank"])
        reserve_rows_json = sorted(arm_data["reserve"], key=lambda r: r["pool_rank"])
        assert [r["pool_rank"] for r in primary_rows_json] == list(range(1, 101))
        assert [r["pool_rank"] for r in reserve_rows_json] == list(range(1, 31))

        cva_rows = []
        primary_new_rows = []
        reserve_new_rows = []
        reassigned_count = 0
        role_changed_count = 0

        for br in primary_rows_json + reserve_rows_json:
            qid = br["base_question_id"]
            orig = orig_lookup[qid]
            new_row = verify_and_build_row(arm, br, orig, root_by_entity)
            cva_rows.append(tuple(new_row[h] for h in CVA_HEADER))
            if br["pool_role"] == "PRIMARY":
                primary_new_rows.append((new_row, sel_lookup[qid]))
            else:
                reserve_new_rows.append((new_row, sel_lookup[qid]))
            if br.get("reassigned_from"):
                reassigned_count += 1
            if br.get("role_changed"):
                role_changed_count += 1

        assert len(cva_rows) == 130
        assert len(primary_new_rows) == 100
        assert len(reserve_new_rows) == 30
        print(f"  verified & built 130 rows (100 primary + 30 reserve); "
              f"reassigned={reassigned_count} role_changed={role_changed_count}")

        # Primary/reserve counts per entity, cross-checked against baseline.json's own counts.
        primary_counts = {}
        for nr, _ in primary_new_rows:
            e = nr["fabricated_entity"]
            primary_counts[e] = primary_counts.get(e, 0) + 1
        reserve_counts = {}
        for nr, _ in reserve_new_rows:
            e = nr["fabricated_entity"]
            reserve_counts[e] = reserve_counts.get(e, 0) + 1

        assert primary_counts == arm_data["primary_counts"], (
            f"[{arm}] primary_counts mismatch: built={primary_counts} "
            f"baseline.json={arm_data['primary_counts']}"
        )
        assert reserve_counts == arm_data["reserve_counts"], (
            f"[{arm}] reserve_counts mismatch"
        )
        assert primary_counts.get(entity_a, 0) == 50 and primary_counts.get(entity_b, 0) == 50, (
            f"[{arm}] primary split is not exactly 50/50: {primary_counts}"
        )
        assert set(primary_counts) | set(reserve_counts) <= {entity_a, entity_b}, (
            f"[{arm}] more than 2 distinct fabricated_entity values present"
        )
        print(f"  primary counts: {primary_counts}  (exact 50/50 confirmed)")
        print(f"  reserve counts: {reserve_counts}")

        # --- Build workbook ---
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        write_data_sheet(wb, "control_vs_alteration", CVA_HEADER, cva_rows)

        primary_control_rows = [build_row_schema_row(nr, sel, "control") for nr, sel in primary_new_rows]
        primary_alteration_rows = [build_row_schema_row(nr, sel, "alteration") for nr, sel in primary_new_rows]
        reserve_control_rows = [build_row_schema_row(nr, sel, "control") for nr, sel in reserve_new_rows]
        reserve_alteration_rows = [build_row_schema_row(nr, sel, "alteration") for nr, sel in reserve_new_rows]

        write_data_sheet(wb, "primary_control_rows", ROW_SCHEMA_HEADER, primary_control_rows)
        write_data_sheet(wb, "primary_alteration_rows", ROW_SCHEMA_HEADER, primary_alteration_rows)
        write_data_sheet(wb, "reserve_control_rows", ROW_SCHEMA_HEADER, reserve_control_rows)
        write_data_sheet(wb, "reserve_alteration_rows", ROW_SCHEMA_HEADER, reserve_alteration_rows)

        churn = {"reassigned": reassigned_count, "role_changed": role_changed_count}
        write_readme_sheet(wb, arm, arm_data, kept_fakes, baseline["derived_from"], churn, generated_at)

        out_path = OUTPUT_PATHS[arm]
        wb.save(out_path)
        wb.close()
        print(f"  wrote {out_path}")

        out_sha256 = sha256_file(out_path)
        manifest["outputs"][f"{arm.lower()}_workbook"] = {
            "path": str(out_path),
            "sha256": out_sha256,
            "kept_fakes": {"A": entity_a, "B": entity_b},
            "primary_counts": primary_counts,
            "reserve_counts": reserve_counts,
            "exact_5050_achieved": bool(
                primary_counts.get(entity_a, 0) == 50 and primary_counts.get(entity_b, 0) == 50
            ),
            "reassigned_rows": reassigned_count,
            "role_changed_rows": role_changed_count,
        }

        # --- Re-open and independently verify ---
        print(f"  re-opening {out_path.name} for independent verification ...")
        wb_check = openpyxl.load_workbook(out_path, data_only=True, read_only=True)
        assert set(wb_check.sheetnames) == {
            "control_vs_alteration", "primary_control_rows", "primary_alteration_rows",
            "reserve_control_rows", "reserve_alteration_rows", "README",
        }, f"unexpected sheet set: {wb_check.sheetnames}"

        ws_chk = wb_check["control_vs_alteration"]
        chk_rows = list(ws_chk.iter_rows(values_only=True))
        chk_header = chk_rows[0]
        assert tuple(chk_header) == CVA_HEADER
        chk_data = [dict(zip(chk_header, r)) for r in chk_rows[1:]]
        assert len(chk_data) == 130

        n_primary = sum(1 for d in chk_data if d["pool_role"] == "PRIMARY")
        n_reserve = sum(1 for d in chk_data if d["pool_role"] == "RESERVE")
        assert n_primary == 100 and n_reserve == 30, f"role counts wrong: primary={n_primary} reserve={n_reserve}"

        distinct_entities = set(d["fabricated_entity"] for d in chk_data)
        assert distinct_entities == {entity_a, entity_b}, f"expected exactly 2 fabricated_entity values, got {distinct_entities}"

        chk_primary_counts = {}
        for d in chk_data:
            if d["pool_role"] == "PRIMARY":
                e = d["fabricated_entity"]
                chk_primary_counts[e] = chk_primary_counts.get(e, 0) + 1
        assert chk_primary_counts.get(entity_a, 0) == 50 and chk_primary_counts.get(entity_b, 0) == 50, (
            f"re-opened workbook primary split is not 50/50: {chk_primary_counts}"
        )

        # row-for-row match against baseline.json (via field-name mapping)
        by_qid_json = {r["base_question_id"]: r for r in primary_rows_json + reserve_rows_json}
        field_map = {
            "pool_role": "pool_role", "pool_rank": "pool_rank",
            "base_question_id": "base_question_id", "region": "region", "year": "year",
            "specialty": "specialty", "exam_part": "exam_part",
            "question_number": "question_number", "variant_id": "variant_id",
            "fabricated_entity": "fabricated_entity", "inserted_sentence": "inserted_sentence",
            "control_question_text": "control_question_text",
            "altered_question_text": "altered_question_text",
            "inserted_character_offset": "insertion_offset", "cluster": "cluster",
            "option_a": "option_a", "option_b": "option_b", "option_c": "option_c",
            "option_d": "option_d", "correct_letter": "correct_letter",
            "correct_option_text": "correct_option_text", "flags": "flags",
            "page_in_exam_pdf": "page_in_exam_pdf", "source_exam_pdf": "source_exam_pdf",
            "source_answer_key_pdf": "source_answer_key_pdf", "source_key": "source_key",
            "control_text_sha256": "control_text_sha256",
            "altered_text_sha256": "altered_text_sha256",
        }
        mismatches = 0
        for d in chk_data:
            qid = d["base_question_id"]
            jr = by_qid_json[qid]
            for wb_field, json_field in field_map.items():
                if d[wb_field] != jr[json_field]:
                    mismatches += 1
                    print(f"    MISMATCH qid={qid} field={wb_field}: wb={d[wb_field]!r} json={jr[json_field]!r}")
            # inserted_segment_repr re-derivation check
            sep = ast.literal_eval(jr["separator_repr"])
            expected_repr = repr(jr["inserted_sentence"] + sep)
            if d["inserted_segment_repr"] != expected_repr:
                mismatches += 1
                print(f"    MISMATCH qid={qid} inserted_segment_repr")
        assert mismatches == 0, f"{mismatches} field mismatches vs baseline.json"
        print(f"  control_vs_alteration matches baseline.json row-for-row: True (0 mismatches over {len(chk_data)} rows)")

        for sheet_name, expected_len in [
            ("primary_control_rows", 100), ("primary_alteration_rows", 100),
            ("reserve_control_rows", 30), ("reserve_alteration_rows", 30),
        ]:
            ws2 = wb_check[sheet_name]
            r2 = list(ws2.iter_rows(values_only=True))
            assert tuple(r2[0]) == ROW_SCHEMA_HEADER, f"{sheet_name} header mismatch"
            assert len(r2) - 1 == expected_len, f"{sheet_name} expected {expected_len} rows, got {len(r2)-1}"

        wb_check.close()
        print(f"  independent verification PASSED for {arm}")

        manifest["verification"][arm] = {
            "rows_total": len(chk_data),
            "rows_primary": n_primary,
            "rows_reserve": n_reserve,
            "distinct_fabricated_entities": sorted(distinct_entities),
            "primary_split_5050": chk_primary_counts,
            "matches_baseline_json_row_for_row": mismatches == 0,
            "sheet_names": wb_check.sheetnames if False else [
                "control_vs_alteration", "primary_control_rows", "primary_alteration_rows",
                "reserve_control_rows", "reserve_alteration_rows", "README",
            ],
        }

    manifest["matches_baseline_json"] = all(
        manifest["verification"][arm]["matches_baseline_json_row_for_row"] for arm in ["BM", "AN"]
    )

    with open(MANIFEST_PATH, "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2, ensure_ascii=False, sort_keys=False)
    print(f"\nWrote {MANIFEST_PATH}")
    print("\nALL DONE.")


if __name__ == "__main__":
    main()
