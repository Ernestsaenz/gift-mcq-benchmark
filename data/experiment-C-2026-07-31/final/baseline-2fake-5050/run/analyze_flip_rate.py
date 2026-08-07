"""Experiment C (2-fake/50-50 baseline) — CTRL vs ALT answer-flip-rate analysis.

Primary outcome is the answer-FLIP rate, not accuracy: for each (provider, model),
pair every question_id's control-condition answer with its altered-condition answer
and ask whether the model's selected_letter changed. Flip rate is computed per
model x arm (BM, AN), with cluster-robust statistics using the `cluster` field
(the exam/case grouping baked into the run-ready workbooks as an extra column —
the medrag_eval DB schema itself has no cluster column, so it is re-joined here by
question_id straight from one of the two workbooks used to import that arm).

Two cluster-robust estimates are reported, matching the convention already used for
the sibling 2026-07-31 balanced-MCQ analysis in this repo
(data/experiment-31-07-26/analysis/00_ORGANIZED_VIEW/06_exploratory/03_statistical_foundations/):
  1. Analytic CR1 sandwich SE (statsmodels OLS, cov_type="cluster") with a
     cluster-count-corrected t interval (df = n_clusters - 1).
  2. Nonparametric whole-cluster percentile bootstrap 95% CI (resample clusters
     with replacement, carrying every question in a sampled cluster along).

Only question_ids where BOTH the control and altered logical calls parsed to a
usable letter (parse_status in {"ok", "ok_conflict"}) are counted as paired
observations; anything else is reported separately as excluded/unparsed so it is
never silently dropped from the denominator without a trace.

Usage:
  PYTHONPATH=code .venv/bin/python run/analyze_flip_rate.py \
      --db runs/expC_2fake_5050.sqlite \
      --arm BM \
      --control-experiment expC-bm-control \
      --altered-experiment expC-bm-altered \
      --cluster-workbook run/expC-bm-control.xlsx \
      --out run/flip_rate_BM.csv

Run once per arm (BM, AN). Requires numpy, statsmodels (both already project
"analysis"-extra dependencies; see pyproject.toml).
"""
from __future__ import annotations

import argparse
import csv
import sqlite3
import sys
from collections import defaultdict
from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from statsmodels.regression.linear_model import OLS

# Import the harness's own DB module so the summary query (latest attempt / parse /
# score per logical call) exactly matches what `medrag-eval export --format csv`
# would produce -- do not hand-roll that SQL a second time here.
sys.path.insert(0, str(Path(__file__).resolve().parents[4] / "code"))
from medrag_eval import db  # noqa: E402

COMPLETE_PARSE_STATUSES = {"ok", "ok_conflict"}


def load_cluster_map(workbook_path: Path) -> dict[str, str]:
    """question_id -> cluster, read from the extra `cluster` column this builder adds."""
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = wb.active
    rows = sheet.iter_rows(values_only=True)
    header = next(rows)
    idx = {name: i for i, name in enumerate(header)}
    if "cluster" not in idx or "question_id" not in idx:
        raise SystemExit(f"{workbook_path}: missing question_id/cluster columns")
    out: dict[str, str] = {}
    for row in rows:
        qid = row[idx["question_id"]]
        if qid is None:
            continue
        out[str(qid)] = row[idx["cluster"]]
    return out


def load_answers(conn: sqlite3.Connection, experiment_name: str) -> dict[tuple[str, str], dict]:
    """(provider, model) -> {question_id: row_dict} for the given experiment.

    Uses db.summary_rows, the same "latest attempt/parse/score per logical call"
    query the CLI's `export --format csv` uses, so this script and a plain CSV
    export can never silently disagree about what "the" answer for a cell was.
    """
    out: dict[tuple[str, str], dict] = defaultdict(dict)
    for row in db.summary_rows(conn, experiment_name):
        r = dict(row)
        key = (r["provider"], r["model"])
        out[key][r["question_id"]] = r
    return out


def cluster_robust_mean(y: np.ndarray, clusters: np.ndarray) -> dict:
    """CR1 sandwich SE + t-interval for the mean of a 0/1 vector, statsmodels OLS."""
    n = len(y)
    groups = np.asarray(clusters)
    n_clusters = len(np.unique(groups))
    x = np.ones((n, 1))
    if n_clusters < 2:
        return {
            "mean": float(y.mean()) if n else float("nan"),
            "se": float("nan"),
            "ci_lo": float("nan"),
            "ci_hi": float("nan"),
            "n_clusters": n_clusters,
            "note": "fewer than 2 clusters; cluster-robust SE undefined",
        }
    fit = OLS(y, x).fit(cov_type="cluster", cov_kwds={"groups": groups})
    ci = fit.conf_int(alpha=0.05)
    return {
        "mean": float(fit.params[0]),
        "se": float(fit.bse[0]),
        "ci_lo": float(ci[0][0]),
        "ci_hi": float(ci[0][1]),
        "n_clusters": n_clusters,
        "note": "",
    }


def cluster_bootstrap_ci(
    y: np.ndarray, clusters: np.ndarray, *, n_boot: int = 10000, seed: int = 20260731
) -> tuple[float, float]:
    """Whole-cluster percentile bootstrap 95% CI for the mean of a 0/1 vector.

    Resamples clusters (not items) with replacement, carrying every item in a
    sampled cluster along -- the standard nonparametric cross-check for a
    cluster-robust proportion (same idea as cluster_ratio_bootstrap /
    cluster_accuracy_bootstrap in the sibling 2026-07-31 analysis).
    """
    unique_clusters = np.unique(clusters)
    n_clusters = len(unique_clusters)
    if n_clusters < 2:
        return (float("nan"), float("nan"))
    by_cluster = {c: y[clusters == c] for c in unique_clusters}
    sizes = np.array([by_cluster[c].size for c in unique_clusters])
    sums = np.array([by_cluster[c].sum() for c in unique_clusters])
    rng = np.random.default_rng(seed)
    draws = rng.integers(0, n_clusters, size=(n_boot, n_clusters))
    boot_means = (sums[draws].sum(axis=1)) / (sizes[draws].sum(axis=1))
    lo, hi = np.percentile(boot_means, [2.5, 97.5])
    return (float(lo), float(hi))


def compute_flip_table(
    control_answers: dict[tuple[str, str], dict],
    altered_answers: dict[tuple[str, str], dict],
    cluster_map: dict[str, str],
) -> list[dict]:
    provider_models = sorted(set(control_answers) & set(altered_answers))
    missing_only_control = sorted(set(control_answers) - set(altered_answers))
    missing_only_altered = sorted(set(altered_answers) - set(control_answers))
    for pm in missing_only_control:
        print(f"WARNING: {pm} present in control experiment but not altered; skipped", file=sys.stderr)
    for pm in missing_only_altered:
        print(f"WARNING: {pm} present in altered experiment but not control; skipped", file=sys.stderr)

    results = []
    for provider, model in provider_models:
        ctrl_by_q = control_answers[(provider, model)]
        alt_by_q = altered_answers[(provider, model)]
        question_ids = sorted(set(ctrl_by_q) & set(alt_by_q))
        missing_ids = (set(ctrl_by_q) | set(alt_by_q)) - set(question_ids)

        flips: list[int] = []
        clusters: list[str] = []
        excluded_unparsed = 0
        excluded_no_cluster = 0
        ctrl_correct = []
        alt_correct = []
        for qid in question_ids:
            c = ctrl_by_q[qid]
            a = alt_by_q[qid]
            if c["parse_status"] not in COMPLETE_PARSE_STATUSES or a["parse_status"] not in COMPLETE_PARSE_STATUSES:
                excluded_unparsed += 1
                continue
            cluster = cluster_map.get(qid)
            if cluster is None:
                excluded_no_cluster += 1
                continue
            flips.append(1 if c["selected_letter"] != a["selected_letter"] else 0)
            clusters.append(cluster)
            if c["strict_correct"] is not None:
                ctrl_correct.append(int(c["strict_correct"]))
            if a["strict_correct"] is not None:
                alt_correct.append(int(a["strict_correct"]))

        y = np.array(flips, dtype=float)
        g = np.array(clusters)
        analytic = cluster_robust_mean(y, g)
        boot_lo, boot_hi = cluster_bootstrap_ci(y, g)

        results.append(
            {
                "provider": provider,
                "model": model,
                "n_paired": len(question_ids),
                "n_scored": len(flips),
                "n_excluded_unparsed": excluded_unparsed,
                "n_excluded_no_cluster": excluded_no_cluster,
                "n_missing_one_side": len(missing_ids),
                "flip_rate": analytic["mean"],
                "cluster_robust_se": analytic["se"],
                "ci95_lo_analytic": analytic["ci_lo"],
                "ci95_hi_analytic": analytic["ci_hi"],
                "ci95_lo_bootstrap": boot_lo,
                "ci95_hi_bootstrap": boot_hi,
                "n_clusters": analytic["n_clusters"],
                "accuracy_control": (sum(ctrl_correct) / len(ctrl_correct)) if ctrl_correct else float("nan"),
                "accuracy_altered": (sum(alt_correct) / len(alt_correct)) if alt_correct else float("nan"),
                "note": analytic["note"],
            }
        )
    return results


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--db", required=True, type=Path, help="Path to the sqlite DB the two experiments were run into.")
    parser.add_argument("--arm", required=True, choices=["BM", "AN"], help="Arm label, for the report header only.")
    parser.add_argument("--control-experiment", required=True, help="experiment_name used for the control run.")
    parser.add_argument("--altered-experiment", required=True, help="experiment_name used for the altered run.")
    parser.add_argument(
        "--cluster-workbook",
        required=True,
        type=Path,
        help="Either expC-{bm,an}-control.xlsx or -altered.xlsx for this arm (same question_id -> cluster map either way).",
    )
    parser.add_argument("--out", type=Path, default=None, help="Optional CSV path for the result table.")
    args = parser.parse_args()

    cluster_map = load_cluster_map(args.cluster_workbook)

    conn = db.connect(args.db)
    try:
        control_answers = load_answers(conn, args.control_experiment)
        altered_answers = load_answers(conn, args.altered_experiment)
    finally:
        conn.close()

    if not control_answers:
        raise SystemExit(f"No summary rows found for experiment {args.control_experiment!r} in {args.db}")
    if not altered_answers:
        raise SystemExit(f"No summary rows found for experiment {args.altered_experiment!r} in {args.db}")

    results = compute_flip_table(control_answers, altered_answers, cluster_map)

    fieldnames = [
        "arm",
        "provider",
        "model",
        "n_paired",
        "n_scored",
        "n_excluded_unparsed",
        "n_excluded_no_cluster",
        "n_missing_one_side",
        "flip_rate",
        "cluster_robust_se",
        "ci95_lo_analytic",
        "ci95_hi_analytic",
        "ci95_lo_bootstrap",
        "ci95_hi_bootstrap",
        "n_clusters",
        "accuracy_control",
        "accuracy_altered",
        "note",
    ]
    print(",".join(fieldnames))
    rows_out = []
    for r in results:
        row = {"arm": args.arm, **r}
        rows_out.append(row)
        print(",".join(str(row[f]) for f in fieldnames))

    if args.out:
        args.out.parent.mkdir(parents=True, exist_ok=True)
        with args.out.open("w", newline="", encoding="utf-8") as fh:
            writer = csv.DictWriter(fh, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows_out)
        print(f"\nWrote {args.out}", file=sys.stderr)

    return 0


if __name__ == "__main__":
    sys.exit(main())
