"""Build expC-items.xlsx — the researcher-facing working artifact for Experiment C.

Reads canonical.json (authority: tools/build_expC.py) plus balanced-flat-A.xlsx
(read-only, only to count camouflage-root occurrences across all 474 rows).

Values only. No formulas are written, so nothing can recalculate differently on
another machine.
"""
from __future__ import annotations

import collections
import json
import re
import unicodedata
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
CANON = ROOT / "canonical.json"
SRC = Path("/Users/ernestsaenz/Programming/GIFT-abstract-dossier/tier1_mcq/data"
           "/experiment-31-07-26/balanced-flat-A.xlsx")
OUT = ROOT / "expC-items.xlsx"
GEN_DATE = "2026-07-31"

# ---------------------------------------------------------------- palette ----
C_HDR = "FF44505E"        # slate header band
C_HDR_TXT = "FFFFFFFF"
C_TITLE = "FFE9E5DD"      # warm parchment, sheet title strip
C_MOD = "FFE7EFE6"        # muted sage  -> a fabricated sentence was inserted here
C_CTRL = "FFF6F5F2"       # muted stone -> unmodified control text
C_OUT = "FFF1EEEA"        # muted clay  -> row dropped from the pool
C_STAGE = "FFDDE4EA"      # muted steel -> funnel stage separator
C_RULE = "FFB9C0C7"

F_HDR = Font(bold=True, color=C_HDR_TXT, size=10)
F_TITLE = Font(bold=True, size=11, color="FF2E3742")
F_SUB = Font(italic=True, size=9, color="FF56606B")
F_BODY = Font(size=10)
F_MONO = Font(size=9, name="Menlo")
F_SECT = Font(bold=True, size=11, color="FF2E3742")

FILL_HDR = PatternFill("solid", fgColor=C_HDR)
FILL_TITLE = PatternFill("solid", fgColor=C_TITLE)
FILL_MOD = PatternFill("solid", fgColor=C_MOD)
FILL_CTRL = PatternFill("solid", fgColor=C_CTRL)
FILL_OUT = PatternFill("solid", fgColor=C_OUT)
FILL_STAGE = PatternFill("solid", fgColor=C_STAGE)

TOP = Alignment(vertical="top", wrap_text=True)
TOPL = Alignment(vertical="top", horizontal="left", wrap_text=True)
TOPC = Alignment(vertical="center", horizontal="center", wrap_text=True)
HDR_AL = Alignment(vertical="center", horizontal="left", wrap_text=True)
RULE = Border(bottom=Side(style="thin", color=C_RULE))


def nf(s: str) -> str:
    """Same normalisation as build_expC.py, so root counts are comparable."""
    s = unicodedata.normalize("NFD", s or "")
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"\s+", " ", s).lower()
    return re.sub(r"normal\s*:", "REFRANGE:", s)


# ------------------------------------------------------------------ sheet ----
def table_sheet(wb, name, title, subtitle, headers, widths, freeze_col=1,
                row_height=None):
    """Row 1 title, row 2 subtitle, row 3 column headers, data from row 4."""
    ws = wb.create_sheet(name)
    ws.sheet_properties.outlinePr.summaryBelow = False
    ncol = len(headers)
    ws.cell(1, 1, title).font = F_TITLE
    ws.cell(2, 1, subtitle).font = F_SUB
    for c in range(1, ncol + 1):
        ws.cell(1, c).fill = FILL_TITLE
        ws.cell(2, c).fill = FILL_TITLE
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 26
    ws.cell(2, 1).alignment = Alignment(vertical="center", wrap_text=False)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(3, c, h)
        cell.font = F_HDR
        cell.fill = FILL_HDR
        cell.alignment = HDR_AL
        cell.border = RULE
    ws.row_dimensions[3].height = 30
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = ws.cell(4, freeze_col + 1)
    ws.auto_filter.ref = f"A3:{get_column_letter(ncol)}3"
    ws._expc_rowheight = row_height
    return ws


def put(ws, r, values, fills=None, aligns=None, fonts=None):
    for c, v in enumerate(values, 1):
        cell = ws.cell(r, c, v)
        cell.alignment = TOP
        cell.font = F_BODY
        if fills and fills.get(c):
            cell.fill = fills[c]
        if aligns and aligns.get(c):
            cell.alignment = aligns[c]
        if fonts and fonts.get(c):
            cell.font = fonts[c]
    if getattr(ws, "_expc_rowheight", None):
        ws.row_dimensions[r].height = ws._expc_rowheight


def finish(ws, last_row, ncol):
    ws.auto_filter.ref = f"A3:{get_column_letter(ncol)}{last_row}"
    ws.print_title_rows = "1:3"


# ------------------------------------------------------------------- main ----
def main():
    d = json.loads(CANON.read_text(encoding="utf-8"))
    items, gen = d["items"], d["generated"]
    glossary = d["guard_glossary"]
    tiers = d["tiers"]
    stats = d["tier_stats"]
    BM = d["sentences"]["biomarker"]
    AN = d["sentences"]["anatomy"]
    BM_ROT = d["sentences"]["biomarker_rotation"]
    AN_ROT = d["sentences"]["anatomy_rotation"]

    # source workbook, read-only: order of the 474 and camouflage-root counts
    wb_src = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    rows = list(wb_src["questions"].iter_rows(values_only=True))
    hdr = rows[0]
    recs = [dict(zip(hdr, r)) for r in rows[1:] if r[0] is not None]
    assert len(recs) == 474, len(recs)
    order = [r["question_id"] for r in recs]
    norm_texts = [nf(r["question_text"]) for r in recs]
    wb_src.close()
    root_hits = {}
    for tbl in (BM, AN):
        for vid, (_ent, _sent, root) in tbl.items():
            root_hits[vid] = sum(root in t for t in norm_texts)

    strict_pair = set(tiers["strict"]["PAIR"])
    relaxed_pair = set(tiers["relaxed"]["PAIR"])
    strict_bm, strict_an = set(tiers["strict"]["BM"]), set(tiers["strict"]["AN"])
    relaxed_bm, relaxed_an = set(tiers["relaxed"]["BM"]), set(tiers["relaxed"]["AN"])
    bm_rows = [q for q in order if "BM_text" in gen.get(q, {})]
    an_rows = [q for q in order if "AN_text" in gen.get(q, {})]
    in_bm_arm, in_an_arm = set(bm_rows), set(an_rows)
    eligible = in_bm_arm | in_an_arm

    def yn(b):
        return "yes" if b else "no"

    def neg(q):
        return yn("negated" in (items[q]["flags"] or ""))

    def yr(q):
        v = items[q]["year"]
        return int(v) if str(v).isdigit() else v

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ------------------------------------------------------------ 1 README ---
    ws = wb.create_sheet("README")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 118
    ws.sheet_view.showGridLines = False
    r = 1

    def sect(label):
        nonlocal r
        ws.cell(r, 1, label).font = F_SECT
        ws.cell(r, 1).fill = FILL_TITLE
        ws.cell(r, 2).fill = FILL_TITLE
        ws.cell(r, 1).alignment = Alignment(vertical="center")
        ws.row_dimensions[r].height = 22
        r += 1

    def line(label, body, height=None):
        nonlocal r
        a = ws.cell(r, 1, label)
        a.font = Font(bold=True, size=10, color="FF2E3742")
        a.alignment = Alignment(vertical="top", wrap_text=True)
        b = ws.cell(r, 2, body)
        b.font = F_BODY
        b.alignment = TOP
        if height:
            ws.row_dimensions[r].height = height
        r += 1

    def gap():
        nonlocal r
        r += 1

    ws.cell(r, 1, "Experiment C — fabricated-entity item set").font = Font(bold=True, size=14,
                                                                          color="FF2E3742")
    r += 1
    ws.cell(r, 1, "expC-items.xlsx — working artifact, one row per question, "
                  "every modification shown in full").font = F_SUB
    r += 2

    sect("What this file is")
    line("Purpose",
         "The human-readable companion to canonical.json. It shows every one of the 474 source "
         "questions, the verdict on each, and — for the questions that survived — the exact "
         "control text and the exact modified text with the fabricated entity inserted. Nothing "
         "is truncated: the CTRL / BM / AN columns hold the complete question string.", 62)
    line("Generated", GEN_DATE)
    line("Row heights",
         "Text columns are wrapped and row height is capped so the sheets stay navigable. To read "
         "a long question in full, widen the row (Format > Row > AutoFit Height) or click the cell "
         "and read the formula bar.", 46)
    line("Filters", "Every table sheet has an autofilter on its header row (row 3) and frozen "
                    "panes, so the header and the question_id column stay visible while scrolling.", 32)
    gap()

    sect("What Experiment C is")
    line("The manipulation",
         "Exactly ONE fabricated biomedical entity — a biomarker that does not exist, or an "
         "anatomical structure that does not exist — is inserted as a single additional sentence "
         "at the END of a question's clinical narrative, immediately BEFORE the interrogative. "
         "Nothing else about the item changes: not a character of the original narrative, not the "
         "stem, not the options, not the answer key.", 76)
    line("What it measures",
         "How the model handles a plausible-sounding finding that cannot be real. Three outcomes "
         "are of interest: (1) the model IGNORES it and answers as if it were not there; (2) the "
         "model FLAGS it as unknown/fabricated/not a real entity; (3) the model TREATS IT AS REAL "
         "and reasons from it, e.g. lets it drive the diagnosis or change the answer.", 76)
    line("Two arms",
         "BM (biomarker) — a fake serum/plasma analyte, e.g. 'fibroquelina-X3 sérica se encuentra "
         "aumentada'. AN (anatomy) — a fake structure found on examination, e.g. 'dolor a la "
         "palpación del saco orfalónico'. Items where BOTH arms are admissible form the PAIRED "
         "set, which supports a within-item comparison of the two arms.", 76)
    line("Two tiers",
         "STRICT applies guard J, which removes decision-type stems ('¿cuál es el siguiente "
         "paso?'). Those are exactly the items where a model is most likely to USE a fabricated "
         "finding, so dropping them protects the answer key but costs sensitivity. RELAXED keeps "
         "them. Both tiers are reported rather than one being silently chosen.", 76)
    gap()

    sect("Provenance chain")
    line("1. Source",
         "balanced-flat-A.xlsx (experiment-31-07-26/) — 474 flattened MIR-style clinical "
         "questions, opened READ-ONLY and never written to.", 34)
    line("2. Derivation",
         "tools/build_expC.py — applies the guard funnel, the key-preservation guards, the "
         "per-arm admissibility gates and the deterministic insertion, and writes canonical.json. "
         "That script is the authority for every number in this workbook.", 46)
    line("3. This file",
         "tools/build_items_xlsx.py — reads canonical.json (and re-opens the source workbook "
         "read-only only to count camouflage-root occurrences for the Sentences sheet) and lays "
         "it out for a human reader.", 46)
    line("Chain",
         "balanced-flat-A.xlsx  ->  canonical.json  ->  expC-items.xlsx")
    gap()

    sect("No language model in the generation path")
    line("Deterministic",
         "Every modified string in this workbook was produced by string surgery, not by a language "
         "model. The twenty fabricated sentences are literal constants in build_expC.py. Insertion "
         "is new = pre + SENTENCE + separator + tail, where the seam is the last sentence boundary "
         "before the interrogative, and byte-level assertions prove that no other character moved: "
         "the prefix is identical, the suffix is identical, the new length is exactly the old "
         "length plus the sentence plus the separator, and the sentence occurs exactly once.", 92)
    line("Variant choice",
         "Also deterministic: the variant is chosen by rotating through the arm's rotation list by "
         "the item's rank within its cluster, then skipping any variant whose camouflage root "
         "already appears in the item's own text (so the fake term can never be assimilated as an "
         "echo of the vignette's own vocabulary).", 62)
    line("Reproducibility",
         "This workbook contains values only — no formulas — so nothing can recalculate to a "
         "different number on another machine.", 32)
    gap()

    sect("Headline numbers")
    for lab, val in (
        ("Source items", "474 (balanced-flat-A.xlsx)"),
        ("Mechanically insertable", "182 — a clean seam exists and a patient anchor exists"),
        ("Key-preserving (strict)", "92 base  ->  BM 59, AN 54, paired 37"),
        ("Key-preserving (relaxed)", "148 base  ->  BM 99, AN 86, paired 58"),
        ("Generated rows", "127 questions carry at least one variant (99 biomarker texts, "
                           "86 anatomy texts)"),
        ("Variant use", "BM07 44, BM02 21, BM08 19, BM04 15  |  AN04 56, AN10 30"),
        ("Cluster dependence", "the 37 strict pairs collapse to 13 independent clusters; "
                               "the 58 relaxed pairs to 25"),
    ):
        line(lab, val)
    line("Ceiling",
         "200 items per arm is NOT achievable from this 474-item bank. The ceiling is 99 "
         "(relaxed biomarker arm). Guard A alone rejects 232 of the 474: those items are bare "
         "theory one-liners with no patient in them at all, so there is nothing for a clinical "
         "finding to attach to.", 62)
    gap()

    sect("Sheet guide")
    for lab, val in (
        ("Funnel", "where all 474 went, guard by guard, for both tiers and both arms"),
        ("Eligible - paired (strict)", "the 37 items admissible in BOTH arms under the strict tier, "
                                       "with full CTRL / BM / AN text"),
        ("Eligible - paired (relaxed)", "the 58 items admissible in BOTH arms under the relaxed tier"),
        ("Biomarker arm (99)", "every question that carries a fabricated biomarker sentence"),
        ("Anatomy arm (86)", "every question that carries a fabricated anatomical structure"),
        ("All 474 - verdict", "one row per source question: eligible or not, and the exact guard "
                              "that rejected it"),
        ("Sentences", "the twenty fabricated sentences, their camouflage roots, and how often each "
                      "was used"),
        ("Clusters", "the relaxed paired set grouped by cluster, so the non-independence is visible "
                     "rather than buried"),
    ):
        line(lab, val)
    gap()

    sect("Colour key")
    ws.cell(r, 1, "inserted variant").fill = FILL_MOD
    ws.cell(r, 1).font = F_BODY
    ws.cell(r, 2, "muted sage — this cell holds text that CONTAINS a fabricated entity, or a row "
                  "that carries one").font = F_BODY
    r += 1
    ws.cell(r, 1, "control").fill = FILL_CTRL
    ws.cell(r, 1).font = F_BODY
    ws.cell(r, 2, "muted stone — unmodified control text, exactly as it appears in the source "
                  "workbook").font = F_BODY
    r += 1
    ws.cell(r, 1, "dropped").fill = FILL_OUT
    ws.cell(r, 1).font = F_BODY
    ws.cell(r, 2, "muted clay — this row was rejected by a guard or a gate and carries no "
                  "variant").font = F_BODY
    r += 1
    readme_rows = r - 1

    # ------------------------------------------------------------ 2 Funnel ---
    fh = ["stage", "guard", "what it rejects", "rejects", "remaining"]
    fw = [30, 8, 92, 10, 12]
    ws = table_sheet(wb, "Funnel", "Funnel — where all 474 source items went",
                     "Guards A–I are mechanical (is insertion possible at all?). J–N protect the "
                     "answer key. The per-arm gates ask whether that KIND of finding is admissible.",
                     fh, fw, freeze_col=1, row_height=30)
    r = 4
    num_al = {4: TOPC, 5: TOPC, 2: TOPC}

    def frow(stage, guard, why, rej, rem, fill=None):
        nonlocal r
        fills = {i: fill for i in range(1, 6)} if fill else None
        put(ws, r, [stage, guard, why, rej, rem], fills=fills, aligns=num_al)
        r += 1

    def fstage(text):
        nonlocal r
        for c in range(1, 6):
            ws.cell(r, c).fill = FILL_STAGE
        cell = ws.cell(r, 1, text)
        cell.font = Font(bold=True, size=10, color="FF2E3742")
        cell.alignment = Alignment(vertical="center", wrap_text=False)
        ws.row_dimensions[r].height = 20
        r += 1

    fstage("START")
    frow("source workbook", "", "474 rows in balanced-flat-A.xlsx (sheet 'questions')", None, 474)

    fstage("STAGE 1 — mechanical insertability (guards A–I, applied in this order)")
    for f in d["mechanical_funnel"]:
        frow("mechanical", f["guard"], f["why"], f["rejects"], f["remaining"])
    frow("mechanical pool", "", "a clean seam exists before the interrogative AND a patient anchor "
                                "exists, so ONE sentence can be appended as a pure append",
         None, d["mechanical_pool"])

    mech = [q for q in order if items[q]["mechanical_guard"] is None]
    for tier, guards, label in (
        ("strict", ["J", "K", "L", "M", "N"], "STAGE 2a — key preservation, STRICT tier "
                                              "(guard J applied): 182 -> 92"),
        ("relaxed", ["K", "L", "M", "N"], "STAGE 2b — key preservation, RELAXED tier "
                                          "(guard J dropped): 182 -> 148"),
    ):
        fstage(label)
        cnt = collections.Counter(items[q][f"{tier}_guard"] for q in mech)
        rem = d["mechanical_pool"]
        if tier == "relaxed":
            frow(f"{tier} key-preservation", "J",
                 "NOT APPLIED in this tier — decision-type stems are deliberately kept, because "
                 "they are the items where a model is most likely to USE the fabricated finding",
                 0, rem)
        for g in guards:
            rem -= cnt.get(g, 0)
            frow(f"{tier} key-preservation", g, glossary[g], cnt.get(g, 0), rem)
        frow(f"{tier} base pool", "", "insertion is possible AND the documented correct answer "
                                      "cannot be disturbed by it", None, rem)
        assert rem == len(tiers[tier]["base"])

    fstage("STAGE 3 — per-arm admissibility gates (each branches from its tier's base pool)")
    for tier in ("strict", "relaxed"):
        base = len(tiers[tier]["base"])
        nbm, nan = len(tiers[tier]["BM"]), len(tiers[tier]["AN"])
        npair = len(tiers[tier]["PAIR"])
        frow(f"{tier} — BM gate", "",
             "REJECTED: the vignette already states the labs are normal, or an option is itself "
             "about lab/biomarker results — a new abnormal analyte would contradict the text or "
             "move an option's referent",
             base - nbm, nbm)
        frow(f"{tier} — AN gate", "",
             "REJECTED: the vignette already states the examination is normal / painless, or an "
             "option is itself about the physical examination",
             base - nan, nan)
        frow(f"{tier} — paired", "",
             "items that survive BOTH gates, so the same question can be run in both arms",
             None, npair)

    fstage("STAGE 4 — generation (deterministic variant choice)")
    frow("BM variant collision", "",
         "camouflage refusal: no biomarker variant left whose root is absent from the item's own "
         "text (0 items lost)", len(relaxed_bm) - len(bm_rows), len(bm_rows))
    frow("AN variant collision", "",
         "camouflage refusal: no anatomy variant left whose root is absent from the item's own "
         "text (0 items lost)", len(relaxed_an) - len(an_rows), len(an_rows))
    frow("questions with >=1 variant", "",
         "distinct questions that carry at least one fabricated sentence (BM arm UNION AN arm, "
         "relaxed tier)", None, len(eligible))

    fstage("FINAL — the four arm counts")
    for lab, key, note in (
        ("STRICT  biomarker arm", "strict_BM", "guard J applied"),
        ("STRICT  anatomy arm", "strict_AN", "guard J applied"),
        ("RELAXED biomarker arm", "relaxed_BM",
         "THE CEILING FOR THIS BANK — 200 items per arm is not reachable from 474 source items"),
        ("RELAXED anatomy arm", "relaxed_AN", "guard J dropped"),
    ):
        s = stats[key]
        frow(lab, "", f"{note}; {s['n']} items spread over {s['clusters']} clusters, "
                      f"so these are not {s['n']} independent observations", None, s["n"])
    for lab, key, note in (
        ("STRICT  paired", "strict_PAIR", "both arms run on the same item"),
        ("RELAXED paired", "relaxed_PAIR", "both arms run on the same item"),
    ):
        s = stats[key]
        frow(lab, "", f"{note}; {s['n']} items collapse to {s['clusters']} independent clusters "
                      f"(see the Clusters sheet for the relaxed set)", None, s["n"])

    fstage("CLUSTER-ADJUSTED EFFECTIVE N (Kish, from canonical.json tier_stats)")
    for key in ("strict_BM", "strict_AN", "strict_PAIR", "relaxed_BM", "relaxed_AN", "relaxed_PAIR"):
        s = stats[key]
        frow(key.replace("_", " "), "",
             f"{s['clusters']} clusters, mean cluster size (Kish) {s['kish_mean_cluster']}; "
             f"effective n = {s['n_eff_icc_0.1']} at ICC 0.1, {s['n_eff_icc_0.3']} at ICC 0.3, "
             f"{s['n_eff_icc_0.5']} at ICC 0.5",
             None, s["n"])
    funnel_rows = r - 4
    finish(ws, r - 1, 5)

    # ------------------------------------------- 3 & 4 paired eligible -------
    pair_h = ["question_id", "region", "year", "exam_part", "cluster", "correct_letter",
              "negated?", "narrative_chars", "CTRL question_text", "BM_variant", "BM_entity",
              "BM_text", "AN_variant", "AN_entity", "AN_text"]
    pair_w = [13, 20, 7, 12, 34, 8, 10, 11, 78, 10, 18, 78, 10, 20, 78]
    pair_counts = {}
    for sheet, ids, tier_lab in (
        ("Eligible - paired (strict)", tiers["strict"]["PAIR"], "STRICT"),
        ("Eligible - paired (relaxed)", tiers["relaxed"]["PAIR"], "RELAXED"),
    ):
        ordered = [q for q in order if q in set(ids)]
        ws = table_sheet(
            wb, sheet,
            f"Paired set — {tier_lab} tier ({len(ordered)} items admissible in BOTH arms)",
            "One row per question. CTRL is the untouched source text; BM_text and AN_text are the "
            "same string with one fabricated sentence appended before the interrogative. "
            "Text is complete, not truncated.",
            pair_h, pair_w, freeze_col=1, row_height=108)
        r = 4
        for q in ordered:
            it, g = items[q], gen[q]
            fills = {9: FILL_CTRL}
            for c in (10, 11, 12, 13, 14, 15):
                fills[c] = FILL_MOD
            put(ws, r, [q, it["region"], yr(q), it["exam_part"], it["cluster"],
                        it["correct_letter"], neg(q), it["narrative_chars"], g["CTRL"],
                        g.get("BM_variant"), g.get("BM_entity"), g.get("BM_text"),
                        g.get("AN_variant"), g.get("AN_entity"), g.get("AN_text")],
                fills=fills,
                aligns={3: TOPC, 6: TOPC, 7: TOPC, 8: TOPC, 10: TOPC, 13: TOPC})
            r += 1
        finish(ws, r - 1, len(pair_h))
        pair_counts[sheet] = r - 4

    # ------------------------------------------------- 5 & 6 arm sheets ------
    arm_h = ["question_id", "region", "year", "cluster", "tier", "variant", "entity",
             "inserted sentence", "CTRL text", "%s text"]
    arm_w = [13, 20, 7, 34, 10, 9, 22, 60, 82, 82]
    arm_counts = {}
    for sheet, ids, arm, strict_set, title, sub in (
        ("Biomarker arm (99)", bm_rows, "BM", strict_bm,
         "Biomarker arm — 99 questions carrying a fabricated serum/plasma analyte",
         "Relaxed tier. The 59 rows marked tier=strict are also in the strict tier; "
         "tier=relaxed rows are the ones guard J would have removed."),
        ("Anatomy arm (86)", an_rows, "AN", strict_an,
         "Anatomy arm — 86 questions carrying a fabricated anatomical structure",
         "Relaxed tier. The 54 rows marked tier=strict are also in the strict tier; "
         "tier=relaxed rows are the ones guard J would have removed."),
    ):
        headers = [h % arm if "%s" in h else h for h in arm_h]
        ws = table_sheet(wb, sheet, title, sub, headers, arm_w, freeze_col=1, row_height=108)
        r = 4
        for q in ids:
            it, g = items[q], gen[q]
            fills = {9: FILL_CTRL, 6: FILL_MOD, 7: FILL_MOD, 8: FILL_MOD, 10: FILL_MOD}
            put(ws, r, [q, it["region"], yr(q), it["cluster"],
                        "strict" if q in strict_set else "relaxed",
                        g[f"{arm}_variant"], g[f"{arm}_entity"], g[f"{arm}_sentence"],
                        g["CTRL"], g[f"{arm}_text"]],
                fills=fills, aligns={3: TOPC, 5: TOPC, 6: TOPC})
            r += 1
        finish(ws, r - 1, len(headers))
        arm_counts[sheet] = r - 4

    # --------------------------------------------------- 7 All 474 verdict ---
    v_h = ["question_id", "region", "year", "exam_part", "correct_letter", "n_chars",
           "eligible?", "rejecting_guard", "what that guard means", "in_strict_pair?",
           "in_relaxed_pair?", "in_BM_arm?", "in_AN_arm?"]
    v_w = [13, 20, 7, 12, 8, 9, 10, 14, 92, 13, 14, 11, 11]
    ws = table_sheet(
        wb, "All 474 - verdict",
        "All 474 source items — verdict and rejecting guard",
        "eligible? = yes when the item carries at least one fabricated sentence in the relaxed "
        "tier (either arm). rejecting_guard is the FIRST test the item failed: a mechanical guard "
        "A–I, then a key-preservation guard K–N, then the per-arm gates.",
        v_h, v_w, freeze_col=1, row_height=30)
    r = 4
    gate_note = ("passed every guard, but failed BOTH per-arm gates: the vignette already "
                 "declares the labs normal (or an option is about labs) AND already declares "
                 "the examination normal (or an option is about the examination)")
    for q in order:
        it = items[q]
        elig = q in eligible
        mg = it["mechanical_guard"]
        rg = it.get("relaxed_guard")
        if mg:
            guard, why = mg, glossary[mg]
        elif rg:
            guard, why = rg, glossary[rg]
        elif not elig:
            guard, why = "gate:BM+AN", gate_note
        else:
            guard, why = "(none)", ("survived every guard and at least one per-arm gate — "
                                    "this item carries a fabricated sentence")
        fills = None if elig else {c: FILL_OUT for c in range(1, len(v_h) + 1)}
        if elig:
            fills = {7: FILL_MOD}
        put(ws, r, [q, it["region"], yr(q), it["exam_part"], it["correct_letter"], it["n_chars"],
                    yn(elig), guard, why, yn(q in strict_pair), yn(q in relaxed_pair),
                    yn(q in in_bm_arm), yn(q in in_an_arm)],
            fills=fills,
            aligns={3: TOPC, 5: TOPC, 6: TOPC, 7: TOPC, 8: TOPC,
                    10: TOPC, 11: TOPC, 12: TOPC, 13: TOPC})
        r += 1
    verdict_rows = r - 4
    finish(ws, r - 1, len(v_h))

    # --------------------------------------------------------- 8 Sentences ---
    s_h = ["id", "arm", "entity", "fabricated sentence", "camouflage root",
           "rows of 474 containing that root", "in rotation?", "used?", "times used"]
    s_w = [8, 12, 24, 86, 16, 14, 12, 9, 11]
    ws = table_sheet(
        wb, "Sentences",
        "The twenty fabricated sentences",
        "Literal constants in build_expC.py — never generated by a model. The camouflage root is "
        "the morpheme used to refuse a pairing: if the item's own text already contains it, the "
        "fake term could be read as an echo of the vignette's vocabulary, so that variant is "
        "skipped. Root counts are computed over the normalised text of all 474 source rows.",
        s_h, s_w, freeze_col=1, row_height=42)
    bm_use = collections.Counter(gen[q]["BM_variant"] for q in bm_rows)
    an_use = collections.Counter(gen[q]["AN_variant"] for q in an_rows)
    r = 4
    sent_rows = 0
    for tbl, arm, rot, use in ((BM, "biomarker", BM_ROT, bm_use), (AN, "anatomy", AN_ROT, an_use)):
        for vid in sorted(tbl):
            ent, sent, root = tbl[vid]
            n = use.get(vid, 0)
            fills = {c: FILL_MOD for c in range(1, len(s_h) + 1)} if n else \
                {c: FILL_OUT for c in range(1, len(s_h) + 1)}
            put(ws, r, [vid, arm, ent, sent, root, root_hits[vid],
                        yn(vid in rot), yn(bool(n)), n],
                fills=fills, aligns={2: TOPC, 5: TOPL, 6: TOPC, 7: TOPC, 8: TOPC, 9: TOPC})
            r += 1
            sent_rows += 1
    finish(ws, r - 1, len(s_h))

    # ---------------------------------------------------------- 9 Clusters ---
    c_h = ["cluster", "n items in relaxed paired set", "region(s)", "year(s)", "exam_part(s)",
           "question_ids"]
    c_w = [40, 14, 22, 12, 18, 74]
    ws = table_sheet(
        wb, "Clusters",
        "Clusters in the relaxed paired set — 58 items, 25 clusters",
        "Items sharing a cluster share a clinical vignette (context_ids), so they are NOT "
        "independent observations. Kish mean cluster size 9.93; effective n 30.6 at ICC 0.1, "
        "15.8 at ICC 0.3, 10.6 at ICC 0.5. Two clusters carry 20 and 12 of the 58 items.",
        c_h, c_w, freeze_col=1, row_height=46)
    groups = collections.OrderedDict()
    for q in order:
        if q in relaxed_pair:
            groups.setdefault(items[q]["cluster"], []).append(q)
    r = 4
    for cl, qs in sorted(groups.items(), key=lambda kv: (-len(kv[1]), kv[0])):
        regions = sorted({items[q]["region"] for q in qs})
        years = sorted({str(items[q]["year"]) for q in qs})
        parts = sorted({str(items[q]["exam_part"]) for q in qs})
        fills = {2: FILL_MOD} if len(qs) > 1 else None
        put(ws, r, [cl, len(qs), "; ".join(regions), "; ".join(years), "; ".join(parts),
                    ", ".join(qs)],
            fills=fills, aligns={2: TOPC, 4: TOPC})
        r += 1
    cluster_rows = r - 4
    finish(ws, r - 1, len(c_h))

    for s in wb.worksheets:
        s.sheet_view.zoomScale = 100
    wb.active = 0
    wb.save(OUT)

    counts = {
        "README": readme_rows,
        "Funnel": funnel_rows,
        "Eligible - paired (strict)": pair_counts["Eligible - paired (strict)"],
        "Eligible - paired (relaxed)": pair_counts["Eligible - paired (relaxed)"],
        "Biomarker arm (99)": arm_counts["Biomarker arm (99)"],
        "Anatomy arm (86)": arm_counts["Anatomy arm (86)"],
        "All 474 - verdict": verdict_rows,
        "Sentences": sent_rows,
        "Clusters": cluster_rows,
    }
    print(json.dumps(counts, indent=1))
    print("wrote", OUT)


if __name__ == "__main__":
    main()
