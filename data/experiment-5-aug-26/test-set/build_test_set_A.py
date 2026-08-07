"""Build the Condition-A counterpart of hard10-flat-B.xlsx (for the over-abstention regression).

A and B are the SAME question; they differ only in the correct slot: B holds the sentinel
"Ninguna...", A holds the real correct option. So we take each verified B row and swap the correct
slot back to the real answer (correct_option_text_A), taken from the authoritative adjusted catalog.
The three distractors are asserted identical between our B row and the catalog's A form (paired item).
"""
from __future__ import annotations
import csv, json
from pathlib import Path
from openpyxl import Workbook, load_workbook

HERE = Path(__file__).resolve().parent
CAT = (HERE.parent.parent / "experiment-4-aug-26" / "replacements"
       / "ab520-replacement-22-2026-08-04" / "exports" / "benchmark-500-question-catalog-adjusted.csv")
NOTA = "Ninguna de las respuestas anteriores es correcta."
COLS = ["question_id","region","year","specialty","exam_part","question_number",
    "question_text","option_a","option_b","option_c","option_d",
    "correct_letter","correct_option_text","flags","page_in_exam_pdf",
    "source_exam_pdf","source_answer_key_pdf","content_sha256","source_key",
    "selection_score","context_ids","origin","negated_stem",
    "deterministic_rank","difficulty_tier_wrong_models","wrong_models_B"]

def read_flat_b():
    wb = load_workbook(HERE / "hard10-flat-B.xlsx", read_only=True); ws = wb["questions"]
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = [dict(zip(hdr, r)) for r in ws.iter_rows(min_row=2, values_only=True)]
    wb.close(); return rows

def read_catalog():
    idx = {}
    for r in csv.DictReader(CAT.open(encoding="utf-8")):
        idx[r["question_id"]] = r
    return idx

def main() -> None:
    b_rows = read_flat_b()
    cat = read_catalog()
    a_rows = []
    for b in b_rows:
        qid = b["question_id"]; L = str(b["correct_letter"]).strip().lower()
        c = cat[qid]
        assert str(c["correct_letter"]).strip().lower() == L, f"{qid}: correct_letter mismatch"
        assert c[f"option_{L}_B"] == NOTA, f"{qid}: catalog B correct slot != NOTA"
        real = c["correct_option_text_A"]
        assert real and real != NOTA, f"{qid}: catalog A correct text missing/NOTA"
        # distractors (x != L) must match between our B row and the catalog A form (paired item)
        for x in "abcd":
            if x != L:
                assert str(b[f"option_{x}"]) == c[f"option_{x}_A"], \
                    f"{qid}: distractor {x} mismatch B-row vs catalog-A"
        a = dict(b)
        a[f"option_{L}"] = real
        a["correct_option_text"] = real
        a_rows.append({k: a.get(k, "") for k in COLS})

    assert len(a_rows) == 10
    wb = Workbook(); ws = wb.active; ws.title = "questions"; ws.append(COLS)
    for r in a_rows: ws.append([r[c] for c in COLS])
    wb.save(HERE / "hard10-flat-A.xlsx")
    with (HERE / "hard10-flat-A.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=COLS); w.writeheader(); w.writerows(a_rows)
    print("WROTE hard10-flat-A.xlsx/.csv (Condition-A counterpart of the 10 B questions)")
    for r in a_rows:
        print(f"  {r['question_id']}: key={r['correct_letter']} A-answer={r['correct_option_text'][:52]!r}")

if __name__ == "__main__":
    main()
