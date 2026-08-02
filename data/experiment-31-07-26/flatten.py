"""Flatten balanced-clinical-questionnaire-500-no-image.xlsx to the medrag_eval import contract.

Emits two workbooks:

  balanced-flat-A.xlsx   474 items, source option text untouched
  balanced-flat-B.xlsx   423 items, correct option's TEXT replaced with
                         "Ninguna de las respuestas anteriores es correcta."
                         (correct_letter unchanged)

The source workbook is opened read-only and never modified.

Run from the repo root:
    python data/experiment-31-07-26/flatten.py
"""

from __future__ import annotations

import json
import re
import unicodedata
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook

HERE = Path(__file__).resolve().parent
SOURCE = HERE / "balanced-clinical-questionnaire-500-no-image.xlsx"
OUT_A = HERE / "balanced-flat-A.xlsx"
OUT_B = HERE / "balanced-flat-B.xlsx"
REPORT = HERE / "flatten_report.json"

SWAP_TEXT = "Ninguna de las respuestas anteriores es correcta."
SPECIALTY = "aparato-digestivo"
EXPECTED_A = 474
EXPECTED_B = 423

# excel_io.REQUIRED_COLUMNS, in order. Extra columns after these are ignored by the
# importer (only *missing* ones raise), so audit fields are appended at the end.
REQUIRED_COLUMNS = [
    "question_id", "region", "year", "specialty", "exam_part", "question_number",
    "question_text", "option_a", "option_b", "option_c", "option_d",
    "correct_letter", "correct_option_text", "flags",
    "page_in_exam_pdf", "source_exam_pdf", "source_answer_key_pdf",
]
AUDIT_COLUMNS = ["content_sha256", "source_key", "selection_score", "context_ids"]

# ---------------------------------------------------------------------------------
# QA exclusions — 2026-07-31
#
# Found by two 12-shard adversarial QA fleets (163 agents, every item read by a
# reviewer, every finding then attacked by an independent refuter). Only findings
# that survived refutation appear here: the A fleet upheld 8 of 54 flags, the B fleet
# 18 of 85. A final deliberately-broad regex sweep for meta-option answers caught two
# more that both fleets and the original narrow pattern missed.
#
# ITEM_DEFECTS are wrong in the SOURCE material, independent of any manipulation, so
# they are removed from BOTH datasets. B_ONLY_DEFECTS are items the Experiment-B swap
# specifically breaks; they stay in A, which is a faithful reproduction of the exam.
# ---------------------------------------------------------------------------------

ITEM_DEFECTS = {
    "b10":  ("dangling_reference_no_context", "'control endoscopico de la lesion descrita' — the 15 mm adenoma is described only in b9's stem, never in the shared context"),
    "b14":  ("dangling_reference_no_context", "'dado el estadio y la ubicacion' — T1N0M0 is in b13, '30 cm del margen anal' is in b9; neither is present here"),
    "b16":  ("dangling_reference_no_context", "'seguimiento endoscopico de este paciente' — the curative pT1 resection justifying 6-month control is in b13-b15"),
    "b18":  ("dangling_reference_no_context", "'En este paciente' — the twelve polyps justifying high-risk referral are in b17; on the visible text, option a) is true"),
    "b19":  ("context_missing_for_answerability", "'el cuadro descrito' is b17's twelve sub-centimetre polyps; the vignette states there is NO family history"),
    "b23":  ("dangling_reference_no_context", "'la entidad que afecta a esta familia' — the Lynch-like diagnosis is in b22; the vignette says 'sin antecedentes'"),
    "b202": ("multiple_correct_answers", "stem asks for the single false statement about functional rectal pain; options c and d are both false"),
    "b228": ("references_another_question", "'Teniendo en cuenta su anterior respuesta' — depends on an answer the model never gave"),
    "b408": ("source_answer_key_error", "key says 'Tos' is NOT a Zenker's symptom; chronic cough is a well-documented Zenker's symptom. Defect in the source exam key."),
}

B_ONLY_DEFECTS = {
    "b152": ("aggregator_correct_answer", "'Todos los anteriores son objetivos esperables...' — removing the aggregator leaves all three survivors individually true. Missed by the narrow pattern (masculine 'todos')."),
    "b188": ("aggregator_correct_answer", "'Todos los anteriores' — ascites, encephalopathy and hypoglycaemia each independently indicate admission"),
    "b191": ("surviving_distractor_is_true", "'Todo paciente con hepatitis aguda por VHC debe recibir tratamiento antiviral' is true under current guidance"),
    "b204": ("aggregator_correct_answer", "'Todas son correctas.' — indocyanine green, aminopyrine breath test etc. are each valid"),
    "b245": ("surviving_distractor_is_true", "with 'Mama' removed, 'Melanoma' is defensibly correct; the literature is split"),
    "b257": ("swap_is_noop", "correct answer was already 'Ninguno de las anteriores' — the swap changes nothing and tests nothing"),
    "b292": ("semantic_duplicate", "option d was already 'Ninguna es correcta.'; the swap creates two none-of-the-above options"),
    "b303": ("surviving_distractor_is_true", "with primary sclerosing cholangitis removed, 'Cirrosis de cualquier etiologia' is itself correct"),
    "b342": ("aggregator_correct_answer", "'Todas las afirmaciones anteriores son correctas.' — the exam itself certifies every survivor as true"),
    "b398": ("surviving_distractor_is_true", "removed key was 'No recomendado como monoterapia'; a survivor is defensibly correct"),
    "b428": ("surviving_distractor_is_true", "most frequent gastric polypoid lesion — a survivor is defensible once fundic-gland polyp is removed"),
    "b430": ("aggregator_correct_answer", "'Todas las afirmaciones anteriores son correctas' — every survivor is certified true by the key"),
    "b437": ("aggregator_correct_answer", "'Todas son ciertas.' — every surviving distractor is individually true"),
    "b458": ("swap_is_noop", "correct answer was already 'Todas las anteriores son falsas.' — semantically identical to the swap string"),
    "b491": ("surviving_distractor_is_true", "stem's bar is very low ('se han relacionado con'); survivors also qualify"),
    "b497": ("surviving_distractor_is_true", "acute severe UC with fever after infliximab — a surviving option is also correct"),
}


# A correct option that asserts the OTHER options are true. Replacing it with
# "none of the above" leaves the item with no correct answer at all.
AFFIRM = re.compile(
    r"(todas? las? (respuestas |opciones )?anteriores son (correctas?|ciertas?|verdaderas?)"
    r"|todas? (las )?(respuestas|opciones|anteriores)"
    r"|respuestas?\s+[a-d]\)?\s*y\s*[a-d]\)?\s*son\s*(correctas?|ciertas?|verdaderas?)"
    r"|[a-d]\)\s*y\s*[a-d]\)\s*son\s*(correctas?|ciertas?))"
)
# An option that already means "none of the above".
NEGATE = re.compile(
    r"(todas? las? (respuestas |opciones )?anteriores son (falsas?|incorrectas?)"
    r"|ninguna de (las anteriores|las respuestas|ellas|las opciones))"
)


def s(value) -> str:
    return "" if value is None else str(value).strip()


def norm(value) -> str:
    return re.sub(r"\s+", " ", unicodedata.normalize("NFC", s(value))).strip().casefold()


def read_table(wb, sheet: str, header_row: int) -> list[dict]:
    rows = list(wb[sheet].iter_rows(values_only=True))
    header = [s(h) for h in rows[header_row - 1]]
    out = []
    for raw in rows[header_row:]:
        if all(v is None or s(v) == "" for v in raw):
            continue
        out.append({h: v for h, v in zip(header, raw) if h})
    return out


def build_context_index(case_contexts: list[dict]) -> dict[str, str]:
    """Concatenate each context's parts in '<n>/<total>' order into one narrative."""
    parts: dict[str, list[tuple[int, str]]] = defaultdict(list)
    for row in case_contexts:
        cid = s(row.get("Context ID"))
        text = s(row.get("Clinical context"))
        if not cid or not text:
            continue
        match = re.search(r"(\d+)\s*/\s*(\d+)", s(row.get("Kind / part")))
        order = int(match.group(1)) if match else 1
        parts[cid].append((order, text))

    resolved = {}
    for cid, chunks in parts.items():
        chunks.sort(key=lambda pair: pair[0])
        resolved[cid] = "\n".join(text for _, text in chunks)
    return resolved


def main() -> None:
    wb = load_workbook(SOURCE, read_only=True, data_only=True)
    question_data = {int(s(r["Item"])): r for r in read_table(wb, "Question Data", 4)}
    answer_key = {int(s(r["Item"])): r for r in read_table(wb, "Answer Key", 4)}
    ledger = {int(s(r["Item"])): r for r in read_table(wb, "Selection Ledger", 3)}
    contexts = build_context_index(read_table(wb, "Case Contexts", 4))
    wb.close()

    assert len(question_data) == 500, f"expected 500 items, got {len(question_data)}"

    records: list[dict] = []
    dispositions: dict[int, dict] = {}

    for item in sorted(question_data):
        qd, ak, sl = question_data[item], answer_key[item], ledger[item]
        letter = s(ak["Correct letter"]).lower()
        options = {x: s(qd.get(f"Option {x.upper()}")) for x in "abcd"}

        # --- context resolution: prepend the full narrative chain -----------------
        cids = [c.strip() for c in re.split(r"[;,]", s(qd.get("Context IDs"))) if c.strip()]
        chain = [contexts[c] for c in cids if c in contexts]
        missing = [c for c in cids if c not in contexts]
        assert not missing, f"item {item}: unresolved context id(s) {missing}"

        question = s(qd["Question"])
        question_text = "\n\n".join(chain + [question]) if chain else question

        reasons = []
        if not options["d"]:
            reasons.append("three_option_no_option_d")

        correct_norm = norm(options[letter])
        others = [v for k, v in options.items() if k != letter and v]
        if NEGATE.search(correct_norm):
            b_note = "negate_equivalent_noop"
        elif AFFIRM.search(correct_norm):
            reasons.append("affirm_broken_no_correct_answer_after_swap")
            b_note = None
        elif any(NEGATE.search(norm(v)) for v in others):
            reasons.append("distractor_already_none_of_the_above")
            b_note = None
        else:
            b_note = None

        record = {
            "question_id": f"b{item}",
            "region": s(qd["Region"]),
            "year": int(s(qd["Year"])),
            "specialty": SPECIALTY,
            "exam_part": s(qd["Exam part"]),
            "question_number": item,
            "question_text": question_text,
            "option_a": options["a"],
            "option_b": options["b"],
            "option_c": options["c"],
            "option_d": options["d"],
            "correct_letter": letter,
            "correct_option_text": s(ak["Correct option text"]),
            "flags": s(sl.get("Flags")),
            "page_in_exam_pdf": s(qd.get("PDF page")),
            "source_exam_pdf": s(sl.get("Source exam PDF")),
            "source_answer_key_pdf": s(sl.get("Source answer-key PDF")),
            "content_sha256": s(sl.get("Content SHA256")),
            "source_key": s(sl.get("Source key")),
            "selection_score": s(sl.get("Selection score")),
            "context_ids": ",".join(cids),
        }
        records.append(record)
        dispositions[item] = {
            "question_id": record["question_id"],
            "in_a": "three_option_no_option_d" not in reasons,
            "exclusion_reasons": reasons,
            "b_note": b_note,
            "context_resolved": bool(chain),
            "context_chars": sum(len(c) for c in chain),
        }

    # --- record QA verdicts on the dispositions ---------------------------------
    for item, disp in dispositions.items():
        qid = disp["question_id"]
        if qid in ITEM_DEFECTS:
            code, evidence = ITEM_DEFECTS[qid]
            disp["qa_exclusion"] = {"scope": "both", "code": code, "evidence": evidence}
        elif qid in B_ONLY_DEFECTS:
            code, evidence = B_ONLY_DEFECTS[qid]
            disp["qa_exclusion"] = {"scope": "b_only", "code": code, "evidence": evidence}
        else:
            disp["qa_exclusion"] = None

    # --- dataset A: drop 3-option items and source-level defects ----------------
    rows_a = [
        r for r in records
        if dispositions[r["question_number"]]["in_a"] and r["question_id"] not in ITEM_DEFECTS
    ]

    # --- dataset B: A minus items the swap would break --------------------------
    rows_b = []
    for row in rows_a:
        if row["question_id"] in B_ONLY_DEFECTS:
            continue
        reasons = dispositions[row["question_number"]]["exclusion_reasons"]
        if any(r != "three_option_no_option_d" for r in reasons):
            continue
        swapped = dict(row)
        swapped[f"option_{row['correct_letter']}"] = SWAP_TEXT
        swapped["correct_option_text"] = SWAP_TEXT
        rows_b.append(swapped)

    validate(rows_a, "A", expect=EXPECTED_A)
    validate(rows_b, "B", expect=EXPECTED_B)

    write_workbook(OUT_A, rows_a)
    write_workbook(OUT_B, rows_b)

    report = {
        "source_workbook": SOURCE.name,
        "swap_text": SWAP_TEXT,
        "counts": {
            "source": len(records),
            "dataset_a": len(rows_a),
            "dataset_b": len(rows_b),
            "dropped_three_option": sum(
                1 for d in dispositions.values() if "three_option_no_option_d" in d["exclusion_reasons"]
            ),
            "dropped_affirm_broken": sum(
                1 for d in dispositions.values()
                if "affirm_broken_no_correct_answer_after_swap" in d["exclusion_reasons"]
            ),
            "dropped_distractor_none": sum(
                1 for d in dispositions.values()
                if "distractor_already_none_of_the_above" in d["exclusion_reasons"]
            ),
            "b_semantic_noops": sum(1 for d in dispositions.values() if d["b_note"] == "negate_equivalent_noop"),
            "items_with_resolved_context": sum(1 for d in dispositions.values() if d["context_resolved"]),
            "qa_dropped_from_both": len(ITEM_DEFECTS),
            "qa_dropped_from_b_only": len(B_ONLY_DEFECTS),
        },
        "qa_exclusions": {
            "method": (
                "Two 12-shard adversarial QA fleets (163 agents): every item read by a reviewer, "
                "every finding attacked by an independent refuter. A fleet upheld 8/54 flags, "
                "B fleet 18/85. A final broad regex sweep for meta-option answers caught 2 more "
                "(b152, b458) that both fleets and the original narrow pattern missed."
            ),
            "both_datasets": {k: {"code": v[0], "evidence": v[1]} for k, v in sorted(ITEM_DEFECTS.items())},
            "b_only": {k: {"code": v[0], "evidence": v[1]} for k, v in sorted(B_ONLY_DEFECTS.items())},
        },
        "items": {str(k): v for k, v in sorted(dispositions.items())},
    }
    REPORT.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    c = report["counts"]
    print(f"source={c['source']}  A={c['dataset_a']}  B={c['dataset_b']}")
    print(f"  dropped: 3-option={c['dropped_three_option']}  "
          f"affirm_broken={c['dropped_affirm_broken']}  "
          f"distractor_none={c['dropped_distractor_none']}")
    print(f"  contexts resolved into question_text: {c['items_with_resolved_context']} items")
    print(f"  B semantic no-ops retained: {c['b_semantic_noops']}")
    print(f"wrote {OUT_A.name}, {OUT_B.name}, {REPORT.name}")


def validate(rows: list[dict], label: str, *, expect: int) -> None:
    assert len(rows) == expect, f"{label}: expected {expect} rows, got {len(rows)}"

    ids = [r["question_id"] for r in rows]
    assert len(ids) == len(set(ids)), f"{label}: duplicate question_id"

    for row in rows:
        qid = row["question_id"]
        for column in REQUIRED_COLUMNS:
            assert column in row, f"{label} {qid}: missing column {column}"

        letter = row["correct_letter"]
        assert letter in "abcd", f"{label} {qid}: bad correct_letter {letter!r}"

        # excel_io.py:200 enforces this at import; fail here instead, with context.
        assert row["correct_option_text"] == row[f"option_{letter}"], (
            f"{label} {qid}: correct_option_text != option_{letter}"
        )

        options = [row[f"option_{x}"] for x in "abcd"]
        assert all(o for o in options), f"{label} {qid}: empty option"
        # parser.py builds letters_by_option as a dict comprehension, so two identical
        # option strings would silently collapse and the later letter would win.
        assert len(options) == len(set(options)), f"{label} {qid}: duplicate option text"

        assert "CONTINUES THE ACTIVE CONTEXT" not in row["question_text"], (
            f"{label} {qid}: unresolved context pointer survived into question_text"
        )
        assert row["question_text"].strip(), f"{label} {qid}: empty question_text"
        assert isinstance(row["year"], int), f"{label} {qid}: year not int"
        assert isinstance(row["question_number"], int), f"{label} {qid}: question_number not int"

    print(f"  [{label}] {len(rows)} rows validated")


def write_workbook(path: Path, rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "questions"
    columns = REQUIRED_COLUMNS + AUDIT_COLUMNS
    ws.append(columns)
    for row in rows:
        ws.append([row.get(c, "") for c in columns])
    wb.save(path)


if __name__ == "__main__":
    main()
