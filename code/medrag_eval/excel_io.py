from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterator, Mapping
import warnings

from openpyxl import load_workbook


REQUIRED_COLUMNS = [
    "question_id",
    "region",
    "year",
    "specialty",
    "exam_part",
    "question_number",
    "question_text",
    "option_a",
    "option_b",
    "option_c",
    "option_d",
    "correct_letter",
    "correct_option_text",
    "flags",
    "page_in_exam_pdf",
    "source_exam_pdf",
    "source_answer_key_pdf",
]

OPTION_LETTERS = ("a", "b", "c", "d")


class ExcelImportError(ValueError):
    """Raised when a scoreable workbook row is invalid."""


class ExcelImportWarning(UserWarning):
    """Warning emitted when a non-scoreable workbook row is skipped."""


@dataclass(frozen=True)
class QuestionRow(Mapping[str, Any]):
    question_id: str
    region: str | None
    year: int
    specialty: str | None
    exam_part: str | None
    question_number: int
    question_text: str
    options: dict[str, str]
    correct_letter: str
    correct_option_text: str
    source_row_json: dict[str, Any]

    @property
    def option_a(self) -> str:
        return self.options["a"]

    @property
    def option_b(self) -> str:
        return self.options["b"]

    @property
    def option_c(self) -> str:
        return self.options["c"]

    @property
    def option_d(self) -> str:
        return self.options["d"]

    def as_dict(self) -> dict[str, Any]:
        return {
            "question_id": self.question_id,
            "region": self.region,
            "year": self.year,
            "specialty": self.specialty,
            "exam_part": self.exam_part,
            "question_number": self.question_number,
            "question_text": self.question_text,
            "option_a": self.option_a,
            "option_b": self.option_b,
            "option_c": self.option_c,
            "option_d": self.option_d,
            "correct_letter": self.correct_letter,
            "correct_option_text": self.correct_option_text,
            "source_row_json": self.source_row_json,
        }

    def __getitem__(self, key: str) -> Any:
        return self.as_dict()[key]

    def __iter__(self) -> Iterator[str]:
        return iter(self.as_dict())

    def __len__(self) -> int:
        return len(self.as_dict())


@dataclass(frozen=True)
class ImportResult:
    questions: list[QuestionRow]
    warnings: list[str]


def load_questions_from_excel(path: str | Path, limit: int | None = None) -> list[QuestionRow]:
    questions, _warning_messages = _load_questions(path, limit=limit, emit_warnings=True)
    return questions


def import_questions_from_workbook(path: str | Path, limit: int | None = None) -> ImportResult:
    questions, warning_messages = _load_questions(path, limit=limit, emit_warnings=False)
    return ImportResult(questions=questions, warnings=warning_messages)


def _load_questions(
    path: str | Path,
    *,
    limit: int | None,
    emit_warnings: bool,
) -> tuple[list[QuestionRow], list[str]]:
    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        header_values = next(rows)
    except StopIteration as exc:
        raise ExcelImportError("Workbook is empty") from exc

    headers = [_clean_header(value) for value in header_values]
    column_indexes = _required_column_indexes(headers)

    questions: list[QuestionRow] = []
    warning_messages: list[str] = []
    seen_question_ids: set[str] = set()

    data_rows_seen = 0
    for row_number, raw_values in enumerate(rows, start=2):
        data_rows_seen += 1
        if limit is not None and data_rows_seen > limit:
            break

        source_row = {
            column: _clean_cell(raw_values[index] if index < len(raw_values) else None)
            for column, index in column_indexes.items()
        }
        if all(value is None for value in source_row.values()):
            continue

        question_id = _required_text(source_row, "question_id", row_number)
        if question_id in seen_question_ids:
            raise ExcelImportError(f"Row {row_number} {question_id}: Duplicate question_id")

        if _is_non_scoreable_skip(source_row):
            message = f"Skipped non-scoreable row {row_number} ({question_id}): missing correct_letter and annulled/damaged"
            warning_messages.append(message)
            if emit_warnings:
                warnings.warn(message, ExcelImportWarning, stacklevel=3)
            continue

        question = _validated_question(source_row, row_number=row_number, question_id=question_id)
        seen_question_ids.add(question.question_id)
        questions.append(question)

    return questions, warning_messages


def _required_column_indexes(headers: list[str | None]) -> dict[str, int]:
    indexes: dict[str, int] = {}
    for index, header in enumerate(headers):
        if header is None:
            continue
        if header in indexes:
            raise ExcelImportError(f"Duplicate workbook column: {header}")
        indexes[header] = index

    missing = [column for column in REQUIRED_COLUMNS if column not in indexes]
    if missing:
        raise ExcelImportError(f"Workbook is missing required columns: {', '.join(missing)}")
    return {column: indexes[column] for column in REQUIRED_COLUMNS}


def _validated_question(source_row: dict[str, Any], *, row_number: int, question_id: str) -> QuestionRow:
    year = _required_int(source_row, "year", row_number, question_id)
    question_number = _required_int(source_row, "question_number", row_number, question_id)
    question_text = _required_text(source_row, "question_text", row_number, question_id)
    options = {
        letter: _required_text(source_row, f"option_{letter}", row_number, question_id)
        for letter in OPTION_LETTERS
    }

    correct_letter_value = source_row.get("correct_letter")
    if correct_letter_value is None:
        raise ExcelImportError(f"Row {row_number} {question_id}: missing correct_letter")
    correct_letter = str(correct_letter_value).strip().casefold()
    if correct_letter not in OPTION_LETTERS:
        raise ExcelImportError(f"Row {row_number} {question_id}: correct_letter must be one of a, b, c, d")

    correct_option_text = _required_text(source_row, "correct_option_text", row_number, question_id)
    if correct_option_text != options[correct_letter]:
        raise ExcelImportError(
            f"Row {row_number} {question_id}: correct_option_text must exactly equal option_{correct_letter}"
        )

    source_row_json = dict(source_row)
    source_row_json["year"] = year
    source_row_json["question_number"] = question_number
    source_row_json["correct_letter"] = correct_letter
    source_row_json["correct_option_text"] = correct_option_text
    for letter, option_text in options.items():
        source_row_json[f"option_{letter}"] = option_text
    source_row_json["question_id"] = question_id
    source_row_json["question_text"] = question_text

    return QuestionRow(
        question_id=question_id,
        region=_optional_text(source_row.get("region")),
        year=year,
        specialty=_optional_text(source_row.get("specialty")),
        exam_part=_optional_text(source_row.get("exam_part")),
        question_number=question_number,
        question_text=question_text,
        options=options,
        correct_letter=correct_letter,
        correct_option_text=correct_option_text,
        source_row_json=source_row_json,
    )


def _is_non_scoreable_skip(source_row: dict[str, Any]) -> bool:
    if source_row.get("correct_letter") is not None:
        return False
    flags = str(source_row.get("flags") or "").casefold()
    correct_option_text = str(source_row.get("correct_option_text") or "").casefold()
    flag_tokens = {token.strip() for token in flags.replace(";", ",").split(",")}
    flagged_as_non_scoreable = bool(flag_tokens & {"nullified", "annulled", "anulada", "anulado", "damaged"})
    text_marks_annulled = "pregunta anulada" in correct_option_text
    return flagged_as_non_scoreable or text_marks_annulled


def _required_text(source_row: dict[str, Any], column: str, row_number: int, question_id: str | None = None) -> str:
    value = source_row.get(column)
    if value is None:
        raise ExcelImportError(_row_message(row_number, question_id, f"{column} is required"))
    text = str(value).strip()
    if not text:
        raise ExcelImportError(_row_message(row_number, question_id, f"{column} is required"))
    source_row[column] = text
    return text


def _required_int(source_row: dict[str, Any], column: str, row_number: int, question_id: str) -> int:
    value = source_row.get(column)
    parsed = _parse_int(value)
    if parsed is None:
        raise ExcelImportError(_row_message(row_number, question_id, f"{column} must be a parseable int"))
    source_row[column] = parsed
    return parsed


def _parse_int(value: Any) -> int | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        try:
            number = float(text) if "." in text else int(text)
        except ValueError:
            return None
        if isinstance(number, float):
            return int(number) if number.is_integer() else None
        return number
    return None


def _optional_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _clean_header(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _clean_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _row_message(row_number: int, question_id: str | None, detail: str) -> str:
    if question_id:
        return f"Row {row_number} {question_id}: {detail}"
    return f"Row {row_number}: {detail}"
